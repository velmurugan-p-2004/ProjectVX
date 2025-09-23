# excel_reports.py
"""
Comprehensive Excel Reporting System for Staff Management Attendance System

This module provides advanced Excel reporting capabilities with:
- Multiple sheet reports
- Charts and graphs
- Formatted tables
- Summary statistics
- Data visualization
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime, timedelta
import sqlite3
from database import get_db
import io
import base64
from flask import make_response


class ExcelReportGenerator:
    """Generate comprehensive Excel reports for staff attendance system"""
    
    def __init__(self):
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.title_font = Font(bold=True, size=16, color="2F5597")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def create_staff_attendance_report(self, school_id, start_date, end_date):
        """Create comprehensive staff attendance report"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create multiple sheets
        self._create_summary_sheet(wb, school_id, start_date, end_date)
        self._create_detailed_attendance_sheet(wb, school_id, start_date, end_date)
        self._create_staff_profile_sheet(wb, school_id)
        self._create_department_analysis_sheet(wb, school_id, start_date, end_date)
        self._create_charts_sheet(wb, school_id, start_date, end_date)
        
        return self._save_workbook_to_response(wb, f"Staff_Attendance_Report_{start_date}_to_{end_date}.xlsx")
    
    def create_individual_staff_report(self, staff_id, start_date, end_date):
        """Create individual staff attendance report"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        self._create_individual_summary_sheet(wb, staff_id, start_date, end_date)
        self._create_individual_detailed_sheet(wb, staff_id, start_date, end_date)
        self._create_individual_charts_sheet(wb, staff_id, start_date, end_date)
        
        db = get_db()
        staff = db.execute('SELECT full_name FROM staff WHERE id = ?', (staff_id,)).fetchone()
        staff_name = staff['full_name'].replace(' ', '_') if staff else 'Unknown'
        
        return self._save_workbook_to_response(wb, f"Individual_Report_{staff_name}_{start_date}_to_{end_date}.xlsx")
    
    def create_company_report(self, start_date, end_date):
        """Create company-wide report across all schools"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        self._create_company_summary_sheet(wb, start_date, end_date)
        self._create_school_comparison_sheet(wb, start_date, end_date)
        self._create_company_charts_sheet(wb, start_date, end_date)
        
        return self._save_workbook_to_response(wb, f"Company_Report_{start_date}_to_{end_date}.xlsx")
    
    def create_monthly_report(self, school_id, year, month):
        """Create monthly attendance report with individual staff records"""
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create Staff Records sheet FIRST (main data users want to see)
        self._create_monthly_staff_records_sheet(wb, school_id, year, month)
        
        # Create summary sheets
        self._create_monthly_summary_sheet(wb, school_id, year, month)
        self._create_monthly_calendar_sheet(wb, school_id, year, month)
        self._create_monthly_trends_sheet(wb, school_id, year, month)
        
        return self._save_workbook_to_response(wb, f"Monthly_Report_{year}_{month:02d}.xlsx")
    
    def create_overtime_report(self, school_id, year, month):
        """Create comprehensive overtime report with individual staff overtime data"""
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create Overtime Records sheet FIRST (main data users want to see)
        self._create_overtime_records_sheet(wb, school_id, year, month)
        
        # Create summary sheets
        self._create_overtime_summary_sheet(wb, school_id, year, month)
        self._create_overtime_trends_sheet(wb, school_id, year, month)
        
        return self._save_workbook_to_response(wb, f"Overtime_Report_{year}_{month:02d}.xlsx")
    
    def create_staff_profile_report(self, school_id):
        """Create comprehensive staff profile report"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        self._create_staff_profile_sheet(wb, school_id)
        
        return self._save_workbook_to_response(wb, f"Staff_Profile_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    
    def _create_summary_sheet(self, wb, school_id, start_date, end_date):
        """Create summary sheet with key metrics"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Staff Attendance Summary Report"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # Report details
        ws['A3'] = f"Report Period: {start_date} to {end_date}"
        ws['A4'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        db = get_db()
        
        # Get school info
        school = db.execute('SELECT name FROM schools WHERE id = ?', (school_id,)).fetchone()
        ws['A5'] = f"School: {school['name'] if school else 'Unknown'}"
        
        # Summary statistics
        ws['A7'] = "Summary Statistics"
        ws['A7'].font = Font(bold=True, size=14)
        
        # Headers
        headers = ['Metric', 'Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Calculate metrics
        total_staff = db.execute('SELECT COUNT(*) as count FROM staff WHERE school_id = ?', (school_id,)).fetchone()['count']
        
        total_attendance_records = db.execute('''
            SELECT COUNT(*) as count FROM attendance 
            WHERE school_id = ? AND date BETWEEN ? AND ?
        ''', (school_id, start_date, end_date)).fetchone()['count']
        
        present_days = db.execute('''
            SELECT COUNT(*) as count FROM attendance 
            WHERE school_id = ? AND date BETWEEN ? AND ? AND status IN ('present', 'late', 'on_duty')
        ''', (school_id, start_date, end_date)).fetchone()['count']
        
        absent_days = db.execute('''
            SELECT COUNT(*) as count FROM attendance 
            WHERE school_id = ? AND date BETWEEN ? AND ? AND status = 'absent'
        ''', (school_id, start_date, end_date)).fetchone()['count']
        
        late_arrivals = db.execute('''
            SELECT COUNT(*) as count FROM attendance 
            WHERE school_id = ? AND date BETWEEN ? AND ? AND status = 'late'
        ''', (school_id, start_date, end_date)).fetchone()['count']
        
        # Add metrics to sheet
        metrics = [
            ('Total Staff', total_staff),
            ('Total Attendance Records', total_attendance_records),
            ('Present Days', present_days),
            ('Absent Days', absent_days),
            ('Late Arrivals', late_arrivals),
            ('Attendance Rate', f"{(present_days / total_attendance_records * 100):.1f}%" if total_attendance_records > 0 else "0%")
        ]
        
        for row, (metric, value) in enumerate(metrics, 9):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        
        # Add borders
        for row in range(8, 15):
            for col in range(1, 3):
                ws.cell(row=row, column=col).border = self.border
    
    def _create_detailed_attendance_sheet(self, wb, school_id, start_date, end_date):
        """Create detailed attendance sheet"""
        ws = wb.create_sheet("Detailed Attendance")
        
        # Title
        ws['A1'] = "Detailed Attendance Records"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = ['Staff ID', 'Full Name', 'Department', 'Date', 'Time In', 'Time Out', 'Status', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Get data
        db = get_db()
        attendance_data = db.execute('''
            SELECT s.staff_id, s.full_name, s.department, a.date, a.time_in, a.time_out, a.status, a.notes
            FROM attendance a
            JOIN staff s ON a.staff_id = s.id
            WHERE a.school_id = ? AND a.date BETWEEN ? AND ?
            ORDER BY a.date DESC, s.full_name
        ''', (school_id, start_date, end_date)).fetchall()
        
        # Add data
        for row, record in enumerate(attendance_data, 4):
            ws.cell(row=row, column=1, value=record['staff_id'])
            ws.cell(row=row, column=2, value=record['full_name'])
            ws.cell(row=row, column=3, value=record['department'] or 'N/A')
            ws.cell(row=row, column=4, value=record['date'])
            ws.cell(row=row, column=5, value=record['time_in'])
            ws.cell(row=row, column=6, value=record['time_out'])
            ws.cell(row=row, column=7, value=record['status'].title())
            ws.cell(row=row, column=8, value=record['notes'] or '')
        
        # Format columns
        column_widths = [12, 20, 15, 12, 10, 10, 12, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Add borders to all data
        for row in range(3, len(attendance_data) + 4):
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.border
    
    def _save_workbook_to_response(self, wb, filename):
        """Save workbook and return Flask response"""
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

    def _create_staff_profile_sheet(self, wb, school_id):
        """Create staff profile sheet"""
        ws = wb.create_sheet("Staff Profiles")

        # Title
        ws['A1'] = "Staff Profiles"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:G1')

        # Headers
        headers = ['Staff ID', 'Full Name', 'Department', 'Position', 'Email', 'Phone', 'Join Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        # Get data
        db = get_db()
        staff_data = db.execute('''
            SELECT staff_id, full_name, department, destination, email, phone, created_at
            FROM staff
            WHERE school_id = ?
            ORDER BY full_name
        ''', (school_id,)).fetchall()

        # Add data
        for row, staff in enumerate(staff_data, 4):
            ws.cell(row=row, column=1, value=staff['staff_id'])
            ws.cell(row=row, column=2, value=staff['full_name'])
            ws.cell(row=row, column=3, value=staff['department'] or 'N/A')
            ws.cell(row=row, column=4, value=staff['destination'] or 'N/A')
            ws.cell(row=row, column=5, value=staff['email'] or 'N/A')
            ws.cell(row=row, column=6, value=staff['phone'] or 'N/A')
            ws.cell(row=row, column=7, value=staff['created_at'][:10] if staff['created_at'] else 'N/A')

        # Format columns
        column_widths = [12, 20, 15, 15, 25, 15, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Add borders
        for row in range(3, len(staff_data) + 4):
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.border

    def _create_department_analysis_sheet(self, wb, school_id, start_date, end_date):
        """Create department-wise analysis sheet"""
        ws = wb.create_sheet("Department Analysis")

        # Title
        ws['A1'] = "Department-wise Attendance Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')

        # Headers
        headers = ['Department', 'Total Staff', 'Present Days', 'Absent Days', 'Attendance Rate', 'Late Arrivals']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        # Get data
        db = get_db()
        dept_data = db.execute('''
            SELECT
                COALESCE(s.department, 'Unassigned') as department,
                COUNT(DISTINCT s.id) as total_staff,
                COUNT(CASE WHEN a.status IN ('present', 'late', 'on_duty') THEN 1 END) as present_days,
                COUNT(CASE WHEN a.status = 'absent' THEN 1 END) as absent_days,
                COUNT(CASE WHEN a.status = 'late' THEN 1 END) as late_arrivals
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id AND a.date BETWEEN ? AND ?
            WHERE s.school_id = ?
            GROUP BY s.department
            ORDER BY department
        ''', (start_date, end_date, school_id)).fetchall()

        # Add data
        for row, dept in enumerate(dept_data, 4):
            total_records = dept['present_days'] + dept['absent_days']
            attendance_rate = (dept['present_days'] / total_records * 100) if total_records > 0 else 0

            ws.cell(row=row, column=1, value=dept['department'])
            ws.cell(row=row, column=2, value=dept['total_staff'])
            ws.cell(row=row, column=3, value=dept['present_days'])
            ws.cell(row=row, column=4, value=dept['absent_days'])
            ws.cell(row=row, column=5, value=f"{attendance_rate:.1f}%")
            ws.cell(row=row, column=6, value=dept['late_arrivals'])

        # Format columns
        column_widths = [15, 12, 12, 12, 15, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Add borders
        for row in range(3, len(dept_data) + 4):
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.border

    def _create_charts_sheet(self, wb, school_id, start_date, end_date):
        """Create charts and visualizations sheet"""
        ws = wb.create_sheet("Charts & Analytics")

        # Title
        ws['A1'] = "Attendance Analytics & Charts"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:H1')

        # Create attendance status pie chart
        db = get_db()
        status_data = db.execute('''
            SELECT status, COUNT(*) as count
            FROM attendance
            WHERE school_id = ? AND date BETWEEN ? AND ?
            GROUP BY status
        ''', (school_id, start_date, end_date)).fetchall()

        # Add data for pie chart
        ws['A3'] = "Attendance Status Distribution"
        ws['A3'].font = Font(bold=True, size=12)

        ws['A4'] = "Status"
        ws['B4'] = "Count"
        for col in [1, 2]:
            cell = ws.cell(row=4, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill

        for row, status in enumerate(status_data, 5):
            ws.cell(row=row, column=1, value=status['status'].title())
            ws.cell(row=row, column=2, value=status['count'])

        # Create pie chart
        pie_chart = PieChart()
        labels = Reference(ws, min_col=1, min_row=5, max_row=4 + len(status_data))
        data = Reference(ws, min_col=2, min_row=4, max_row=4 + len(status_data))
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        pie_chart.title = "Attendance Status Distribution"
        ws.add_chart(pie_chart, "D3")

        # Daily attendance trend
        daily_data = db.execute('''
            SELECT date,
                   COUNT(CASE WHEN status IN ('present', 'late', 'on_duty') THEN 1 END) as present_count,
                   COUNT(CASE WHEN status = 'absent' THEN 1 END) as absent_count
            FROM attendance
            WHERE school_id = ? AND date BETWEEN ? AND ?
            GROUP BY date
            ORDER BY date
        ''', (school_id, start_date, end_date)).fetchall()

        # Add daily trend data
        ws['A15'] = "Daily Attendance Trend"
        ws['A15'].font = Font(bold=True, size=12)

        headers = ['Date', 'Present', 'Absent']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=16, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill

        for row, day in enumerate(daily_data, 17):
            ws.cell(row=row, column=1, value=day['date'])
            ws.cell(row=row, column=2, value=day['present_count'])
            ws.cell(row=row, column=3, value=day['absent_count'])

        # Create line chart for daily trend
        line_chart = LineChart()
        line_chart.title = "Daily Attendance Trend"
        line_chart.y_axis.title = "Number of Staff"
        line_chart.x_axis.title = "Date"

        data = Reference(ws, min_col=2, min_row=16, max_col=3, max_row=16 + len(daily_data))
        categories = Reference(ws, min_col=1, min_row=17, max_row=16 + len(daily_data))
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(categories)
        ws.add_chart(line_chart, "D15")

    def _create_individual_summary_sheet(self, wb, staff_id, start_date, end_date):
        """Create individual staff summary sheet"""
        ws = wb.create_sheet("Personal Summary")

        db = get_db()
        staff = db.execute('SELECT * FROM staff WHERE id = ?', (staff_id,)).fetchone()

        # Title
        ws['A1'] = f"Personal Attendance Report - {staff['full_name']}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')

        # Staff details
        ws['A3'] = "Staff Information"
        ws['A3'].font = Font(bold=True, size=12)

        staff_info = [
            ('Staff ID:', staff['staff_id']),
            ('Full Name:', staff['full_name']),
            ('Department:', staff['department'] or 'N/A'),
            ('Position:', staff['destination'] or 'N/A'),
            ('Email:', staff['email'] or 'N/A'),
            ('Phone:', staff['phone'] or 'N/A')
        ]

        for row, (label, value) in enumerate(staff_info, 4):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)

        # Attendance summary
        ws['A11'] = "Attendance Summary"
        ws['A11'].font = Font(bold=True, size=12)

        # Calculate attendance metrics
        attendance_data = db.execute('''
            SELECT status, COUNT(*) as count
            FROM attendance
            WHERE staff_id = ? AND date BETWEEN ? AND ?
            GROUP BY status
        ''', (staff_id, start_date, end_date)).fetchall()

        total_days = sum(record['count'] for record in attendance_data)
        present_days = sum(record['count'] for record in attendance_data if record['status'] in ['present', 'late', 'on_duty'])

        summary_data = [
            ('Total Working Days:', total_days),
            ('Present Days:', present_days),
            ('Attendance Rate:', f"{(present_days/total_days*100):.1f}%" if total_days > 0 else "0%")
        ]

        for row, (label, value) in enumerate(summary_data, 12):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)

        # Format columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20

    def _create_individual_detailed_sheet(self, wb, staff_id, start_date, end_date):
        """Create individual detailed attendance sheet"""
        ws = wb.create_sheet("Detailed Records")

        # Title
        ws['A1'] = "Detailed Attendance Records"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')

        # Headers
        headers = ['Date', 'Time In', 'Time Out', 'Status', 'Hours Worked', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        # Get data
        db = get_db()
        attendance_records = db.execute('''
            SELECT date, time_in, time_out, status, notes
            FROM attendance
            WHERE staff_id = ? AND date BETWEEN ? AND ?
            ORDER BY date DESC
        ''', (staff_id, start_date, end_date)).fetchall()

        # Add data
        for row, record in enumerate(attendance_records, 4):
            ws.cell(row=row, column=1, value=record['date'])
            ws.cell(row=row, column=2, value=record['time_in'])
            ws.cell(row=row, column=3, value=record['time_out'])
            ws.cell(row=row, column=4, value=record['status'].title())

            # Calculate hours worked
            if record['time_in'] and record['time_out']:
                try:
                    time_in = datetime.strptime(record['time_in'], '%H:%M:%S').time()
                    time_out = datetime.strptime(record['time_out'], '%H:%M:%S').time()
                    time_in_dt = datetime.combine(datetime.today(), time_in)
                    time_out_dt = datetime.combine(datetime.today(), time_out)
                    if time_out_dt < time_in_dt:  # Next day checkout
                        time_out_dt += timedelta(days=1)
                    hours_worked = (time_out_dt - time_in_dt).total_seconds() / 3600
                    ws.cell(row=row, column=5, value=f"{hours_worked:.1f}h")
                except:
                    ws.cell(row=row, column=5, value="N/A")
            else:
                ws.cell(row=row, column=5, value="N/A")

            ws.cell(row=row, column=6, value=record['notes'] or '')

        # Format columns
        column_widths = [12, 10, 10, 12, 12, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Add borders
        for row in range(3, len(attendance_records) + 4):
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.border

    def _create_individual_charts_sheet(self, wb, staff_id, start_date, end_date):
        """Create individual charts sheet"""
        ws = wb.create_sheet("Personal Analytics")

        # Title
        ws['A1'] = "Personal Attendance Analytics"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')

        # Monthly attendance trend
        db = get_db()
        monthly_data = db.execute('''
            SELECT strftime('%Y-%m', date) as month,
                   COUNT(CASE WHEN status IN ('present', 'late', 'on_duty') THEN 1 END) as present_days,
                   COUNT(*) as total_days
            FROM attendance
            WHERE staff_id = ? AND date BETWEEN ? AND ?
            GROUP BY strftime('%Y-%m', date)
            ORDER BY month
        ''', (staff_id, start_date, end_date)).fetchall()

        # Add monthly trend data
        ws['A3'] = "Monthly Attendance Trend"
        ws['A3'].font = Font(bold=True, size=12)

        headers = ['Month', 'Present Days', 'Total Days', 'Attendance %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill

        for row, month in enumerate(monthly_data, 5):
            attendance_pct = (month['present_days'] / month['total_days'] * 100) if month['total_days'] > 0 else 0
            ws.cell(row=row, column=1, value=month['month'])
            ws.cell(row=row, column=2, value=month['present_days'])
            ws.cell(row=row, column=3, value=month['total_days'])
            ws.cell(row=row, column=4, value=f"{attendance_pct:.1f}%")

        # Format columns
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 15

    def _create_company_summary_sheet(self, wb, start_date, end_date):
        """Create company-wide summary sheet"""
        ws = wb.create_sheet("Company Summary")

        # Title
        ws['A1'] = "Company-wide Attendance Summary"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')

        ws['A3'] = f"Report Period: {start_date} to {end_date}"
        ws['A4'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Headers
        headers = ['School Name', 'Total Staff', 'Total Records', 'Present Days', 'Attendance Rate', 'Late Arrivals']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        # Get data
        db = get_db()
        school_data = db.execute('''
            SELECT
                s.name as school_name,
                COUNT(DISTINCT st.id) as total_staff,
                COUNT(a.id) as total_records,
                COUNT(CASE WHEN a.status IN ('present', 'late', 'on_duty') THEN 1 END) as present_days,
                COUNT(CASE WHEN a.status = 'late' THEN 1 END) as late_arrivals
            FROM schools s
            LEFT JOIN staff st ON s.id = st.school_id
            LEFT JOIN attendance a ON st.id = a.staff_id AND a.date BETWEEN ? AND ?
            GROUP BY s.id, s.name
            ORDER BY s.name
        ''', (start_date, end_date)).fetchall()

        # Add data
        for row, school in enumerate(school_data, 7):
            attendance_rate = (school['present_days'] / school['total_records'] * 100) if school['total_records'] > 0 else 0

            ws.cell(row=row, column=1, value=school['school_name'])
            ws.cell(row=row, column=2, value=school['total_staff'])
            ws.cell(row=row, column=3, value=school['total_records'])
            ws.cell(row=row, column=4, value=school['present_days'])
            ws.cell(row=row, column=5, value=f"{attendance_rate:.1f}%")
            ws.cell(row=row, column=6, value=school['late_arrivals'])

        # Format columns
        column_widths = [25, 12, 15, 12, 15, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Add borders
        for row in range(6, len(school_data) + 7):
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.border

    def _create_school_comparison_sheet(self, wb, start_date, end_date):
        """Create school comparison sheet"""
        ws = wb.create_sheet("School Comparison")

        # Title
        ws['A1'] = "School Performance Comparison"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')

        # Get comparative data
        db = get_db()
        comparison_data = db.execute('''
            SELECT
                s.name as school_name,
                COUNT(DISTINCT st.id) as total_staff,
                ROUND(AVG(CASE WHEN a.status IN ('present', 'late', 'on_duty') THEN 1.0 ELSE 0.0 END) * 100, 2) as avg_attendance_rate,
                COUNT(CASE WHEN a.status = 'late' THEN 1 END) as total_late_arrivals,
                COUNT(CASE WHEN a.status = 'absent' THEN 1 END) as total_absences
            FROM schools s
            LEFT JOIN staff st ON s.id = st.school_id
            LEFT JOIN attendance a ON st.id = a.staff_id AND a.date BETWEEN ? AND ?
            GROUP BY s.id, s.name
            ORDER BY avg_attendance_rate DESC
        ''', (start_date, end_date)).fetchall()

        # Headers
        headers = ['Rank', 'School Name', 'Staff Count', 'Attendance Rate', 'Performance']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        # Add ranked data
        for row, (rank, school) in enumerate(enumerate(comparison_data, 1), 4):
            performance = "Excellent" if school['avg_attendance_rate'] >= 95 else \
                         "Good" if school['avg_attendance_rate'] >= 85 else \
                         "Average" if school['avg_attendance_rate'] >= 75 else "Needs Improvement"

            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=school['school_name'])
            ws.cell(row=row, column=3, value=school['total_staff'])
            ws.cell(row=row, column=4, value=f"{school['avg_attendance_rate']:.1f}%")
            ws.cell(row=row, column=5, value=performance)

        # Format columns
        column_widths = [8, 25, 12, 15, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

    def _create_company_charts_sheet(self, wb, start_date, end_date):
        """Create company charts sheet"""
        ws = wb.create_sheet("Company Analytics")

        # Title
        ws['A1'] = "Company-wide Analytics"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')

        # School performance comparison chart data
        db = get_db()
        chart_data = db.execute('''
            SELECT
                s.name as school_name,
                ROUND(AVG(CASE WHEN a.status IN ('present', 'late', 'on_duty') THEN 1.0 ELSE 0.0 END) * 100, 1) as attendance_rate
            FROM schools s
            LEFT JOIN staff st ON s.id = st.school_id
            LEFT JOIN attendance a ON st.id = a.staff_id AND a.date BETWEEN ? AND ?
            GROUP BY s.id, s.name
            ORDER BY attendance_rate DESC
        ''', (start_date, end_date)).fetchall()

        # Add chart data
        ws['A3'] = "School Performance Comparison"
        ws['A3'].font = Font(bold=True, size=12)

        headers = ['School', 'Attendance Rate (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill

        for row, school in enumerate(chart_data, 5):
            ws.cell(row=row, column=1, value=school['school_name'])
            ws.cell(row=row, column=2, value=school['attendance_rate'])

        # Create bar chart
        bar_chart = BarChart()
        bar_chart.title = "School Attendance Rate Comparison"
        bar_chart.y_axis.title = "Attendance Rate (%)"
        bar_chart.x_axis.title = "Schools"

        data = Reference(ws, min_col=2, min_row=4, max_row=4 + len(chart_data))
        categories = Reference(ws, min_col=1, min_row=5, max_row=4 + len(chart_data))
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(categories)
        ws.add_chart(bar_chart, "D3")

    def _create_monthly_staff_records_sheet(self, wb, school_id, year, month):
        """Create individual staff records sheet for monthly attendance"""
        ws = wb.create_sheet("Staff Records")
        
        # Title
        ws['A1'] = f"Monthly Staff Attendance Records - {year}/{month:02d}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:I1')
        
        # Date range info
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        ws['A2'] = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        ws['A2'].font = Font(bold=True, size=11)
        
        db = get_db()
        school = db.execute('SELECT name FROM schools WHERE id = ?', (school_id,)).fetchone()
        ws['A3'] = f"School: {school['name'] if school else 'Unknown'}"
        
        # Headers
        headers = [
            'Staff ID', 'Staff Name', 'Department', 'Position',
            'Total Working Days', 'Absent (count)', 'Leave (count)',
            'On Duty (OD) (count)', 'Total Present Days'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Calculate total working days in the month (business days)
        import calendar
        total_days_in_month = (end_date - start_date).days + 1
        
        # Count business days (excluding weekends) for more accurate working days
        business_days_count = 0
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() < 5:  # Monday = 0, Sunday = 6, so < 5 means weekdays
                business_days_count += 1
            current_date += timedelta(days=1)
        
        # Get individual staff monthly attendance summary with corrected calculations
        staff_records = db.execute('''
            SELECT 
                s.staff_id,
                s.full_name,
                s.department,
                COALESCE(s.destination, '') as position,
                COUNT(CASE WHEN a.status = 'absent' THEN 1 END) as recorded_absent_count,
                COUNT(CASE WHEN a.status = 'leave' THEN 1 END) as leave_count,
                COUNT(CASE WHEN a.status = 'on_duty' THEN 1 END) as on_duty_count,
                COUNT(CASE WHEN a.status IN ('present', 'late', 'early_departure') THEN 1 END) as recorded_present_count,
                COUNT(CASE WHEN a.date IS NOT NULL THEN 1 END) as total_recorded_days
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id 
                AND a.date BETWEEN ? AND ?
                AND a.school_id = ?
            WHERE s.school_id = ? AND s.is_active = 1
            GROUP BY s.id, s.staff_id, s.full_name, s.department, s.destination
            ORDER BY CAST(s.staff_id AS INTEGER) ASC
        ''', (start_date, end_date, school_id, school_id)).fetchall()
        
        # Store corrected values for totals calculation
        corrected_staff_data = []
        
        # Add staff data with corrected calculations
        row = 6
        for staff in staff_records:
            # Calculate corrected values
            working_days = business_days_count
            leave_count = staff['leave_count']
            on_duty_count = staff['on_duty_count']
            recorded_present = staff['recorded_present_count']
            
            # Present count includes actual present + late + on_duty days
            total_present_days = recorded_present + on_duty_count
            
            # Absent count = Working days - (Present + Leave + On Duty)
            # But we need to account for days without attendance records
            recorded_days = staff['total_recorded_days']
            unrecorded_days = working_days - recorded_days
            
            # If there are unrecorded days, we assume they are absent
            actual_absent_count = staff['recorded_absent_count'] + unrecorded_days
            
            # Store corrected data for totals
            corrected_data = {
                'working_days': working_days,
                'absent_count': actual_absent_count,
                'leave_count': leave_count,
                'on_duty_count': on_duty_count,
                'present_count': total_present_days
            }
            corrected_staff_data.append(corrected_data)
            
            values = [
                staff['staff_id'],
                staff['full_name'],
                staff['department'] or 'Unassigned',
                staff['position'],
                working_days,  # Total working days (business days)
                actual_absent_count,  # Corrected absent count
                leave_count,
                on_duty_count,
                total_present_days  # Corrected present count
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                
                # Add conditional formatting for better readability
                if col == 6 and value > 5:  # High absence count
                    cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                elif col == 9 and working_days > 0:  # Attendance rate coloring based on present days
                    rate = value / working_days
                    if rate >= 0.95:  # Excellent attendance
                        cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
                    elif rate < 0.80:  # Poor attendance
                        cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            
            row += 1
        
        # Add summary row
        if corrected_staff_data:
            row += 1
            
            # Calculate totals using corrected data
            total_staff = len(corrected_staff_data)
            total_working_days_display = total_staff * business_days_count
            total_absent = sum(record['absent_count'] for record in corrected_staff_data)
            total_leave = sum(record['leave_count'] for record in corrected_staff_data)
            total_on_duty = sum(record['on_duty_count'] for record in corrected_staff_data)
            total_present = sum(record['present_count'] for record in corrected_staff_data)
            
            # Create summary row values with proper structure
            summary_values = [
                "TOTALS:",
                f"{total_staff} Staff",
                "",
                "",
                total_working_days_display,
                total_absent,
                total_leave,
                total_on_duty,
                total_present
            ]
            
            for col, value in enumerate(summary_values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = Font(bold=True)
                cell.border = self.border
                if col == 1:  # "TOTALS:" label
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                else:  # Data cells
                    cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        # Format columns
        column_widths = [12, 25, 18, 18, 16, 14, 12, 16, 16]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Add notes
        notes_row = row + 3
        ws.cell(row=notes_row, column=1, value="Notes:").font = Font(bold=True)
        ws.cell(row=notes_row + 1, column=1, value=f"• Total Working Days: {business_days_count} business days in {year}/{month:02d}")
        ws.cell(row=notes_row + 2, column=1, value="• Staff ID sorted in ascending numerical order") 
        ws.cell(row=notes_row + 3, column=1, value="• Absent count = Working days - (Present + Leave + On Duty)")
        ws.cell(row=notes_row + 4, column=1, value="• Total Present Days = Present + Late + On Duty")
        ws.cell(row=notes_row + 5, column=1, value="• Days without attendance records are counted as absent")
        ws.cell(row=notes_row + 6, column=1, value="• Green highlighting = Excellent attendance (≥95%)")
        ws.cell(row=notes_row + 7, column=1, value="• Red highlighting = High absence count (>5) or poor attendance (<80%)")

    def _create_monthly_summary_sheet(self, wb, school_id, year, month):
        """Create monthly summary sheet"""
        ws = wb.create_sheet("Monthly Summary")

        # Title
        ws['A1'] = f"Monthly Attendance Report - {year}/{month:02d}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')

        db = get_db()
        school = db.execute('SELECT name FROM schools WHERE id = ?', (school_id,)).fetchone()
        ws['A3'] = f"School: {school['name'] if school else 'Unknown'}"

        # Monthly statistics
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)

        # Get monthly stats
        monthly_stats = db.execute('''
            SELECT
                COUNT(DISTINCT staff_id) as active_staff,
                COUNT(*) as total_records,
                COUNT(CASE WHEN status IN ('present', 'late', 'on_duty') THEN 1 END) as present_days,
                COUNT(CASE WHEN status = 'absent' THEN 1 END) as absent_days,
                COUNT(CASE WHEN status = 'late' THEN 1 END) as late_days,
                COUNT(CASE WHEN status = 'leave' THEN 1 END) as leave_days
            FROM attendance
            WHERE school_id = ? AND date BETWEEN ? AND ?
        ''', (school_id, start_date, end_date)).fetchone()

        # Display stats
        ws['A5'] = "Monthly Statistics"
        ws['A5'].font = Font(bold=True, size=12)

        stats = [
            ('Active Staff:', monthly_stats['active_staff']),
            ('Total Records:', monthly_stats['total_records']),
            ('Present Days:', monthly_stats['present_days']),
            ('Absent Days:', monthly_stats['absent_days']),
            ('Late Arrivals:', monthly_stats['late_days']),
            ('Leave Days:', monthly_stats['leave_days']),
            ('Attendance Rate:', f"{(monthly_stats['present_days']/monthly_stats['total_records']*100):.1f}%" if monthly_stats['total_records'] > 0 else "0%")
        ]

        for row, (label, value) in enumerate(stats, 6):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)

        # Format columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

    def _create_monthly_calendar_sheet(self, wb, school_id, year, month):
        """Create monthly calendar view sheet"""
        ws = wb.create_sheet("Monthly Calendar")

        # Title
        ws['A1'] = f"Monthly Calendar View - {year}/{month:02d}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:H1')

        # Calendar headers
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        for col, day in enumerate(days, 1):
            cell = ws.cell(row=3, column=col, value=day)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        # Get daily attendance summary for the month
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)

        db = get_db()
        daily_summary = db.execute('''
            SELECT date,
                   COUNT(CASE WHEN status IN ('present', 'late', 'on_duty') THEN 1 END) as present_count,
                   COUNT(CASE WHEN status = 'absent' THEN 1 END) as absent_count,
                   COUNT(*) as total_count
            FROM attendance
            WHERE school_id = ? AND date BETWEEN ? AND ?
            GROUP BY date
            ORDER BY date
        ''', (school_id, start_date, end_date)).fetchall()

        # Create calendar grid
        current_date = start_date
        row = 4
        col = current_date.weekday() + 1  # Monday = 0, so add 1 for column

        daily_data = {record['date']: record for record in daily_summary}

        while current_date <= end_date:
            if col > 7:
                col = 1
                row += 3  # Leave space for attendance data

            # Date
            ws.cell(row=row, column=col, value=current_date.day).font = Font(bold=True)

            # Attendance data
            if current_date.strftime('%Y-%m-%d') in daily_data:
                data = daily_data[current_date.strftime('%Y-%m-%d')]
                ws.cell(row=row+1, column=col, value=f"P: {data['present_count']}")
                ws.cell(row=row+2, column=col, value=f"A: {data['absent_count']}")

            current_date += timedelta(days=1)
            col += 1

        # Format columns
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 12

    def _create_monthly_trends_sheet(self, wb, school_id, year, month):
        """Create monthly trends analysis sheet"""
        ws = wb.create_sheet("Monthly Trends")

        # Title
        ws['A1'] = f"Monthly Trends Analysis - {year}/{month:02d}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')

        # Weekly breakdown
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)

        db = get_db()
        weekly_data = db.execute('''
            SELECT
                strftime('%W', date) as week_number,
                COUNT(CASE WHEN status IN ('present', 'late', 'on_duty') THEN 1 END) as present_count,
                COUNT(CASE WHEN status = 'absent' THEN 1 END) as absent_count,
                COUNT(*) as total_count
            FROM attendance
            WHERE school_id = ? AND date BETWEEN ? AND ?
            GROUP BY strftime('%W', date)
            ORDER BY week_number
        ''', (school_id, start_date, end_date)).fetchall()

        # Add weekly data
        ws['A3'] = "Weekly Breakdown"
        ws['A3'].font = Font(bold=True, size=12)

        headers = ['Week', 'Present', 'Absent', 'Total', 'Attendance %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill

        for row, week in enumerate(weekly_data, 5):
            attendance_pct = (week['present_count'] / week['total_count'] * 100) if week['total_count'] > 0 else 0
            ws.cell(row=row, column=1, value=f"Week {week['week_number']}")
            ws.cell(row=row, column=2, value=week['present_count'])
            ws.cell(row=row, column=3, value=week['absent_count'])
            ws.cell(row=row, column=4, value=week['total_count'])
            ws.cell(row=row, column=5, value=f"{attendance_pct:.1f}%")

        # Format columns
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 12

    def _create_overtime_records_sheet(self, wb, school_id, year, month):
        """Create individual staff overtime records sheet"""
        ws = wb.create_sheet("Overtime Records")
        
        # Title
        ws['A1'] = f"Staff Overtime Records - {year}/{month:02d}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:G1')
        
        # Date range info
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        ws['A2'] = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        ws['A2'].font = Font(bold=True, size=11)
        
        db = get_db()
        school = db.execute('SELECT name FROM schools WHERE id = ?', (school_id,)).fetchone()
        ws['A3'] = f"School: {school['name'] if school else 'Unknown'}"
        
        # Headers
        headers = [
            'Staff ID', 'Staff Name', 'Department', 'Position',
            'Total Overtime Days', 'Total Overtime Hours', 'Overtime Details'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Helper function to calculate overtime hours
        def calculate_overtime_hours(overtime_in, overtime_out):
            if not overtime_in or not overtime_out:
                return 0.0
            try:
                # Convert time strings to datetime objects for calculation
                if isinstance(overtime_in, str):
                    in_time = datetime.strptime(overtime_in, '%H:%M:%S').time()
                else:
                    in_time = overtime_in
                
                if isinstance(overtime_out, str):
                    out_time = datetime.strptime(overtime_out, '%H:%M:%S').time()
                else:
                    out_time = overtime_out
                
                # Calculate duration in hours
                in_minutes = in_time.hour * 60 + in_time.minute
                out_minutes = out_time.hour * 60 + out_time.minute
                
                # Handle overnight overtime
                if out_minutes < in_minutes:
                    out_minutes += 24 * 60
                
                duration_minutes = out_minutes - in_minutes
                return round(duration_minutes / 60.0, 2)
            except:
                return 0.0
        
        # Get individual staff overtime data with detailed breakdown
        staff_overtime_data = db.execute('''
            SELECT 
                s.staff_id,
                s.full_name,
                s.department,
                COALESCE(s.destination, '') as position,
                COUNT(CASE WHEN a.overtime_in IS NOT NULL AND a.overtime_out IS NOT NULL THEN 1 END) as overtime_days,
                GROUP_CONCAT(
                    CASE 
                        WHEN a.overtime_in IS NOT NULL AND a.overtime_out IS NOT NULL 
                        THEN a.date || ' (' || a.overtime_in || '-' || a.overtime_out || ')'
                    END, 
                    '; '
                ) as overtime_details,
                a.overtime_in,
                a.overtime_out
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id 
                AND a.date BETWEEN ? AND ?
                AND a.school_id = ?
                AND (a.overtime_in IS NOT NULL OR a.overtime_out IS NOT NULL)
            WHERE s.school_id = ? AND s.is_active = 1
            GROUP BY s.id, s.staff_id, s.full_name, s.department, s.destination
            ORDER BY CAST(s.staff_id AS INTEGER) ASC
        ''', (start_date, end_date, school_id, school_id)).fetchall()
        
        # Calculate total hours for each staff member and collect data for totals
        corrected_staff_data = []
        row = 6
        
        for staff in staff_overtime_data:
            # Get detailed overtime records for this staff member
            overtime_records = db.execute('''
                SELECT overtime_in, overtime_out, date
                FROM attendance a
                JOIN staff s ON s.id = a.staff_id
                WHERE s.staff_id = ? AND s.school_id = ?
                    AND a.date BETWEEN ? AND ?
                    AND a.overtime_in IS NOT NULL 
                    AND a.overtime_out IS NOT NULL
            ''', (staff['staff_id'], school_id, start_date, end_date)).fetchall()
            
            # Calculate total overtime hours
            total_overtime_hours = 0.0
            for record in overtime_records:
                hours = calculate_overtime_hours(record['overtime_in'], record['overtime_out'])
                total_overtime_hours += hours
            
            overtime_days = len(overtime_records)
            
            # Format overtime details for display
            overtime_details = ""
            if overtime_records:
                details_list = []
                for record in overtime_records:
                    details_list.append(f"{record['date']} ({record['overtime_in']}-{record['overtime_out']})")
                overtime_details = "; ".join(details_list[:3])  # Show max 3 details
                if len(overtime_records) > 3:
                    overtime_details += f" and {len(overtime_records)-3} more..."
            else:
                overtime_details = "No overtime recorded"
            
            # Store data for totals calculation
            corrected_data = {
                'overtime_days': overtime_days,
                'total_hours': total_overtime_hours
            }
            corrected_staff_data.append(corrected_data)
            
            values = [
                staff['staff_id'],
                staff['full_name'],
                staff['department'] or 'Unassigned',
                staff['position'],
                overtime_days,
                f"{total_overtime_hours:.2f}",
                overtime_details
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                
                # Add conditional formatting
                if col == 5 and overtime_days > 5:  # High overtime days
                    cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                elif col == 6 and total_overtime_hours > 20:  # High overtime hours
                    cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                elif col == 5 and overtime_days > 0:  # Has overtime
                    cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            
            row += 1
        
        # Add summary row
        if corrected_staff_data:
            row += 1
            
            # Calculate totals
            total_staff_with_overtime = len([d for d in corrected_staff_data if d['overtime_days'] > 0])
            total_overtime_days = sum(record['overtime_days'] for record in corrected_staff_data)
            total_overtime_hours = sum(record['total_hours'] for record in corrected_staff_data)
            
            summary_values = [
                "TOTALS:",
                f"{total_staff_with_overtime} Staff with Overtime",
                "",
                "",
                total_overtime_days,
                f"{total_overtime_hours:.2f}",
                f"{total_staff_with_overtime} out of {len(corrected_staff_data)} staff worked overtime"
            ]
            
            for col, value in enumerate(summary_values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = Font(bold=True)
                cell.border = self.border
                if col == 1:  # "TOTALS:" label
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                else:  # Data cells
                    cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        # Format columns
        column_widths = [12, 25, 18, 18, 16, 16, 50]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Add notes
        notes_row = row + 3
        ws.cell(row=notes_row, column=1, value="Notes:").font = Font(bold=True)
        ws.cell(row=notes_row + 1, column=1, value=f"• Overtime data for {year}/{month:02d}")
        ws.cell(row=notes_row + 2, column=1, value="• Staff ID sorted in ascending numerical order")
        ws.cell(row=notes_row + 3, column=1, value="• Hours calculated from overtime_in and overtime_out times")
        ws.cell(row=notes_row + 4, column=1, value="• Green highlighting = Staff with overtime recorded")
        ws.cell(row=notes_row + 5, column=1, value="• Red highlighting = High overtime (>5 days or >20 hours)")

    def _create_overtime_summary_sheet(self, wb, school_id, year, month):
        """Create overtime summary sheet with statistics"""
        ws = wb.create_sheet("Overtime Summary")
        
        # Title
        ws['A1'] = f"Overtime Summary Report - {year}/{month:02d}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        db = get_db()
        school = db.execute('SELECT name FROM schools WHERE id = ?', (school_id,)).fetchone()
        ws['A3'] = f"School: {school['name'] if school else 'Unknown'}"
        
        # Monthly statistics
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        # Get overtime statistics
        overtime_stats = db.execute('''
            SELECT
                COUNT(DISTINCT s.id) as total_staff,
                COUNT(DISTINCT CASE WHEN a.overtime_in IS NOT NULL AND a.overtime_out IS NOT NULL THEN s.id END) as staff_with_overtime,
                COUNT(CASE WHEN a.overtime_in IS NOT NULL AND a.overtime_out IS NOT NULL THEN 1 END) as total_overtime_days,
                AVG(CASE 
                    WHEN a.overtime_in IS NOT NULL AND a.overtime_out IS NOT NULL 
                    THEN (strftime('%H', a.overtime_out) * 60 + strftime('%M', a.overtime_out)) - 
                         (strftime('%H', a.overtime_in) * 60 + strftime('%M', a.overtime_in))
                END) / 60.0 as avg_overtime_hours_per_day
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id 
                AND a.date BETWEEN ? AND ?
                AND a.school_id = ?
            WHERE s.school_id = ? AND s.is_active = 1
        ''', (start_date, end_date, school_id, school_id)).fetchone()
        
        # Summary statistics
        ws['A5'] = "Summary Statistics"
        ws['A5'].font = Font(bold=True, size=12)
        
        stats = [
            ["Total Active Staff", overtime_stats['total_staff'] or 0],
            ["Staff with Overtime", overtime_stats['staff_with_overtime'] or 0],
            ["Total Overtime Days", overtime_stats['total_overtime_days'] or 0],
            ["Average Hours per Overtime Day", f"{overtime_stats['avg_overtime_hours_per_day']:.2f}" if overtime_stats['avg_overtime_hours_per_day'] else "0.00"],
            ["Overtime Coverage %", f"{((overtime_stats['staff_with_overtime'] or 0) / max(overtime_stats['total_staff'], 1) * 100):.1f}%"]
        ]
        
        for row, (label, value) in enumerate(stats, 7):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            
            # Add borders
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=2).border = self.border
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def _create_overtime_trends_sheet(self, wb, school_id, year, month):
        """Create overtime trends analysis sheet"""
        ws = wb.create_sheet("Overtime Trends")
        
        # Title
        ws['A1'] = f"Overtime Trends - {year}/{month:02d}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # Date range
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        db = get_db()
        
        # Department-wise overtime analysis
        dept_overtime = db.execute('''
            SELECT
                COALESCE(s.department, 'Unassigned') as department,
                COUNT(DISTINCT s.id) as total_staff,
                COUNT(DISTINCT CASE WHEN a.overtime_in IS NOT NULL AND a.overtime_out IS NOT NULL THEN s.id END) as staff_with_overtime,
                COUNT(CASE WHEN a.overtime_in IS NOT NULL AND a.overtime_out IS NOT NULL THEN 1 END) as overtime_instances
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id 
                AND a.date BETWEEN ? AND ?
                AND a.school_id = ?
            WHERE s.school_id = ? AND s.is_active = 1
            GROUP BY s.department
            ORDER BY overtime_instances DESC
        ''', (start_date, end_date, school_id, school_id)).fetchall()
        
        # Department analysis
        ws['A3'] = "Department-wise Overtime Analysis"
        ws['A3'].font = Font(bold=True, size=12)
        
        headers = ['Department', 'Total Staff', 'Staff with OT', 'OT Instances', 'Coverage %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        for row, dept in enumerate(dept_overtime, 5):
            coverage_pct = (dept['staff_with_overtime'] / max(dept['total_staff'], 1) * 100)
            
            values = [
                dept['department'],
                dept['total_staff'],
                dept['staff_with_overtime'],
                dept['overtime_instances'],
                f"{coverage_pct:.1f}%"
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                
                # Color coding for coverage
                if col == 5:  # Coverage percentage
                    if coverage_pct >= 50:
                        cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                    elif coverage_pct >= 25:
                        cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        
        # Format columns
        column_widths = [18, 12, 14, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
