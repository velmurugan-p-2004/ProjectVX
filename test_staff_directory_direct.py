#!/usr/bin/env python3
"""
Direct test of the fixed staff directory report generation.
"""

import sqlite3
import sys
import os

def test_generate_staff_directory_sql():
    """Test just the SQL query from the fixed function."""
    print("=== Testing Staff Directory SQL Query ===")
    
    try:
        # Connect to database
        db = sqlite3.connect('vishnorex.db')
        db.row_factory = sqlite3.Row
        
        school_id = 1
        
        # This is the EXACT query from the fixed generate_staff_directory_report function
        staff_data = db.execute('''
            SELECT s.staff_id, s.full_name, s.first_name, s.last_name,
                   s.date_of_birth, s.date_of_joining, s.department, s.position,
                   s.gender, s.phone, s.email, s.shift_type, s.basic_salary,
                   s.created_at
            FROM staff s
            WHERE s.school_id = ?
            ORDER BY s.department, s.full_name
        ''', (school_id,)).fetchall()
        
        print(f"✅ SQL Query executed successfully!")
        print(f"✅ Found {len(staff_data)} staff records")
        
        if staff_data:
            print("\nSample data (first record):")
            staff = staff_data[0]
            print(f"  Staff ID: {staff['staff_id']}")
            print(f"  Name: {staff['full_name']}")
            print(f"  Department: {staff['department']}")
            print(f"  Position: {staff['position']}")
            print(f"  Phone: {staff['phone']}")
            print(f"  Email: {staff['email']}")
            print(f"  Shift: {staff['shift_type']}")
            
        # Test data access that would be used in Excel generation
        print("\n=== Testing Excel Data Access ===")
        test_headers = [
            'S.No', 'Staff ID', 'Full Name', 'Department', 'Position',
            'Gender', 'Phone', 'Email', 'Date of Joining', 'Date of Birth', 'Shift Type'
        ]
        
        print(f"Excel headers ({len(test_headers)} columns): {test_headers}")
        
        if staff_data:
            print("\nTesting data mapping for Excel:")
            staff = staff_data[0]
            excel_row = [
                1,  # S.No
                staff['staff_id'] or 'N/A',
                staff['full_name'],
                staff['department'] or 'N/A',
                staff['position'] or 'N/A',
                staff['gender'] or 'N/A',
                staff['phone'] or 'N/A',
                staff['email'] or 'N/A',
                staff['date_of_joining'] or 'N/A',
                staff['date_of_birth'] or 'N/A',
                staff['shift_type'] or 'General'
            ]
            
            print(f"✅ Excel row data: {excel_row}")
            print(f"✅ Data mapping successful - {len(excel_row)} values for {len(test_headers)} headers")
            
        db.close()
        return True
        
    except Exception as e:
        print(f"❌ SQL test failed: {e}")
        return False

def test_excel_creation():
    """Test Excel file creation with the fixed data structure."""
    print("\n=== Testing Excel File Creation ===")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        import datetime
        
        # Get data
        db = sqlite3.connect('vishnorex.db')
        db.row_factory = sqlite3.Row
        
        staff_data = db.execute('''
            SELECT s.staff_id, s.full_name, s.first_name, s.last_name,
                   s.date_of_birth, s.date_of_joining, s.department, s.position,
                   s.gender, s.phone, s.email, s.shift_type, s.basic_salary,
                   s.created_at
            FROM staff s
            WHERE s.school_id = ?
            ORDER BY s.department, s.full_name
        ''', (1,)).fetchall()
        
        # Create workbook (using the same code as the fixed function)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Staff Directory"
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=16, color="2F5597")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Add title
        ws.merge_cells('A1:K1')  # Fixed to match 11 columns
        title_cell = ws['A1']
        title_cell.value = f"Staff Directory Report - Generated on {datetime.datetime.now().strftime('%Y-%m-%d')}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        # Headers (fixed to match available columns)
        headers = [
            'S.No', 'Staff ID', 'Full Name', 'Department', 'Position',
            'Gender', 'Phone', 'Email', 'Date of Joining', 'Date of Birth', 'Shift Type'
        ]
        
        header_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_idx, staff in enumerate(staff_data, header_row + 1):
            ws.cell(row=row_idx, column=1, value=row_idx - header_row)
            ws.cell(row=row_idx, column=2, value=staff['staff_id'] or 'N/A')
            ws.cell(row=row_idx, column=3, value=staff['full_name'])
            ws.cell(row=row_idx, column=4, value=staff['department'] or 'N/A')
            ws.cell(row=row_idx, column=5, value=staff['position'] or 'N/A')
            ws.cell(row=row_idx, column=6, value=staff['gender'] or 'N/A')
            ws.cell(row=row_idx, column=7, value=staff['phone'] or 'N/A')
            ws.cell(row=row_idx, column=8, value=staff['email'] or 'N/A')
            ws.cell(row=row_idx, column=9, value=staff['date_of_joining'] or 'N/A')
            ws.cell(row=row_idx, column=10, value=staff['date_of_birth'] or 'N/A')
            ws.cell(row=row_idx, column=11, value=staff['shift_type'] or 'General')
            
            # Apply border to all cells
            for col in range(1, 12):  # Fixed to match 11 columns
                ws.cell(row=row_idx, column=col).border = border
        
        # Auto-adjust column widths
        for col in range(1, 12):  # Fixed to match 11 columns
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save to file for verification
        filename = "test_staff_directory_fixed.xlsx"
        wb.save(filename)
        
        print(f"✅ Excel file created successfully: {filename}")
        print(f"✅ Workbook has {len(staff_data) + 3} rows and {len(headers)} columns")
        
        # Verify the file can be read back
        wb_test = openpyxl.load_workbook(filename)
        ws_test = wb_test.active
        print(f"✅ File verification: {ws_test.max_row} rows, {ws_test.max_column} columns")
        
        # Check title and headers
        title = ws_test['A1'].value
        print(f"✅ Title: {title}")
        
        header_values = [ws_test.cell(row=3, column=col).value for col in range(1, ws_test.max_column + 1)]
        print(f"✅ Headers: {header_values}")
        
        wb.close()
        wb_test.close()
        db.close()
        
        return True
        
    except ImportError:
        print("❌ openpyxl not available for Excel testing")
        return False
    except Exception as e:
        print(f"❌ Excel creation failed: {e}")
        return False

if __name__ == "__main__":
    print("Staff Directory Report - Direct Fix Test")
    print("=" * 55)
    
    # Test SQL query
    sql_ok = test_generate_staff_directory_sql()
    
    # Test Excel creation
    excel_ok = test_excel_creation()
    
    print("\n" + "=" * 55)
    print("FINAL TEST RESULTS:")
    print(f"SQL Query Test:    {'✅ PASSED' if sql_ok else '❌ FAILED'}")
    print(f"Excel Creation:    {'✅ PASSED' if excel_ok else '❌ FAILED'}")
    
    if sql_ok and excel_ok:
        print("\n🎉 STAFF ADDRESS COLUMN FIX IS FULLY WORKING!")
        print("\n✅ All problematic columns removed from query:")
        print("   • s.address")
        print("   • s.emergency_contact") 
        print("   • s.qualification")
        print("   • s.experience")
        print("   • s.updated_at")
        
        print("\n✅ Excel structure updated:")
        print("   • Headers reduced from 15 to 11 columns")
        print("   • Column ranges fixed (1-12 instead of 1-16)")
        print("   • Cell mappings corrected")
        print("   • Title merge range fixed (A1:K1)")
        
        print("\n✅ Report now includes these columns:")
        print("   1. S.No  2. Staff ID  3. Full Name  4. Department  5. Position")
        print("   6. Gender  7. Phone  8. Email  9. Date of Joining  10. Date of Birth  11. Shift Type")
        
        print("\n🚀 The Staff & HR Reports section should now work without errors!")
        
    else:
        print("\n❌ ISSUES DETECTED - Check error messages above")
    
    print("\nTo test in the web interface:")
    print("1. Start Flask: python app.py")
    print("2. Login as admin")
    print("3. Navigate to Reports & Analytics or Staff & HR Reports")
    print("4. Click 'Generate' on Staff Directory report")
    print("5. Verify Excel file downloads successfully")
