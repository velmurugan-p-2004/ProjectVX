from flask import Flask, render_template, request, redirect, url_for, session, jsonify, g, make_response
import sqlite3
from flask_wtf.csrf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import datetime
import calendar
import time
from database import get_db, init_db
from zk_biometric import sync_attendance_from_device, ZKBiometricDevice, verify_staff_biometric, process_device_attendance_automatically
from shift_management import ShiftManager
from excel_reports import ExcelReportGenerator
from staff_management_enhanced import StaffManager
from attendance_advanced import AdvancedAttendanceManager
from reporting_dashboard import ReportingDashboard
from data_visualization import DataVisualization
from notification_system import NotificationManager
from backup_manager import BackupManager
from salary_calculator import SalaryCalculator
import os
import json
import calendar

# Import cloud modules (with fallback for backward compatibility)
try:
    from cloud_api import cloud_api
    from cloud_connector import start_cloud_connector, stop_cloud_connector, get_cloud_connector
    from cloud_config import get_cloud_config, get_device_config
    CLOUD_ENABLED = True
except ImportError:
    print("Cloud modules not available. Running in Ethernet-only mode.")
    CLOUD_ENABLED = False

# Create Flask app instance
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24).hex())

# Initialize CSRF protection
csrf = CSRFProtect(app)

# Initialize database with the app
init_db(app)

# Add custom Jinja2 filters
@app.template_filter('dateformat')
def dateformat_filter(date, format='%Y-%m-%d'):
    """Format a date using strftime"""
    if date is None:
        return ""
    if isinstance(date, str):
        try:
            # Try to parse string date
            date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            try:
                # Try alternative format
                date = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S').date()
            except ValueError:
                return date  # Return as-is if can't parse
    return date.strftime(format)

@app.template_filter('timeformat')
def timeformat_filter(time, format='%I:%M %p'):
    """Format a time using strftime in 12-hour format"""
    if time is None:
        return "--:--"
    if isinstance(time, str):
        try:
            # Try to parse string datetime with full format
            time_obj = datetime.datetime.strptime(time, '%Y-%m-%d %H:%M:%S')
            return time_obj.strftime(format)
        except ValueError:
            try:
                # Try to parse string time only
                time_obj = datetime.datetime.strptime(time, '%H:%M:%S').time()
                return datetime.datetime.combine(datetime.date.today(), time_obj).strftime(format)
            except ValueError:
                try:
                    # Try alternative format
                    time_obj = datetime.datetime.strptime(time, '%H:%M').time()
                    return datetime.datetime.combine(datetime.date.today(), time_obj).strftime(format)
                except ValueError:
                    return time  # Return as-is if can't parse
    # Convert to 12-hour format if it's a time object
    return datetime.datetime.combine(datetime.date.today(), time).strftime(format)

@app.template_filter('datetimeformat')
def datetimeformat_filter(datetime_obj, format='%Y-%m-%d %I:%M %p'):
    """Format a datetime using strftime in 12-hour format"""
    if datetime_obj is None:
        return ""
    if isinstance(datetime_obj, str):
        try:
            # Try to parse string datetime
            datetime_obj = datetime.datetime.strptime(datetime_obj, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            try:
                # Try alternative format
                datetime_obj = datetime.datetime.strptime(datetime_obj, '%Y-%m-%d')
            except ValueError:
                return datetime_obj  # Return as-is if can't parse
    return datetime_obj.strftime(format)

@app.template_filter('capitalize_first')
def capitalize_first_filter(text):
    """Capitalize first letter of text"""
    if not text:
        return ""
    return text[0].upper() + text[1:] if len(text) > 1 else text.upper()

@app.template_filter('simple_date')
def simple_date_filter(date_str):
    """Convert YYYY-MM-DD to readable format"""
    if not date_str or len(date_str) < 10:
        return date_str
    try:
        date_obj = datetime.datetime.strptime(date_str[:10], '%Y-%m-%d')
        return date_obj.strftime('%B %d, %Y')
    except ValueError:
        return date_str

# Register cloud API blueprint if available
if CLOUD_ENABLED:
    app.register_blueprint(cloud_api)
    print("Cloud API endpoints registered")

# Ensure on_duty_permissions table exists
def ensure_on_duty_permission_table():
    db = get_db()
    table_exists = db.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='on_duty_permissions'").fetchone()
    if not table_exists:
        db.execute('''
            CREATE TABLE IF NOT EXISTS on_duty_permissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER,
                school_id INTEGER,
                permission_type TEXT,
                start_datetime TEXT,
                end_datetime TEXT,
                reason TEXT,
                status TEXT DEFAULT 'pending',
                applied_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        db.commit()

with app.app_context():
    ensure_on_duty_permission_table()


# Route for admin to process OD permission (ensure only one definition)
@app.route('/process_on_duty_permission', methods=['POST'])
def process_on_duty_permission():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    permission_id = request.form.get('permission_id')
    decision = request.form.get('decision')  # 'approve' or 'reject'
    admin_id = session['user_id']
    processed_at = datetime.datetime.now()

    db = get_db()
    permission = db.execute('SELECT * FROM on_duty_permissions WHERE id = ?', (permission_id,)).fetchone()
    if not permission:
        return jsonify({'success': False, 'error': 'Permission not found'})

    status = 'approved' if decision == 'approve' else 'rejected'
    db.execute('''
        UPDATE on_duty_permissions
        SET status = ?, processed_by = ?, processed_at = ?
        WHERE id = ?
    ''', (status, admin_id, processed_at, permission_id))

    if status == 'approved':
        staff_id = permission['staff_id']
        school_id = permission['school_id']
        # Accept both '%Y-%m-%d %H:%M' and '%Y-%m-%d %H:%M:%S' formats
        try:
            start_dt = datetime.datetime.strptime(permission['start_datetime'], '%Y-%m-%d %H:%M')
        except ValueError:
            start_dt = datetime.datetime.strptime(permission['start_datetime'], '%Y-%m-%d %H:%M:%S')
        try:
            end_dt = datetime.datetime.strptime(permission['end_datetime'], '%Y-%m-%d %H:%M')
        except ValueError:
            end_dt = datetime.datetime.strptime(permission['end_datetime'], '%Y-%m-%d %H:%M:%S')
        current_dt = start_dt
        while current_dt.date() <= end_dt.date():
            date_str = current_dt.strftime('%Y-%m-%d')
            existing = db.execute('SELECT * FROM attendance WHERE staff_id = ? AND date = ?', (staff_id, date_str)).fetchone()
            # Mark as present and OD
            if existing:
                db.execute('UPDATE attendance SET status = ? WHERE staff_id = ? AND date = ?', ('OD', staff_id, date_str))
            else:
                db.execute('INSERT INTO attendance (staff_id, school_id, date, status) VALUES (?, ?, ?, ?)', (staff_id, school_id, date_str, 'OD'))
            current_dt += datetime.timedelta(days=1)
    db.commit()
    return jsonify({'success': True})



# File upload configuration
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.teardown_appcontext
def close_db(error):
    """Close database connection at the end of request"""
    _ = error  # Suppress unused parameter warning
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

# In app.py, update the index route
@app.route('/')
def index():
    db = get_db()

    # First check if the column exists
    columns = db.execute("PRAGMA table_info(schools)").fetchall()
    has_is_hidden = any(col['name'] == 'is_hidden' for col in columns)

    if has_is_hidden:
        schools = db.execute('SELECT id, name FROM schools WHERE is_hidden = 0 OR is_hidden IS NULL ORDER BY name').fetchall()
    else:
        schools = db.execute('SELECT id, name FROM schools ORDER BY name').fetchall()

    return render_template('index.html', schools=schools)

# Routes
@app.route('/company_login', methods=['GET', 'POST'])
def handle_company_login():
    if request.method == 'GET':
        return render_template('company_login.html')

    # Handle POST request
    username = request.form.get('username')
    password = request.form.get('password')

    db = get_db()
    company_admin = db.execute('''
        SELECT * FROM company_admins WHERE username = ?
    ''', (username,)).fetchone()

    if not company_admin:
        return jsonify({'error': 'Company admin not found'}), 401

    if not check_password_hash(company_admin['password'], password):
        return jsonify({'error': 'Invalid password'}), 401

    session['user_id'] = company_admin['id']
    session['user_type'] = 'company_admin'
    session['full_name'] = company_admin['full_name']
    return jsonify({'redirect': url_for('company_dashboard')})

@app.route('/login', methods=['POST'])
@csrf.exempt  # Exempt from CSRF for easier login handling
def handle_school_login():
    school_id = request.form.get('school_id')
    username = request.form.get('username')
    password = request.form.get('password')

    print(f"Login attempt - School ID: {school_id}, Username: {username}")  # Debug log

    if not school_id:
        return jsonify({'error': 'Please select a school'}), 400

    db = get_db()

    # Check school admin
    admin = db.execute('''
        SELECT * FROM admins
        WHERE school_id = ? AND username = ?
    ''', (school_id, username)).fetchone()

    if admin and check_password_hash(admin['password'], password):
        print("Admin login successful")  # Debug log
        session['user_id'] = admin['id']
        session['school_id'] = admin['school_id']
        session['user_type'] = 'admin'
        session['full_name'] = admin['full_name']
        return jsonify({'redirect': url_for('admin_dashboard')})

    # Check staff - using username as staff_id
    staff = db.execute('''
        SELECT * FROM staff
        WHERE school_id = ? AND staff_id = ?
    ''', (school_id, username)).fetchone()

    if staff:
        password_hash = staff['password_hash'] if staff['password_hash'] is not None else ''
        print(f"Staff found: {staff['full_name']}, Has password hash: {bool(password_hash)}")  # Debug log

        # Check if password hash exists and verify password
        if password_hash and check_password_hash(password_hash, password):
            print("Staff login successful")  # Debug log
            session['user_id'] = staff['id']
            session['school_id'] = staff['school_id']
            session['user_type'] = 'staff'
            session['full_name'] = staff['full_name']
            return jsonify({'redirect': url_for('staff_dashboard')})
        elif not password_hash:
            print("Staff has no password set")  # Debug log
            return jsonify({'error': 'Password not set for this staff member. Please contact admin.'}), 401
        else:
            print("Password verification failed")  # Debug log

    print("Login failed - invalid credentials")  # Debug log
    return jsonify({'error': 'Invalid credentials'}), 401

# Add these new routes to app.py

@app.route('/company/school_details/<int:school_id>')
def school_details(school_id):
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return redirect(url_for('index'))

    db = get_db()

    # Get school info
    school = db.execute('SELECT * FROM schools WHERE id = ?', (school_id,)).fetchone()
    if not school:
        return redirect(url_for('company_dashboard'))

    # Get admins
    admins = db.execute('''
        SELECT id, username, full_name, email
        FROM admins
        WHERE school_id = ?
    ''', (school_id,)).fetchall()

    # Get staff
    staff = db.execute('''
        SELECT id, staff_id, full_name, department, position, email, phone
        FROM staff
        WHERE school_id = ?
        ORDER BY full_name
    ''', (school_id,)).fetchall()

    # Get attendance summary
    today = datetime.date.today()
    attendance_summary = db.execute('''
        SELECT
            COUNT(*) as total_staff,
            SUM(CASE WHEN a.status = 'present' THEN 1 ELSE 0 END) as present,
            SUM(CASE WHEN a.status = 'absent' THEN 1 ELSE 0 END) as absent,
            SUM(CASE WHEN a.status = 'late' THEN 1 ELSE 0 END) as late,
            SUM(CASE WHEN a.status = 'leave' THEN 1 ELSE 0 END) as on_leave,
            SUM(CASE WHEN a.status = 'on_duty' THEN 1 ELSE 0 END) as on_duty
        FROM (
            SELECT s.id, COALESCE(a.status, 'absent') as status
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id AND a.date = ?
            WHERE s.school_id = ?
        ) a
    ''', (today, school_id)).fetchone()

    # Get pending leaves
    pending_leaves = db.execute('''
        SELECT l.id, s.full_name, l.leave_type, l.start_date, l.end_date, l.reason
        FROM leave_applications l
        JOIN staff s ON l.staff_id = s.id
        WHERE l.school_id = ? AND l.status = 'pending'
        ORDER BY l.applied_at
    ''', (school_id,)).fetchall()

    return render_template('school_details.html',
                         school=school,
                         admins=admins,
                         staff=staff,
                         attendance_summary=attendance_summary,
                         pending_leaves=pending_leaves,
                         today=today)

@app.route('/get_attendance_summary')
def get_attendance_summary():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    db = get_db()

    if session['user_type'] == 'staff':
        staff_id = session['user_id']
        today = datetime.date.today()

        # Get current month attendance
        first_day = today.replace(day=1)
        last_day = (today.replace(day=28) + datetime.timedelta(days=4)).replace(day=1) - datetime.timedelta(days=1)

        attendance = db.execute('''
            SELECT status, COUNT(*) as count
            FROM attendance
            WHERE staff_id = ? AND date BETWEEN ? AND ?
            GROUP BY status
        ''', (staff_id, first_day, last_day)).fetchall()

        # Initialize counts
        present = 0
        absent = 0
        late = 0
        leave = 0

        for record in attendance:
            if record['status'] == 'present':
                present = record['count']
            elif record['status'] == 'absent':
                absent = record['count']
            elif record['status'] == 'late':
                late = record['count']
            elif record['status'] == 'leave':
                leave = record['count']

        return jsonify({
            'success': True,
            'present': present,
            'absent': absent,
            'late': late,
            'leave': leave
        })

    return jsonify({'success': False, 'error': 'Unauthorized'})

@app.route('/get_staff_details')
def get_staff_details():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.args.get('id')
    db = get_db()

    staff = db.execute('SELECT * FROM staff WHERE id = ?', (staff_id,)).fetchone()
    if not staff:
        return jsonify({'success': False, 'error': 'Staff not found'})

    # Get attendance records
    attendance = db.execute('''
        SELECT date, time_in, time_out, status
        FROM attendance
        WHERE staff_id = ?
        ORDER BY date DESC
    ''', (staff_id,)).fetchall()

    return jsonify({
        'success': True,
        'staff': dict(staff),
        'attendance': [dict(a) for a in attendance]
    })


@app.route('/export_staff_data')
def export_staff_data():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range is required'})

    try:
        # Validate date format
        datetime.datetime.strptime(start_date, '%Y-%m-%d')
        datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'})

    # Generate Excel report
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)

@app.route('/add_admin', methods=['POST'])
def add_admin():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 401

    # Get form data
    school_id = request.form.get('school_id')
    username = request.form.get('username')
    password = request.form.get('password')
    full_name = request.form.get('full_name')
    email = request.form.get('email')

    # Validate required fields
    if not school_id:
        return jsonify({'success': False, 'error': 'School ID is required'}), 400
    
    if not username or not username.strip():
        return jsonify({'success': False, 'error': 'Username is required'}), 400
    
    if not password or not password.strip():
        return jsonify({'success': False, 'error': 'Password is required'}), 400
        
    if not full_name or not full_name.strip():
        return jsonify({'success': False, 'error': 'Full name is required'}), 400

    # Validate school exists
    db = get_db()
    school = db.execute('SELECT id FROM schools WHERE id = ?', (school_id,)).fetchone()
    if not school:
        return jsonify({'success': False, 'error': 'Invalid school ID'}), 400

    # Clean data
    username = username.strip()
    full_name = full_name.strip()
    email = email.strip() if email else None
    
    # Hash password
    password_hash = generate_password_hash(password)

    try:
        db.execute('''
            INSERT INTO admins (school_id, username, password, full_name, email)
            VALUES (?, ?, ?, ?, ?)
        ''', (school_id, username, password_hash, full_name, email))
        db.commit()
        return jsonify({'success': True, 'message': 'Administrator added successfully'})
    except sqlite3.IntegrityError as e:
        if 'username' in str(e).lower():
            return jsonify({'success': False, 'error': 'Username already exists'}), 400
        else:
            return jsonify({'success': False, 'error': 'Database constraint error'}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': 'An unexpected error occurred'}), 500

@app.route('/delete_admin', methods=['POST'])
def delete_admin():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 401

    admin_id = request.form.get('admin_id')
    
    if not admin_id:
        return jsonify({'success': False, 'error': 'Admin ID is required'}), 400

    db = get_db()

    try:
        # Check if admin exists
        admin = db.execute('SELECT id FROM admins WHERE id = ?', (admin_id,)).fetchone()
        if not admin:
            return jsonify({'success': False, 'error': 'Administrator not found'}), 404

        db.execute('DELETE FROM admins WHERE id = ?', (admin_id,))
        db.commit()
        return jsonify({'success': True, 'message': 'Administrator deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': 'An error occurred while deleting administrator'}), 500



@app.route('/get_biometric_verifications')
def get_biometric_verifications():
    """Get biometric verification history for staff"""
    if 'user_id' not in session or session['user_type'] != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    db = get_db()

    # Get verification history
    query = '''
        SELECT verification_type, verification_time, biometric_method,
               verification_status, notes, device_ip
        FROM biometric_verifications
        WHERE staff_id = ?
    '''
    params = [staff_id]

    if start_date and end_date:
        query += ' AND DATE(verification_time) BETWEEN ? AND ?'
        params.extend([start_date, end_date])

    query += ' ORDER BY verification_time DESC LIMIT 50'

    verifications = db.execute(query, params).fetchall()

    return jsonify({
        'success': True,
        'verifications': [dict(v) for v in verifications]
    })

@app.route('/get_today_attendance_status')
def get_today_attendance_status():
    """Get today's attendance status for staff"""
    if 'user_id' not in session or session['user_type'] != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    today = datetime.date.today()

    db = get_db()

    # Get today's attendance
    attendance = db.execute('''
        SELECT time_in, time_out, status
        FROM attendance
        WHERE staff_id = ? AND date = ?
    ''', (staff_id, today)).fetchone()

    # Get today's verifications
    verifications = db.execute('''
        SELECT verification_type, verification_time, biometric_method, verification_status
        FROM biometric_verifications
        WHERE staff_id = ? AND DATE(verification_time) = ?
        ORDER BY verification_time DESC
    ''', (staff_id, today)).fetchall()

    # Determine available actions based on current status
    available_actions = []
    # Check-in and check-out are always available (can be updated multiple times)
    available_actions.append('check-in')
    available_actions.append('check-out')

    # Format attendance times to 12-hour format
    formatted_attendance = format_attendance_times_to_12hr(attendance) if attendance else None

    return jsonify({
        'success': True,
        'attendance': formatted_attendance,
        'verifications': [dict(v) for v in verifications],
        'available_actions': available_actions
    })

@app.route('/get_realtime_attendance')
def get_realtime_attendance():
    """Get real-time attendance data for admin dashboard"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session.get('school_id', 1)
    today = datetime.date.today()

    db = get_db()

    # Get today's attendance details for all staff
    today_attendance = db.execute('''
        SELECT s.id as staff_id, s.staff_id as staff_number, s.full_name, s.department,
               a.time_in, a.time_out,
               COALESCE(a.status, 'absent') as status
        FROM staff s
        LEFT JOIN attendance a ON s.id = a.staff_id AND a.date = ?
        WHERE s.school_id = ?
        ORDER BY s.full_name
    ''', (today, school_id)).fetchall()

    # Get attendance summary
    attendance_summary = db.execute('''
        SELECT
            COUNT(*) as total_staff,
            SUM(CASE WHEN a.status = 'present' THEN 1 ELSE 0 END) as present,
            SUM(CASE WHEN a.status = 'absent' THEN 1 ELSE 0 END) as absent,
            SUM(CASE WHEN a.status = 'late' THEN 1 ELSE 0 END) as late,
            SUM(CASE WHEN a.status = 'leave' THEN 1 ELSE 0 END) as on_leave
        FROM (
            SELECT s.id, COALESCE(a.status, 'absent') as status
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id AND a.date = ?
            WHERE s.school_id = ?
        ) a
    ''', (today, school_id)).fetchone()

    # Format attendance times to 12-hour format
    formatted_attendance = [format_attendance_times_to_12hr(dict(row)) for row in today_attendance]

    return jsonify({
        'success': True,
        'attendance_data': formatted_attendance,
        'summary': dict(attendance_summary) if attendance_summary else {},
        'timestamp': datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')
    })

@app.route('/export_company_report')
def export_company_report():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range is required'})

    try:
        # Validate date format
        datetime.datetime.strptime(start_date, '%Y-%m-%d')
        datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'})

    # Generate Excel report
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_company_report(start_date, end_date)

@app.route('/export_staff_report')
def export_staff_report():
    if 'user_id' not in session or session['user_type'] != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range required'})

    try:
        # Validate date format
        datetime.datetime.strptime(start_date, '%Y-%m-%d')
        datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'})

    # Generate Excel report
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_individual_staff_report(staff_id, start_date, end_date)

@app.route('/export_monthly_report')
def export_monthly_report():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    if not year or not month:
        return jsonify({'success': False, 'error': 'Year and month are required'})

    if month < 1 or month > 12:
        return jsonify({'success': False, 'error': 'Invalid month. Must be between 1 and 12'})

    # Generate Excel report
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_monthly_report(school_id, year, month)

@app.route('/export_department_report')
def export_department_report():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    department = request.args.get('department')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range is required'})

    try:
        # Validate date format
        datetime.datetime.strptime(start_date, '%Y-%m-%d')
        datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'})

    # Generate Excel report (using staff attendance report filtered by department)
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)

@app.route('/export_yearly_report')
def export_yearly_report():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    year = request.args.get('year', type=int)

    if not year:
        return jsonify({'success': False, 'error': 'Year is required'})

    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"

    # Generate Excel report
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)

# Enhanced Staff Management Routes
@app.route('/bulk_import_staff', methods=['POST'])
def bulk_import_staff():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'})

    # Save uploaded file temporarily
    filename = secure_filename(file.filename)
    temp_path = os.path.join('temp', filename)
    os.makedirs('temp', exist_ok=True)
    file.save(temp_path)

    try:
        staff_manager = StaffManager()
        result = staff_manager.bulk_import_staff(temp_path, school_id)
        return jsonify(result)
    finally:
        # Clean up temp file
        if os.path.exists(temp_path):
            os.remove(temp_path)

@app.route('/advanced_search_staff')
def advanced_search_staff():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']

    filters = {
        'search_term': request.args.get('search_term'),
        'department': request.args.get('department'),
        'position': request.args.get('position'),
        'gender': request.args.get('gender'),
        'date_from': request.args.get('date_from'),
        'date_to': request.args.get('date_to'),
        'limit': request.args.get('limit')  # Remove default limit to match server-side query
    }

    # Remove None values
    filters = {k: v for k, v in filters.items() if v}

    staff_manager = StaffManager()
    staff_list = staff_manager.advanced_search_staff(school_id, filters)

    return jsonify({'success': True, 'staff': staff_list})

@app.route('/upload_staff_photo', methods=['POST'])
def upload_staff_photo():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id')
    if not staff_id:
        return jsonify({'success': False, 'error': 'Staff ID required'})

    if 'photo' not in request.files:
        return jsonify({'success': False, 'error': 'No photo uploaded'})

    photo_file = request.files['photo']
    if photo_file.filename == '':
        return jsonify({'success': False, 'error': 'No photo selected'})

    staff_manager = StaffManager()
    result = staff_manager.manage_staff_photo(staff_id, photo_file)

    return jsonify(result)

@app.route('/get_department_analytics')
def get_department_analytics():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    staff_manager = StaffManager()
    analytics = staff_manager.get_department_analytics(school_id)

    return jsonify({'success': True, 'analytics': analytics})

@app.route('/generate_staff_id')
def generate_staff_id():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    department = request.args.get('department')

    staff_manager = StaffManager()
    staff_id = staff_manager.generate_staff_id(school_id, department)

    return jsonify({'success': True, 'staff_id': staff_id})

@app.route('/bulk_update_staff', methods=['POST'])
def bulk_update_staff():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    updates = request.json.get('updates', [])

    if not updates:
        return jsonify({'success': False, 'error': 'No updates provided'})

    staff_manager = StaffManager()
    result = staff_manager.bulk_update_staff(updates, school_id)

    return jsonify(result)

@app.route('/staff_management_enhanced')
def staff_management_enhanced():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return redirect(url_for('index'))

    return render_template('staff_management_enhanced.html', today=datetime.datetime.now().strftime('%Y-%m-%d'))

@app.route('/download_staff_template')
def download_staff_template():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Create a sample Excel template
    import pandas as pd
    from io import BytesIO

    # Sample data for template
    template_data = {
        'staff_id': ['EMP001', 'EMP002'],
        'full_name': ['John Doe', 'Jane Smith'],
        'email': ['john.doe@example.com', 'jane.smith@example.com'],
        'phone': ['1234567890', '0987654321'],
        'department': ['IT', 'HR'],
        'position': ['Developer', 'Manager'],
        'gender': ['Male', 'Female'],
        'date_of_birth': ['1990-01-01', '1985-05-15'],
        'date_of_joining': ['2023-01-01', '2023-02-01']
    }

    df = pd.DataFrame(template_data)

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Staff Template', index=False)

    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=staff_import_template.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

# Advanced Attendance Management Routes
@app.route('/process_attendance_advanced', methods=['POST'])
def process_attendance_advanced():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id')
    verification_type = request.form.get('verification_type')  # check-in, check-out, overtime-in, overtime-out
    timestamp_str = request.form.get('timestamp')

    if not all([staff_id, verification_type, timestamp_str]):
        return jsonify({'success': False, 'error': 'Missing required parameters'})

    try:
        timestamp = datetime.datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
        school_id = session['school_id']

        attendance_manager = AdvancedAttendanceManager()
        result = attendance_manager.process_attendance_with_overtime(
            int(staff_id), school_id, verification_type, timestamp
        )

        return jsonify(result)

    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid timestamp format'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_overtime_summary')
def get_overtime_summary():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Staff can view their own overtime, admin can view any staff's overtime
    if session['user_type'] == 'staff':
        staff_id = session['user_id']
    else:
        staff_id = request.args.get('staff_id')
        if not staff_id:
            return jsonify({'success': False, 'error': 'Staff ID required'})

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range required'})

    attendance_manager = AdvancedAttendanceManager()
    result = attendance_manager.get_overtime_summary(int(staff_id), start_date, end_date)

    return jsonify(result)

@app.route('/create_regularization_request', methods=['POST'])
def create_regularization_request():
    if 'user_id' not in session or session['user_type'] != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    school_id = session['school_id']
    attendance_id = request.form.get('attendance_id')
    request_type = request.form.get('request_type')  # late_arrival, early_departure
    original_time = request.form.get('original_time')
    expected_time = request.form.get('expected_time')
    reason = request.form.get('reason')

    if not all([attendance_id, request_type, original_time, expected_time, reason]):
        return jsonify({'success': False, 'error': 'All fields are required'})

    attendance_manager = AdvancedAttendanceManager()
    result = attendance_manager.create_attendance_regularization_request(
        staff_id, school_id, int(attendance_id), request_type,
        original_time, expected_time, reason
    )

    return jsonify(result)

@app.route('/process_regularization_request', methods=['POST'])
def process_regularization_request():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    request_id = request.form.get('request_id')
    decision = request.form.get('decision')  # approve, reject
    admin_reason = request.form.get('admin_reason')
    admin_id = session['user_id']

    if not all([request_id, decision]):
        return jsonify({'success': False, 'error': 'Request ID and decision are required'})

    attendance_manager = AdvancedAttendanceManager()
    result = attendance_manager.process_regularization_request(
        int(request_id), admin_id, decision, admin_reason
    )

    return jsonify(result)

@app.route('/get_attendance_analytics')
def get_attendance_analytics():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range required'})

    attendance_manager = AdvancedAttendanceManager()
    result = attendance_manager.get_attendance_analytics(school_id, start_date, end_date)

    return jsonify(result)

@app.route('/auto_mark_leave_attendance', methods=['POST'])
def auto_mark_leave_attendance():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    date = request.form.get('date', datetime.datetime.now().strftime('%Y-%m-%d'))

    attendance_manager = AdvancedAttendanceManager()
    result = attendance_manager.auto_mark_leave_attendance(school_id, date)

    return jsonify(result)

# Reporting Dashboard Routes
@app.route('/reporting_dashboard')
def reporting_dashboard():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return redirect(url_for('index'))

    today = datetime.datetime.now()
    return render_template('reporting_dashboard.html',
                         today=today.strftime('%Y-%m-%d'),
                         current_month=today.strftime('%Y-%m'),
                         current_year=today.year)

@app.route('/salary_management')
def salary_management():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session['user_type'] not in ['admin', 'company_admin']:
        return redirect(url_for('staff_dashboard'))
    return render_template('salary_management.html')

@app.route('/get_summary_dashboard')
def get_summary_dashboard():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    reporting_dashboard = ReportingDashboard()
    summary = reporting_dashboard.get_summary_dashboard(school_id)

    return jsonify({'success': True, 'summary': summary})

@app.route('/generate_report')
def generate_report():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    report_type = request.args.get('report_type')

    if not report_type:
        return jsonify({'success': False, 'error': 'Report type is required'})

    reporting_dashboard = ReportingDashboard()

    try:
        if report_type == 'daily':
            date = request.args.get('date')
            if not date:
                return jsonify({'success': False, 'error': 'Date is required for daily report'})
            report = reporting_dashboard.generate_daily_report(school_id, date)

        elif report_type == 'weekly':
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            if not start_date or not end_date:
                return jsonify({'success': False, 'error': 'Date range is required for weekly report'})
            report = reporting_dashboard.generate_weekly_report(school_id, start_date, end_date)

        elif report_type == 'monthly':
            year = request.args.get('year', type=int)
            month = request.args.get('month', type=int)
            if not year or not month:
                return jsonify({'success': False, 'error': 'Year and month are required for monthly report'})
            report = reporting_dashboard.generate_monthly_report(school_id, year, month)

        elif report_type == 'yearly':
            year = request.args.get('year', type=int)
            if not year:
                return jsonify({'success': False, 'error': 'Year is required for yearly report'})
            report = reporting_dashboard.generate_yearly_report(school_id, year)

        elif report_type == 'department':
            department = request.args.get('department')
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            if not all([department, start_date, end_date]):
                return jsonify({'success': False, 'error': 'Department and date range are required'})
            report = reporting_dashboard.generate_department_report(school_id, department, start_date, end_date)

        elif report_type == 'custom':
            filters = {
                'start_date': request.args.get('start_date'),
                'end_date': request.args.get('end_date'),
                'department': request.args.get('department'),
                'position': request.args.get('position'),
                'status': request.args.get('status')
            }
            # Remove None values
            filters = {k: v for k, v in filters.items() if v}
            report = reporting_dashboard.generate_custom_report(school_id, filters)

        elif report_type == 'trends':
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            if not start_date or not end_date:
                return jsonify({'success': False, 'error': 'Date range is required for trends report'})
            report = reporting_dashboard.generate_trends_report(school_id, start_date, end_date)

        elif report_type == 'summary':
            report = reporting_dashboard.get_summary_dashboard(school_id)
            report['report_type'] = 'summary'

        else:
            return jsonify({'success': False, 'error': 'Invalid report type'})

        return jsonify({'success': True, 'report': report})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_departments')
def get_departments():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    db = get_db()

    departments = db.execute('''
        SELECT DISTINCT department as name
        FROM staff
        WHERE school_id = ? AND department IS NOT NULL AND department != ''
        ORDER BY department
    ''', (school_id,)).fetchall()

    return jsonify({
        'success': True,
        'departments': [dept['name'] for dept in departments]
    })

@app.route('/export_report_excel')
def export_report_excel():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    report_type = request.args.get('report_type')

    if not report_type:
        return jsonify({'success': False, 'error': 'Report type is required'})

    # Generate appropriate Excel report based on type
    excel_generator = ExcelReportGenerator()

    if report_type == 'daily':
        date = request.args.get('date', datetime.datetime.now().strftime('%Y-%m-%d'))
        return excel_generator.create_staff_attendance_report(school_id, date, date)
    elif report_type in ['weekly', 'custom', 'trends']:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        if start_date and end_date:
            return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)
    elif report_type == 'monthly':
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        if year and month:
            return excel_generator.create_monthly_report(school_id, year, month)
    elif report_type == 'yearly':
        year = request.args.get('year', type=int)
        if year:
            start_date = f"{year}-01-01"
            end_date = f"{year}-12-31"
            return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)

    # Fallback to general report
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    return excel_generator.create_staff_attendance_report(school_id, today, today)

@app.route('/generate_admin_report')
def generate_admin_report():
    """Generate and download reports from admin reports page"""
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    try:
        school_id = session['school_id']
        report_type = request.args.get('report_type')
        format_type = request.args.get('format', 'excel').lower()

        if not report_type:
            return jsonify({'success': False, 'error': 'Report type is required'})

        # Get filter parameters
        year = request.args.get('year', datetime.datetime.now().year, type=int)
        month = request.args.get('month', type=int)
        department = request.args.get('department', '')

        # Create Excel generator for all reports
        excel_generator = ExcelReportGenerator()

        # Route to appropriate report generation based on report_type
        if report_type == 'monthly_salary':
            return generate_monthly_salary_report(school_id, year, month, department, format_type)
        elif report_type == 'payroll_summary':
            return generate_payroll_summary_report(school_id, year, month, format_type)
        elif report_type == 'department_salary':
            return generate_department_salary_report(school_id, year, department, format_type)
        elif report_type == 'staff_directory':
            return generate_staff_directory_report(school_id, format_type)
        elif report_type == 'department_report':
            return generate_department_analysis_report(school_id, year, month, format_type)
        elif report_type == 'performance_report':
            return generate_performance_report(school_id, year, month, department, format_type)
        elif report_type == 'daily_attendance':
            date = request.args.get('date', datetime.datetime.now().strftime('%Y-%m-%d'))
            return generate_daily_attendance_report(school_id, date, department, format_type)
        elif report_type == 'monthly_attendance':
            if month:
                return excel_generator.create_monthly_report(school_id, year, month)
            else:
                # Current month
                return excel_generator.create_monthly_report(school_id, year, datetime.datetime.now().month)
        elif report_type == 'overtime_report':
            return generate_overtime_report(school_id, year, month, format_type)
        else:
            return jsonify({'success': False, 'error': f'Unknown report type: {report_type}'})

    except Exception as e:
        print(f"Report generation error: {str(e)}")
        return jsonify({'success': False, 'error': f'Report generation failed: {str(e)}'})

@app.route('/test_performance_report_json')
def test_performance_report_json():
    """Test endpoint for performance report that returns JSON data for validation"""
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})
    
    try:
        school_id = session['school_id']
        department = request.args.get('department', '')
        
        # Temporarily modify the generate_performance_report function to return JSON
        # by setting the format_type to 'json' which we'll handle in the function
        
        # Create a fake request context for the performance report function
        original_args = request.args
        
        # Call the performance report generator directly, but we need to modify it to return JSON
        from flask import g
        import datetime
        
        db = get_db()
        
        # Get from_date and to_date parameters
        from_date_str = request.args.get('from_date')
        to_date_str = request.args.get('to_date')
        
        # Date range logic (same as in the function)
        if from_date_str and to_date_str:
            try:
                start_date = datetime.datetime.strptime(from_date_str, '%Y-%m-%d').date()
                end_date = datetime.datetime.strptime(to_date_str, '%Y-%m-%d').date()
            except ValueError:
                today = datetime.date.today()
                start_date = datetime.date(today.year, today.month, 1)
                end_date = datetime.date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
        else:
            today = datetime.date.today()
            start_date = datetime.date(today.year, today.month, 1) 
            end_date = datetime.date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
        
        start_str, end_str = start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')
        
        # Build department filter
        dept_clause = ''
        staff_params = [school_id]
        if department and str(department).strip():
            dept_clause = ' AND COALESCE(s.department, "") = ?'
            staff_params.append(department)
        
        # Get staff data
        staff_performance_query = f"""
            SELECT 
                s.id as staff_db_id,
                s.staff_id,
                s.full_name,
                COALESCE(NULLIF(TRIM(s.department), ''), 'Unassigned') as department,
                COALESCE(NULLIF(TRIM(s.position), ''), 'Not Specified') as position,
                s.date_of_joining
            FROM staff s
            WHERE s.school_id = ? AND s.is_active = 1 {dept_clause}
            ORDER BY s.department, s.full_name
        """
        
        staff_rows = db.execute(staff_performance_query, tuple(staff_params)).fetchall()
        
        # Calculate working days (moved outside the loop)
        total_days = (end_date - start_date).days + 1
        working_days = 0
        current_dt = start_date
        weekend_days = 0
        
        # Get holidays
        holidays_in_period = db.execute("""
            SELECT start_date, end_date, holiday_name 
            FROM holidays 
            WHERE school_id = ? AND is_active = 1
            AND NOT (end_date < ? OR start_date > ?)
        """, (school_id, start_str, end_str)).fetchall()
        
        holiday_dates = set()
        for holiday in holidays_in_period:
            h_start = datetime.datetime.strptime(holiday['start_date'], '%Y-%m-%d').date()
            h_end = datetime.datetime.strptime(holiday['end_date'], '%Y-%m-%d').date()
            curr_date = h_start
            while curr_date <= h_end:
                holiday_dates.add(curr_date)
                curr_date += datetime.timedelta(days=1)
        
        # Count working days
        current_dt = start_date
        while current_dt <= end_date:
            if current_dt.weekday() < 5:  # Monday = 0, Sunday = 6
                if current_dt not in holiday_dates:
                    working_days += 1
            else:
                weekend_days += 1
            current_dt += datetime.timedelta(days=1)
        
        # Get attendance data for each staff
        final_performance_rows = []
        
        for staff in staff_rows:
            staff_db_id = staff['staff_db_id']
            
            # Attendance data
            attendance_data = db.execute("""
                SELECT 
                    COUNT(*) as total_records,
                    SUM(CASE WHEN status IN ('present', 'late') THEN 1 ELSE 0 END) as days_present,
                    SUM(CASE WHEN status = 'absent' THEN 1 ELSE 0 END) as days_absent,
                    SUM(CASE WHEN status = 'late' THEN 1 ELSE 0 END) as days_late
                FROM attendance 
                WHERE staff_id = ? AND date BETWEEN ? AND ?
            """, (staff_db_id, start_str, end_str)).fetchone()
            
            # Leave data
            leave_data = db.execute("""
                SELECT 
                    COALESCE(COUNT(CASE WHEN status = 'approved' THEN id END), 0) as approved_leave_count,
                    COALESCE(SUM(CASE WHEN status = 'approved' 
                                THEN julianday(end_date) - julianday(start_date) + 1 
                                ELSE 0 END), 0) as days_on_leave
                FROM leave_applications 
                WHERE staff_id = ? AND NOT (end_date < ? OR start_date > ?)
            """, (staff_db_id, start_str, end_str)).fetchone()
            
            # OD data  
            od_data = db.execute("""
                SELECT 
                    COALESCE(COUNT(CASE WHEN status = 'approved' THEN id END), 0) as approved_od_count,
                    COALESCE(SUM(CASE WHEN status = 'approved' 
                                THEN julianday(end_date) - julianday(start_date) + 1 
                                ELSE 0 END), 0) as days_on_od
                FROM on_duty_applications 
                WHERE staff_id = ? AND NOT (end_date < ? OR start_date > ?)
            """, (staff_db_id, start_str, end_str)).fetchone()
            
            # Permission data
            permission_data = db.execute("""
                SELECT 
                    COALESCE(COUNT(CASE WHEN status = 'approved' THEN id END), 0) as approved_permission_count,
                    COALESCE(SUM(CASE WHEN status = 'approved' THEN duration_hours ELSE 0 END), 0) as total_permission_hours
                FROM permission_applications 
                WHERE staff_id = ? AND permission_date BETWEEN ? AND ?
            """, (staff_db_id, start_str, end_str)).fetchone()
            
            permission_days = round(float(permission_data['total_permission_hours'] or 0) / 8, 1)
            
            final_performance_rows.append({
                'staff_id': staff['staff_id'],
                'staff_name': staff['full_name'],
                'department': staff['department'],
                'position': staff['position'],
                'days_present': int(attendance_data['days_present'] or 0),
                'days_absent': int(attendance_data['days_absent'] or 0),
                'days_late': int(attendance_data['days_late'] or 0),
                'days_on_od_applied': int(od_data['days_on_od'] or 0),
                'days_on_leave_applied': int(leave_data['days_on_leave'] or 0),
                'days_with_permission_applied': permission_days,
                'total_working_days': working_days,
                'total_attendance_records': int(attendance_data['total_records'] or 0)
            })
            
            # Only process first 3 staff for quick testing to avoid timeout
            if len(final_performance_rows) >= 3:
                break
        
        return jsonify({
            'success': True,
            'data': final_performance_rows,
            'summary': {
                'total_staff': len(final_performance_rows),
                'date_range': f'{start_str} to {end_str}',
                'department_filter': department if department else 'All Departments',
                'total_working_days': working_days,
                'weekend_days': weekend_days,
                'holiday_days': len(holiday_dates),
                'total_days': total_days
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Test failed: {str(e)}'})

def generate_monthly_salary_report(school_id, year, month, department, format_type):
    """Generate monthly salary report"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    db = get_db()

    # Build query based on filters
    where_conditions = ['s.school_id = ?']
    params = [school_id]

    if department:
        where_conditions.append('s.department = ?')
        params.append(department)

    # Get staff with salary information
    staff_query = f'''
        SELECT s.id, s.staff_id, s.full_name, s.department, s.position,
               s.basic_salary,
               (COALESCE(s.hra, 0) + COALESCE(s.transport_allowance, 0) + COALESCE(s.other_allowances, 0)) as allowances,
               (COALESCE(s.pf_deduction, 0) + COALESCE(s.esi_deduction, 0) + COALESCE(s.professional_tax, 0) + COALESCE(s.other_deductions, 0)) as deductions,
               COALESCE(s.basic_salary, 0) + (COALESCE(s.hra, 0) + COALESCE(s.transport_allowance, 0) + COALESCE(s.other_allowances, 0)) - (COALESCE(s.pf_deduction, 0) + COALESCE(s.esi_deduction, 0) + COALESCE(s.professional_tax, 0) + COALESCE(s.other_deductions, 0)) as net_salary,
               s.date_of_joining
        FROM staff s
        WHERE {' AND '.join(where_conditions)}
        ORDER BY s.department, s.full_name
    '''

    staff_data = db.execute(staff_query, params).fetchall()

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Salary Report"

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16, color="2F5597")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Add title
    month_name = calendar.month_name[month] if month else 'All Months'
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Monthly Salary Report - {month_name} {year}"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    # Add summary
    ws['A3'] = f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    if department:
        ws['A4'] = f"Department: {department}"

    # Headers
    headers = ['S.No', 'Staff ID', 'Name', 'Department', 'Position', 'Basic Salary', 'Allowances', 'Deductions', 'Net Salary']
    header_row = 6

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Add data
    total_basic = total_allowances = total_deductions = total_net = 0

    for row_idx, staff in enumerate(staff_data, header_row + 1):
        ws.cell(row=row_idx, column=1, value=row_idx - header_row)
        ws.cell(row=row_idx, column=2, value=staff['staff_id'] or 'N/A')
        ws.cell(row=row_idx, column=3, value=staff['full_name'])
        ws.cell(row=row_idx, column=4, value=staff['department'] or 'N/A')
        ws.cell(row=row_idx, column=5, value=staff['position'] or 'N/A')
        ws.cell(row=row_idx, column=6, value=staff['basic_salary'] or 0)
        ws.cell(row=row_idx, column=7, value=staff['allowances'] or 0)
        ws.cell(row=row_idx, column=8, value=staff['deductions'] or 0)
        ws.cell(row=row_idx, column=9, value=staff['net_salary'] or 0)

        # Add to totals
        total_basic += staff['basic_salary'] or 0
        total_allowances += staff['allowances'] or 0
        total_deductions += staff['deductions'] or 0
        total_net += staff['net_salary'] or 0

        # Apply border to all cells
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).border = border

    # Add totals row
    total_row = len(staff_data) + header_row + 1
    ws.cell(row=total_row, column=5, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=total_basic).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=total_allowances).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=total_deductions).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=total_net).font = Font(bold=True)

    # Auto-adjust column widths
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.getvalue())
    filename = f'monthly_salary_report_{year}_{month or "all"}_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

    return response

def generate_staff_directory_report(school_id, format_type):
    """Generate comprehensive staff directory report"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    db = get_db()

    # Get comprehensive staff information
    staff_data = db.execute('''
        SELECT s.staff_id, s.full_name, s.first_name, s.last_name,
               s.date_of_birth, s.date_of_joining, s.department, s.position,
               s.gender, s.phone, s.email, s.shift_type, s.basic_salary,
               s.created_at
        FROM staff s
        WHERE s.school_id = ?
        ORDER BY s.department, s.full_name
    ''', (school_id,)).fetchall()

    # Common headers and rows for all formats
    headers = [
        'S.No', 'Staff ID', 'Full Name', 'Department', 'Position',
        'Gender', 'Phone', 'Email', 'Date of Joining', 'Date of Birth', 'Shift Type'
    ]
    rows = []
    for idx, staff in enumerate(staff_data, start=1):
        rows.append([
            idx,
            staff['staff_id'] or 'N/A',
            staff['full_name'],
            staff['department'] or 'N/A',
            staff['position'] or 'N/A',
            staff['gender'] or 'N/A',
            staff['phone'] or 'N/A',
            staff['email'] or 'N/A',
            staff['date_of_joining'] or 'N/A',
            staff['date_of_birth'] or 'N/A',
            staff['shift_type'] or 'General'
        ])

    # CSV export
    if format_type == 'csv':
        import io, csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Staff Directory Report', datetime.datetime.now().strftime('%Y-%m-%d')])
        writer.writerow([])
        writer.writerow(headers)
        writer.writerows(rows)
        response = make_response(output.getvalue())
        filename = f'staff_directory_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return response

    # PDF export
    if format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            import io

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=18)
            styles = getSampleStyleSheet()
            elements = []

            title = Paragraph(f"Staff Directory Report - {datetime.datetime.now().strftime('%Y-%m-%d')}", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 12))

            table_data = [headers] + rows
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            elements.append(table)

            doc.build(elements)
            pdf_value = buffer.getvalue()
            buffer.close()

            response = make_response(pdf_value)
            filename = f'staff_directory_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            response.headers['Content-Disposition'] = f'attachment; filename={filename}'
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            return response
        except ImportError:
            return jsonify({
                'success': False,
                'error': 'PDF generation requires reportlab. Install with: pip install reportlab'
            })
        except Exception as e:
            return jsonify({'success': False, 'error': f'PDF generation failed: {str(e)}'})

    # Default: Excel export
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Directory"

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16, color="2F5597")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Add title
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"Staff Directory Report - Generated on {datetime.datetime.now().strftime('%Y-%m-%d')}"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    # Headers row
    header_row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    row_idx = header_row + 1
    for r in rows:
        for col, value in enumerate(r, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
        row_idx += 1

    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to BytesIO
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.getvalue())
    filename = f'staff_directory_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

    return response

# Placeholder functions for other report types - can be expanded later
def generate_payroll_summary_report(school_id, year, month, format_type):
    """Generate enhanced Payroll Summary Report with detailed breakdown.
    
    Includes:
    1. Summary Section: Total payroll expenses, department breakdown, staff count
    2. Detailed Staff Records: All salary components, allowances, and deductions
    3. Attendance-based deductions for absent/late days
    4. Professional Excel formatting with totals and styling
    """
    import datetime, calendar, io
    from flask import request

    db = get_db()

    # Handle date range from request parameters (enhanced filtering)
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    department = request.args.get('department', '').strip()
    
    # Determine date range for the payroll report
    if from_date_str and to_date_str:
        try:
            start_date = datetime.datetime.strptime(from_date_str, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            # Fallback to year/month if date parsing fails
            if not isinstance(year, int) or year <= 0:
                year = datetime.date.today().year
            if not isinstance(month, int) or not (1 <= month <= 12):
                month = datetime.date.today().month
            start_date = datetime.date(year, month, 1)
            end_date = datetime.date(year, month, calendar.monthrange(year, month)[1])
    elif year and month:
        start_date = datetime.date(year, month, 1)
        end_date = datetime.date(year, month, calendar.monthrange(year, month)[1])
    else:
        # Default to current month
        today = datetime.date.today()
        start_date = datetime.date(today.year, today.month, 1)
        end_date = datetime.date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    
    start_str, end_str = start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')
    period_label = f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"

    # Build department filter clause
    dept_clause = ''
    staff_params = [school_id]
    if department:
        dept_clause = ' AND COALESCE(s.department, "") = ?'
        staff_params.append(department)

    # Get staff data with salary components
    staff_query = f"""
        SELECT 
            s.id as staff_db_id,
            s.staff_id,
            s.full_name,
            COALESCE(NULLIF(TRIM(s.department), ''), 'Unassigned') as department,
            COALESCE(NULLIF(TRIM(s.position), ''), 'Not Specified') as position,
            COALESCE(s.basic_salary, 0) as base_salary,
            COALESCE(s.hra, 0) as hra_allowance,
            COALESCE(s.transport_allowance, 0) as transport_allowance,
            COALESCE(s.other_allowances, 0) as other_allowances,
            COALESCE(s.pf_deduction, 0) as pf_deduction,
            COALESCE(s.esi_deduction, 0) as esi_deduction,
            COALESCE(s.professional_tax, 0) as professional_tax,
            COALESCE(s.other_deductions, 0) as other_deductions,
            s.date_of_joining
        FROM staff s
        WHERE s.school_id = ? AND s.is_active = 1 {dept_clause}
        ORDER BY s.department, s.full_name
    """
    
    staff_rows = db.execute(staff_query, tuple(staff_params)).fetchall()

    # Calculate working days in the period (excluding weekends and holidays)
    total_days = (end_date - start_date).days + 1
    working_days = 0
    current_dt = start_date
    weekend_days = 0
    
    # Get holidays in the period
    holidays_in_period = db.execute("""
        SELECT start_date, end_date, holiday_name 
        FROM holidays 
        WHERE school_id = ? AND is_active = 1
        AND NOT (end_date < ? OR start_date > ?)
    """, (school_id, start_str, end_str)).fetchall()
    
    # Create set of holiday dates
    holiday_dates = set()
    for holiday in holidays_in_period:
        h_start = datetime.datetime.strptime(holiday['start_date'], '%Y-%m-%d').date()
        h_end = datetime.datetime.strptime(holiday['end_date'], '%Y-%m-%d').date()
        curr_date = h_start
        while curr_date <= h_end:
            holiday_dates.add(curr_date)
            curr_date += datetime.timedelta(days=1)
    
    # Count working days (Mon-Fri, excluding holidays)
    while current_dt <= end_date:
        if current_dt.weekday() < 5:  # Monday = 0, Sunday = 6
            if current_dt not in holiday_dates:
                working_days += 1
        else:
            weekend_days += 1
        current_dt += datetime.timedelta(days=1)
    
    holiday_days = len(holiday_dates)

    # Process each staff member for detailed payroll calculation
    payroll_records = []
    total_payroll_expense = 0.0
    department_totals = {}
    
    for staff in staff_rows:
        staff_db_id = staff['staff_db_id']
        staff_id = staff['staff_id']
        
        # Get attendance data for deduction calculations
        attendance_data = db.execute("""
            SELECT 
                COUNT(*) as total_records,
                SUM(CASE WHEN status IN ('present', 'late') THEN 1 ELSE 0 END) as days_present,
                SUM(CASE WHEN status = 'absent' THEN 1 ELSE 0 END) as days_absent,
                SUM(CASE WHEN status = 'late' THEN 1 ELSE 0 END) as days_late
            FROM attendance 
            WHERE staff_id = ? AND date BETWEEN ? AND ?
        """, (staff_db_id, start_str, end_str)).fetchone()
        
        days_present = int(attendance_data['days_present'] or 0)
        days_absent = int(attendance_data['days_absent'] or 0)
        days_late = int(attendance_data['days_late'] or 0)
        
        # Calculate base salary components
        base_salary = float(staff['base_salary'])
        hra_allowance = float(staff['hra_allowance'])
        transport_allowance = float(staff['transport_allowance'])
        other_allowances = float(staff['other_allowances'])
        
        # Calculate total allowances
        total_allowances = hra_allowance + transport_allowance + other_allowances
        gross_pay = base_salary + total_allowances
        
        # Calculate attendance-based deductions
        if working_days > 0:
            daily_salary = base_salary / working_days
            absent_days_deduction = days_absent * daily_salary
            late_arrival_deduction = days_late * (daily_salary * 0.1)  # 10% penalty for late arrivals
        else:
            absent_days_deduction = 0
            late_arrival_deduction = 0
        
        # Static deductions from staff record
        pf_deduction = float(staff['pf_deduction'])
        esi_deduction = float(staff['esi_deduction'])
        professional_tax = float(staff['professional_tax'])
        other_deductions = float(staff['other_deductions'])
        
        # Total deductions
        total_deductions = (absent_days_deduction + late_arrival_deduction + 
                           pf_deduction + esi_deduction + professional_tax + other_deductions)
        
        # Net payroll (final amount to be paid)
        net_payroll = gross_pay - total_deductions
        
        # Add to department totals
        dept = staff['department']
        if dept not in department_totals:
            department_totals[dept] = {'count': 0, 'total': 0.0}
        department_totals[dept]['count'] += 1
        department_totals[dept]['total'] += net_payroll
        
        # Add to total payroll expense
        total_payroll_expense += net_payroll
        
        # Create detailed staff record
        payroll_records.append({
            'staff_id': staff_id,
            'staff_name': staff['full_name'],
            'department': staff['department'],
            'position': staff['position'],
            'base_salary': base_salary,
            'allowances': {
                'hra': hra_allowance,
                'transport': transport_allowance,
                'other': other_allowances,
                'total': total_allowances
            },
            'gross_pay': gross_pay,
            'deductions': {
                'absent_days': {'count': days_absent, 'amount': absent_days_deduction},
                'late_arrivals': {'count': days_late, 'amount': late_arrival_deduction},
                'pf': pf_deduction,
                'esi': esi_deduction,
                'professional_tax': professional_tax,
                'other': other_deductions,
                'total': total_deductions
            },
            'net_payroll': net_payroll,
            'attendance_summary': {
                'days_present': days_present,
                'days_absent': days_absent,
                'days_late': days_late,
                'working_days': working_days
            }
        })

    # Create summary section
    summary_data = {
        'total_payroll_expense': total_payroll_expense,
        'total_staff_count': len(payroll_records),
        'period_covered': period_label,
        'working_days': working_days,
        'department_breakdown': department_totals
    }

    # Generate Excel report with professional formatting
    if format_type == 'excel':
        return _generate_payroll_summary_excel(payroll_records, summary_data, start_date, end_date, department)
    
    # Return JSON for API requests
    return jsonify({
        'success': True,
        'summary': summary_data,
        'payroll_records': payroll_records
    })


def _generate_payroll_summary_excel(payroll_records, summary_data, start_date, end_date, department=None):
    """Generate professionally formatted Excel report for payroll summary"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payroll Summary Report'

    # Define professional styles
    title_font = Font(bold=True, size=16, color='FFFFFF')
    title_fill = PatternFill(start_color='2F4F4F', end_color='2F4F4F', fill_type='solid')
    
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    subheader_font = Font(bold=True, size=11, color='FFFFFF')
    subheader_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    
    summary_font = Font(bold=True, size=11)
    summary_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Row tracker
    current_row = 1

    # TITLE SECTION
    ws.merge_cells(f'A{current_row}:O{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = 'PAYROLL SUMMARY REPORT'
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    title_cell.border = border
    current_row += 2

    # SUMMARY SECTION
    ws.merge_cells(f'A{current_row}:O{current_row}')
    summary_title = ws[f'A{current_row}']
    summary_title.value = 'PAYROLL SUMMARY'
    summary_title.font = header_font
    summary_title.fill = header_fill
    summary_title.alignment = center_align
    summary_title.border = border
    current_row += 1

    # Summary details
    summary_details = [
        ('Period Covered:', summary_data['period_covered']),
        ('Total Staff Count:', str(summary_data['total_staff_count'])),
        ('Total Working Days:', str(summary_data['working_days'])),
        ('Total Payroll Expense:', f"₹{summary_data['total_payroll_expense']:,.2f}"),
        ('Department Filter:', department if department else 'All Departments')
    ]

    for label, value in summary_details:
        ws[f'A{current_row}'] = label
        ws[f'B{current_row}'] = value
        ws[f'A{current_row}'].font = summary_font
        ws[f'B{current_row}'].font = Font(size=11)
        ws[f'A{current_row}'].fill = summary_fill
        ws[f'B{current_row}'].fill = summary_fill
        ws[f'A{current_row}'].border = border
        ws[f'B{current_row}'].border = border
        current_row += 1

    current_row += 1

    # DEPARTMENT BREAKDOWN (if multiple departments)
    if len(summary_data['department_breakdown']) > 1:
        ws.merge_cells(f'A{current_row}:D{current_row}')
        dept_title = ws[f'A{current_row}']
        dept_title.value = 'DEPARTMENT BREAKDOWN'
        dept_title.font = subheader_font
        dept_title.fill = subheader_fill
        dept_title.alignment = center_align
        dept_title.border = border
        current_row += 1

        # Department breakdown headers
        dept_headers = ['Department', 'Staff Count', 'Total Amount', 'Average per Staff']
        for col, header in enumerate(dept_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        current_row += 1

        # Department breakdown data
        for dept, data in summary_data['department_breakdown'].items():
            avg_per_staff = data['total'] / data['count'] if data['count'] > 0 else 0
            row_data = [dept, data['count'], f"₹{data['total']:,.2f}", f"₹{avg_per_staff:,.2f}"]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                if col > 1:  # Right align numbers
                    cell.alignment = right_align
            current_row += 1

        current_row += 1

    # DETAILED STAFF RECORDS SECTION
    ws.merge_cells(f'A{current_row}:O{current_row}')
    details_title = ws[f'A{current_row}']
    details_title.value = 'DETAILED STAFF PAYROLL RECORDS'
    details_title.font = header_font
    details_title.fill = header_fill
    details_title.alignment = center_align
    details_title.border = border
    current_row += 1

    # Main table headers
    headers = [
        'Staff ID', 'Staff Name', 'Department', 'Position', 'Base Salary',
        'HRA', 'Transport', 'Other Allow.', 'Gross Pay',
        'Absent Days', 'Late Count', 'PF', 'Other Ded.', 'Total Ded.', 'Net Pay'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    current_row += 1

    # Data rows with alternating colors
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    total_gross_pay = 0.0
    total_deductions = 0.0
    total_net_pay = 0.0

    for idx, record in enumerate(payroll_records):
        row_fill = light_fill if idx % 2 == 0 else None
        
        # Prepare row data
        row_data = [
            record['staff_id'],
            record['staff_name'],
            record['department'],
            record['position'],
            f"₹{record['base_salary']:,.2f}",
            f"₹{record['allowances']['hra']:,.2f}",
            f"₹{record['allowances']['transport']:,.2f}",
            f"₹{record['allowances']['other']:,.2f}",
            f"₹{record['gross_pay']:,.2f}",
            f"{record['deductions']['absent_days']['count']} (₹{record['deductions']['absent_days']['amount']:,.0f})",
            f"{record['deductions']['late_arrivals']['count']} (₹{record['deductions']['late_arrivals']['amount']:,.0f})",
            f"₹{record['deductions']['pf']:,.2f}",
            f"₹{record['deductions']['other']:,.2f}",
            f"₹{record['deductions']['total']:,.2f}",
            f"₹{record['net_payroll']:,.2f}"
        ]
        
        # Write row data
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            if col >= 5:  # Right align numeric columns
                cell.alignment = right_align
            else:
                cell.alignment = center_align
        
        # Update totals
        total_gross_pay += record['gross_pay']
        total_deductions += record['deductions']['total']
        total_net_pay += record['net_payroll']
        
        current_row += 1

    # TOTALS ROW
    totals_data = [
        'TOTALS', '', '', '', '',
        '', '', '', f"₹{total_gross_pay:,.2f}",
        '', '', '', '', f"₹{total_deductions:,.2f}", f"₹{total_net_pay:,.2f}"
    ]
    
    for col, value in enumerate(totals_data, 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.font = Font(bold=True, size=12)
        cell.fill = summary_fill
        cell.border = border
        cell.alignment = right_align if col >= 5 else center_align

    # Auto-adjust column widths
    column_widths = [12, 25, 18, 18, 15, 12, 12, 12, 15, 15, 15, 12, 12, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Save workbook and return response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import make_response
    resp = make_response(output.getvalue())
    fname = f"payroll_summary_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    
    return resp

def generate_department_salary_report(school_id, year, department, format_type):
    """Generate comprehensive department-wise salary report with detailed breakdown"""
    import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    db = get_db()
    
    # Build query with optional department filtering
    query = """
        SELECT 
            staff_id, full_name, department, position,
            COALESCE(basic_salary, 0) as basic_salary,
            COALESCE(hra, 0) as hra,
            COALESCE(transport_allowance, 0) as transport_allowance,
            COALESCE(other_allowances, 0) as other_allowances,
            COALESCE(pf_deduction, 0) as pf_deduction,
            COALESCE(esi_deduction, 0) as esi_deduction,
            COALESCE(professional_tax, 0) as professional_tax,
            COALESCE(other_deductions, 0) as other_deductions
        FROM staff 
        WHERE school_id = ? AND is_active = 1
    """
    
    params = [school_id]
    
    # Add department filter if specified
    if department and department.strip():
        query += " AND department = ?"
        params.append(department.strip())
    
    query += " ORDER BY department, full_name"
    
    staff_data = db.execute(query, params).fetchall()
    
    # Group staff by department
    departments = {}
    for staff in staff_data:
        dept = staff['department'] or 'Unassigned'
        if dept not in departments:
            departments[dept] = []
        
        # Calculate totals
        total_allowances = (
            float(staff['hra']) + 
            float(staff['transport_allowance']) + 
            float(staff['other_allowances'])
        )
        total_deductions = (
            float(staff['pf_deduction']) + 
            float(staff['esi_deduction']) + 
            float(staff['professional_tax']) + 
            float(staff['other_deductions'])
        )
        base_salary = float(staff['basic_salary'])
        gross_pay = base_salary + total_allowances - total_deductions
        
        staff_record = {
            'staff_id': staff['staff_id'],
            'name': staff['full_name'],
            'position': staff['position'] or 'N/A',
            'base_salary': base_salary,
            'allowances': {
                'hra': float(staff['hra']),
                'transport': float(staff['transport_allowance']),
                'other': float(staff['other_allowances']),
                'total': total_allowances
            },
            'deductions': {
                'pf': float(staff['pf_deduction']),
                'esi': float(staff['esi_deduction']),
                'professional_tax': float(staff['professional_tax']),
                'other': float(staff['other_deductions']),
                'total': total_deductions
            },
            'gross_pay': gross_pay
        }
        departments[dept].append(staff_record)
    
    if format_type == 'json':
        return jsonify({
            'report_type': 'Department Wise Salary Report',
            'school_id': school_id,
            'year': year,
            'generated_at': datetime.datetime.now().isoformat(),
            'departments': departments,
            'summary': {
                'total_departments': len(departments),
                'total_staff': sum(len(dept_staff) for dept_staff in departments.values()),
                'total_gross_pay': sum(
                    sum(staff['gross_pay'] for staff in dept_staff) 
                    for dept_staff in departments.values()
                )
            }
        })
    
    elif format_type == 'excel':
        return _generate_department_salary_excel(departments, school_id, year)
    
    else:
        return jsonify({'error': 'Unsupported format type'}), 400

def _generate_department_salary_excel(departments, school_id, year):
    """Generate professionally formatted Excel file for department-wise salary report"""
    import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import make_response
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Department Wise Salary Report"
    
    # Define styles
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    
    header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    dept_header_font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
    dept_header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    
    data_font = Font(name='Arial', size=11)
    odd_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    current_row = 1
    
    # Title section
    ws.merge_cells(f'A{current_row}:I{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = f"DEPARTMENT WISE SALARY REPORT - {year}"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    title_cell.border = thin_border
    current_row += 2
    
    # Report info
    ws[f'A{current_row}'] = f"Generated on: {datetime.datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws[f'A{current_row}'].font = Font(name='Arial', size=10, italic=True)
    current_row += 1
    
    ws[f'A{current_row}'] = f"School ID: {school_id}"
    ws[f'A{current_row}'].font = Font(name='Arial', size=10, italic=True)
    current_row += 2
    
    # Calculate totals for summary
    total_staff = sum(len(dept_staff) for dept_staff in departments.values())
    total_gross_pay = sum(
        sum(staff['gross_pay'] for staff in dept_staff) 
        for dept_staff in departments.values()
    )
    
    # Summary section
    ws[f'A{current_row}'] = "SUMMARY"
    ws[f'A{current_row}'].font = Font(name='Arial', size=12, bold=True)
    current_row += 1
    
    ws[f'A{current_row}'] = f"Total Departments: {len(departments)}"
    ws[f'C{current_row}'] = f"Total Staff: {total_staff}"
    ws[f'E{current_row}'] = f"Total Gross Pay: ₹{total_gross_pay:,.2f}"
    
    for cell in [ws[f'A{current_row}'], ws[f'C{current_row}'], ws[f'E{current_row}']]:
        cell.font = Font(name='Arial', size=10, bold=True)
    current_row += 2
    
    # Main table headers
    headers = [
        'Staff ID', 'Staff Name', 'Position/Job Title', 'Base Salary', 
        'HRA', 'Transport', 'Other Allow.', 'PF Deduction', 'ESI Deduction', 
        'Prof. Tax', 'Other Ded.', 'Total Deductions', 'Gross Pay'
    ]
    
    # Process each department
    for dept_name in sorted(departments.keys()):
        dept_staff = departments[dept_name]
        
        # Department header
        ws.merge_cells(f'A{current_row}:M{current_row}')
        dept_header_cell = ws[f'A{current_row}']
        dept_header_cell.value = f"{dept_name.upper()} DEPARTMENT ({len(dept_staff)} Staff)"
        dept_header_cell.font = dept_header_font
        dept_header_cell.fill = dept_header_fill
        dept_header_cell.alignment = center_alignment
        dept_header_cell.border = thin_border
        current_row += 1
        
        # Column headers for this department
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # Department staff data
        dept_total_gross = 0
        for i, staff in enumerate(dept_staff):
            row_data = [
                staff['staff_id'],
                staff['name'],
                staff['position'],
                f"₹{staff['base_salary']:,.2f}",
                f"₹{staff['allowances']['hra']:,.2f}",
                f"₹{staff['allowances']['transport']:,.2f}",
                f"₹{staff['allowances']['other']:,.2f}",
                f"₹{staff['deductions']['pf']:,.2f}",
                f"₹{staff['deductions']['esi']:,.2f}",
                f"₹{staff['deductions']['professional_tax']:,.2f}",
                f"₹{staff['deductions']['other']:,.2f}",
                f"₹{staff['deductions']['total']:,.2f}",
                f"₹{staff['gross_pay']:,.2f}"
            ]
            
            dept_total_gross += staff['gross_pay']
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.font = data_font
                cell.border = thin_border
                
                # Alternate row coloring
                if i % 2 == 1:
                    cell.fill = odd_row_fill
                    
                # Alignment
                if col <= 3:  # Text columns
                    cell.alignment = left_alignment
                else:  # Number columns
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            current_row += 1
        
        # Department total row
        ws.merge_cells(f'A{current_row}:L{current_row}')
        total_cell = ws[f'A{current_row}']
        total_cell.value = f"DEPARTMENT TOTAL"
        total_cell.font = Font(name='Arial', size=11, bold=True)
        total_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        total_cell.alignment = center_alignment
        total_cell.border = thin_border
        
        gross_total_cell = ws[f'M{current_row}']
        gross_total_cell.value = f"₹{dept_total_gross:,.2f}"
        gross_total_cell.font = Font(name='Arial', size=11, bold=True)
        gross_total_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        gross_total_cell.alignment = Alignment(horizontal='right', vertical='center')
        gross_total_cell.border = thin_border
        
        current_row += 2  # Space between departments
    
    # Adjust column widths
    column_widths = [12, 25, 20, 15, 12, 12, 12, 12, 12, 12, 12, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=department_salary_report_{year}.xlsx'
    
    return response

def generate_department_analysis_report(school_id, year=None, month=None, format_type='excel'):
    """Generate comprehensive Department Report with multi-format export.
    Includes: total staff per department, gender + position breakdown,
    average tenure, salary min/avg/max, and attendance stats for selected month.
    """
    import datetime, calendar, io

    db = get_db()

    # Resolve period (attendance window + tenure as of end of month)
    today = datetime.date.today()
    if not isinstance(year, int) or year <= 0:
        year = today.year
    if not isinstance(month, int) or not (1 <= month <= 12):
        month = today.month
    start_date = datetime.date(year, month, 1)
    end_date = datetime.date(year, month, calendar.monthrange(year, month)[1])
    start_str, end_str = start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')

    # Departments
    departments = [r['dept'] for r in db.execute(
        """
        SELECT DISTINCT department AS dept
        FROM staff
        WHERE school_id = ? AND COALESCE(department, '') <> ''
        ORDER BY department
        """, (school_id,)
    ).fetchall()]

    # Totals
    total_map = {r['department']: r['total'] for r in db.execute(
        """
        SELECT department, COUNT(*) AS total
        FROM staff
        WHERE school_id = ? AND COALESCE(department, '') <> ''
        GROUP BY department
        """, (school_id,)
    ).fetchall()}

    # Gender breakdown
    gender_map = {}
    for g in db.execute(
        """
        SELECT department, COALESCE(gender, 'Other') AS gender, COUNT(*) AS cnt
        FROM staff
        WHERE school_id = ? AND COALESCE(department, '') <> ''
        GROUP BY department, COALESCE(gender, 'Other')
        """, (school_id,)
    ).fetchall():
        dept = g['department']
        gender_map.setdefault(dept, {'Male': 0, 'Female': 0, 'Other': 0})
        key = g['gender'] if g['gender'] in ('Male', 'Female', 'Other') else 'Other'
        gender_map[dept][key] += g['cnt']

    # Position breakdown rows (for sheet)
    position_rows = [
        {'department': p['department'], 'position': p['position'] or 'Unspecified', 'count': p['cnt']}
        for p in db.execute(
            """
            SELECT department, COALESCE(position, 'Unspecified') AS position, COUNT(*) AS cnt
            FROM staff
            WHERE school_id = ? AND COALESCE(department, '') <> ''
            GROUP BY department, COALESCE(position, 'Unspecified')
            ORDER BY department, position
            """, (school_id,)
        ).fetchall()
    ]

    # Salary stats (convert sqlite3.Row -> plain dict)
    salary_rows = db.execute(
        """
        SELECT department,
               MIN(COALESCE(basic_salary, 0)) AS min_salary,
               AVG(COALESCE(basic_salary, 0)) AS avg_salary,
               MAX(COALESCE(basic_salary, 0)) AS max_salary
        FROM staff
        WHERE school_id = ? AND COALESCE(department, '') <> ''
        GROUP BY department
        """, (school_id,)
    ).fetchall()
    salary_map = {}
    for s in salary_rows:
        dept = s['department']
        salary_map[dept] = {
            'min_salary': float(s['min_salary'] or 0),
            'avg_salary': float(s['avg_salary'] or 0),
            'max_salary': float(s['max_salary'] or 0),
        }

    # Tenure as of end_date
    from collections import defaultdict
    tenure_sum_days, tenure_count = defaultdict(int), defaultdict(int)
    for r in db.execute(
        """
        SELECT department, date_of_joining
        FROM staff
        WHERE school_id = ? AND COALESCE(department, '') <> ''
        """, (school_id,)
    ).fetchall():
        doj = r['date_of_joining']
        if doj:
            try:
                doj_dt = datetime.datetime.strptime(doj, '%Y-%m-%d').date()
                days = (end_date - doj_dt).days
                if days >= 0:
                    tenure_sum_days[r['department']] += days
                    tenure_count[r['department']] += 1
            except Exception:
                pass
    avg_tenure_years = {d: (tenure_sum_days[d]/tenure_count[d]/365.0) if tenure_count[d] else 0.0 for d in departments}

    # Attendance for selected month
    # Build attendance_map as plain dicts to avoid sqlite3.Row .get errors
    attendance_rows = db.execute(
        """
        SELECT s.department AS department,
               SUM(CASE WHEN a.status = 'present' THEN 1 ELSE 0 END) AS present_count,
               SUM(CASE WHEN a.status = 'absent' THEN 1 ELSE 0 END) AS absent_count,
               SUM(CASE WHEN a.status = 'late' THEN 1 ELSE 0 END) AS late_count,
               SUM(CASE WHEN a.status = 'leave' THEN 1 ELSE 0 END) AS leave_count
        FROM staff s
        LEFT JOIN attendance a ON a.staff_id = s.id AND a.date BETWEEN ? AND ?
        WHERE s.school_id = ? AND COALESCE(s.department, '') <> ''
        GROUP BY s.department
        ORDER BY s.department
        """, (start_str, end_str, school_id)
    ).fetchall()
    attendance_map = {}
    for a in attendance_rows:
        attendance_map[a['department']] = {
            'present_count': int(a['present_count'] or 0),
            'absent_count': int(a['absent_count'] or 0),
            'late_count': int(a['late_count'] or 0),
            'leave_count': int(a['leave_count'] or 0),
        }

    # Summary rows
    summary = []
    for dept in departments:
        g = gender_map.get(dept, {'Male': 0, 'Female': 0, 'Other': 0})
        sstat = salary_map.get(dept, {'min_salary': 0, 'avg_salary': 0, 'max_salary': 0})
        att = attendance_map.get(dept, {'present_count': 0, 'absent_count': 0, 'late_count': 0, 'leave_count': 0})
        present = int(att.get('present_count', 0) or 0)
        absent = int(att.get('absent_count', 0) or 0)
        late = int(att.get('late_count', 0) or 0)
        leave = int(att.get('leave_count', 0) or 0)
        denom = present + absent + late + leave
        present_rate = (present/denom*100.0) if denom else 0.0
        summary.append({
            'department': dept,
            'total_staff': int(total_map.get(dept, 0)),
            'male': int(g.get('Male', 0) or 0),
            'female': int(g.get('Female', 0) or 0),
            'other': int(g.get('Other', 0) or 0),
            'avg_tenure_years': round(avg_tenure_years.get(dept, 0.0), 2),
            'min_salary': float(sstat.get('min_salary', 0) or 0),
            'avg_salary': float(sstat.get('avg_salary', 0) or 0),
            'max_salary': float(sstat.get('max_salary', 0) or 0),
            'present': present, 'absent': absent, 'late': late, 'leave': leave,
            'present_rate': round(present_rate, 2)
        })

    # CSV
    if format_type == 'csv':
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([f"Department Report - {calendar.month_name[month]} {year}"])
        writer.writerow([])
        headers = ['Department','Total Staff','Male','Female','Other','Avg Tenure (yrs)',
                   'Min Salary','Avg Salary','Max Salary','Present','Absent','Late','Leave','Present Rate (%)']
        writer.writerow(headers)
        for row in summary:
            writer.writerow([
                row['department'], row['total_staff'], row['male'], row['female'], row['other'], row['avg_tenure_years'],
                f"{row['min_salary']:.2f}", f"{row['avg_salary']:.2f}", f"{row['max_salary']:.2f}",
                row['present'], row['absent'], row['late'], row['leave'], row['present_rate']
            ])
        resp = make_response(output.getvalue())
        fname = f"department_report_{year}_{str(month).zfill(2)}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
        resp.headers['Content-Type'] = 'text/csv'
        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return resp

    # PDF
    if format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=18)
            styles = getSampleStyleSheet()
            elements = []
            elements.append(Paragraph(f"Department Report - {calendar.month_name[month]} {year}", styles['Title']))
            elements.append(Spacer(1, 12))
            headers = ['Department','Total','Male','Female','Other','AvgTenure','MinSal','AvgSal','MaxSal','Present','Absent','Late','Leave','%Present']
            data = [headers]
            for r in summary:
                data.append([
                    r['department'], r['total_staff'], r['male'], r['female'], r['other'],
                    f"{r['avg_tenure_years']:.2f}", f"{r['min_salary']:.0f}", f"{r['avg_salary']:.0f}", f"{r['max_salary']:.0f}",
                    r['present'], r['absent'], r['late'], r['leave'], f"{r['present_rate']:.1f}%"
                ])
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            elements.append(table)
            doc.build(elements)
            pdf = buffer.getvalue(); buffer.close()
            resp = make_response(pdf)
            fname = f"department_report_{year}_{str(month).zfill(2)}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
            resp.headers['Content-Type'] = 'application/pdf'
            resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            return resp
        except ImportError:
            return jsonify({'success': False, 'error': 'PDF generation requires reportlab. Install with: pip install reportlab'})
        except Exception as e:
            return jsonify({'success': False, 'error': f'PDF generation failed: {str(e)}'})

    # Default: Excel with Summary, Positions, Staff Details
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Summary'

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['Department','Total Staff','Male','Female','Other','Avg Tenure (yrs)',
               'Min Salary','Avg Salary','Max Salary','Present','Absent','Late','Leave','Present Rate (%)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h); c.font = header_font; c.fill = header_fill; c.border = border; c.alignment = Alignment(horizontal='center')

    r = 2
    for row in summary:
        vals = [row['department'], row['total_staff'], row['male'], row['female'], row['other'], row['avg_tenure_years'],
                row['min_salary'], row['avg_salary'], row['max_salary'], row['present'], row['absent'], row['late'], row['leave'], row['present_rate']]
        for cidx, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=cidx, value=val); cell.border = border
        r += 1
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Positions sheet
    ws2 = wb.create_sheet(title='Positions')
    pos_headers = ['Department','Position','Count']
    for col, h in enumerate(pos_headers, 1):
        c = ws2.cell(row=1, column=col, value=h); c.font = header_font; c.fill = header_fill; c.border = border; c.alignment = Alignment(horizontal='center')
    r = 2
    for pr in position_rows:
        ws2.cell(row=r, column=1, value=pr['department'])
        ws2.cell(row=r, column=2, value=pr['position'])
        ws2.cell(row=r, column=3, value=pr['count'])
        for c in range(1, 4):
            ws2.cell(row=r, column=c).border = border
        r += 1
    for col in range(1, 4):
        ws2.column_dimensions[get_column_letter(col)].width = 24

    # Staff Details sheet
    staff_details = db.execute(
        """
        SELECT department, staff_id, full_name, COALESCE(position,'Unspecified') AS position,
               COALESCE(gender,'Other') AS gender, date_of_joining, COALESCE(basic_salary,0) AS basic_salary
        FROM staff
        WHERE school_id = ? AND COALESCE(department, '') <> ''
        ORDER BY department, full_name
        """, (school_id,)
    ).fetchall()
    ws3 = wb.create_sheet(title='Staff Details')
    sd_headers = ['Department','Staff ID','Full Name','Position','Gender','Date of Joining','Basic Salary']
    for col, h in enumerate(sd_headers, 1):
        c = ws3.cell(row=1, column=col, value=h); c.font = header_font; c.fill = header_fill; c.border = border; c.alignment = Alignment(horizontal='center')
    r = 2
    for s in staff_details:
        vals = [s['department'], s['staff_id'], s['full_name'], s['position'], s['gender'], s['date_of_joining'], float(s['basic_salary'] or 0)]
        for cidx, val in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=cidx, value=val); cell.border = border
        r += 1
    widths = [18, 14, 26, 18, 12, 16, 14]
    for col, w in enumerate(widths, 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    output = BytesIO(); wb.save(output); output.seek(0)
    resp = make_response(output.getvalue())
    fname = f"department_report_{year}_{str(month).zfill(2)}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp

def generate_performance_report(school_id, year=None, month=None, department=None, format_type='excel'):
    """Generate enhanced Staff Performance Metrics and Evaluations Report.
    
    Includes the specific fields requested by user:
    - Staff ID, Name, Department, Position
    - Days Present (count)
    - Days Absent (count) 
    - Days on OD (Official Duty) applied
    - Days on Leave applied
    - Days with Permission applied
    
    Supports date range filtering and Excel export.
    """
    import datetime, calendar, io
    from flask import request

    db = get_db()

    # Handle date range from request parameters
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    # Determine date range for the report
    if from_date_str and to_date_str:
        try:
            start_date = datetime.datetime.strptime(from_date_str, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            # Fallback to month/year if date parsing fails
            today = datetime.date.today()
            if not isinstance(year, int) or year <= 0:
                year = today.year
            if not isinstance(month, int) or not (1 <= month <= 12):
                month = today.month
            start_date = datetime.date(year, month, 1)
            end_date = datetime.date(year, month, calendar.monthrange(year, month)[1])
    elif year and month:
        start_date = datetime.date(year, month, 1)
        end_date = datetime.date(year, month, calendar.monthrange(year, month)[1])
    else:
        # Default to current month
        today = datetime.date.today()
        start_date = datetime.date(today.year, today.month, 1)
        end_date = datetime.date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    
    start_str, end_str = start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')
    period_label = f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"

    # Build department filter clause
    dept_clause = ''
    params_base = [start_str, end_str, school_id]
    if department and str(department).strip():
        dept_clause = ' AND COALESCE(s.department, "") = ? '
        params_base.append(department)

    # Main query to get staff performance data with all required fields
    staff_performance_query = f"""
        SELECT 
            s.id as staff_db_id,
            s.staff_id,
            s.full_name,
            COALESCE(NULLIF(TRIM(s.department), ''), 'Unassigned') as department,
            COALESCE(NULLIF(TRIM(s.position), ''), 'Not Specified') as position,
            s.date_of_joining
            
        FROM staff s
        WHERE s.school_id = ? AND s.is_active = 1 {dept_clause[24:] if dept_clause else ''}
        ORDER BY s.department, s.full_name
    """

    # Adjust parameters for staff query (remove date parameters)
    staff_params = [school_id]
    if department and str(department).strip():
        staff_params.append(department)
    
    staff_rows = db.execute(staff_performance_query, tuple(staff_params)).fetchall()

    # Now get attendance data for each staff member separately for accuracy
    performance_rows = []
    
    for staff in staff_rows:
        staff_id = staff['staff_id']
        staff_db_id = staff['staff_db_id']
        
        # Get attendance data for this specific staff member
        attendance_data = db.execute("""
            SELECT 
                COUNT(*) as total_records,
                SUM(CASE WHEN status IN ('present', 'late') THEN 1 ELSE 0 END) as days_present,
                SUM(CASE WHEN status = 'absent' THEN 1 ELSE 0 END) as days_absent,
                SUM(CASE WHEN status = 'late' THEN 1 ELSE 0 END) as days_late,
                SUM(CASE WHEN status = 'present' THEN 1 ELSE 0 END) as days_present_only,
                SUM(CASE WHEN status = 'early_departure' THEN 1 ELSE 0 END) as days_early_departure
            FROM attendance 
            WHERE staff_id = ? AND date BETWEEN ? AND ?
        """, (staff_db_id, start_str, end_str)).fetchone()
        
        days_present = int(attendance_data['days_present'] or 0)
        days_absent = int(attendance_data['days_absent'] or 0)
        days_late = int(attendance_data['days_late'] or 0)
        days_present_only = int(attendance_data['days_present_only'] or 0)
        total_attendance_records = int(attendance_data['total_records'] or 0)
        
        performance_rows.append({
            'staff_id': staff['staff_id'],
            'staff_db_id': staff_db_id,
            'staff_name': staff['full_name'],
            'department': staff['department'],
            'position': staff['position'],
            'days_present': days_present,
            'days_absent': days_absent,
            'days_late': days_late,
            'days_present_only': days_present_only,
            'total_attendance_records': total_attendance_records
        })

    # Get Leave Applications data for all staff
    leave_query = f"""
        SELECT 
            s.staff_id,
            COALESCE(COUNT(CASE WHEN l.status = 'approved' THEN l.id END), 0) as approved_leave_count,
            COALESCE(SUM(CASE WHEN l.status = 'approved' 
                        THEN julianday(l.end_date) - julianday(l.start_date) + 1 
                        ELSE 0 END), 0) as days_on_leave
        FROM staff s
        LEFT JOIN leave_applications l ON s.id = l.staff_id 
            AND NOT (l.end_date < ? OR l.start_date > ?)
        WHERE s.school_id = ? AND s.is_active = 1 {dept_clause[24:] if dept_clause else ''}
        GROUP BY s.id, s.staff_id
    """
    
    leave_params = [start_str, end_str, school_id]
    if department and str(department).strip():
        leave_params.append(department)
    
    leave_data = db.execute(leave_query, tuple(leave_params)).fetchall()
    leave_map = {row['staff_id']: {
        'applications': int(row['approved_leave_count']), 
        'days': float(row['days_on_leave'])
    } for row in leave_data}

    # Get Official Duty (OD) Applications data
    od_query = f"""
        SELECT 
            s.staff_id,
            COALESCE(COUNT(CASE WHEN od.status = 'approved' THEN od.id END), 0) as approved_od_count,
            COALESCE(SUM(CASE WHEN od.status = 'approved' 
                        THEN julianday(od.end_date) - julianday(od.start_date) + 1 
                        ELSE 0 END), 0) as days_on_od
        FROM staff s
        LEFT JOIN on_duty_applications od ON s.id = od.staff_id 
            AND NOT (od.end_date < ? OR od.start_date > ?)
        WHERE s.school_id = ? AND s.is_active = 1 {dept_clause[24:] if dept_clause else ''}
        GROUP BY s.id, s.staff_id
    """
    
    od_params = [start_str, end_str, school_id]
    if department and str(department).strip():
        od_params.append(department)
    
    od_data = db.execute(od_query, tuple(od_params)).fetchall()
    od_map = {row['staff_id']: {
        'applications': int(row['approved_od_count']), 
        'days': float(row['days_on_od'])
    } for row in od_data}

    # Get Permission Applications data
    permission_query = f"""
        SELECT 
            s.staff_id,
            COALESCE(COUNT(CASE WHEN p.status = 'approved' THEN p.id END), 0) as approved_permission_count,
            COALESCE(SUM(CASE WHEN p.status = 'approved' THEN p.duration_hours ELSE 0 END), 0) as total_permission_hours
        FROM staff s
        LEFT JOIN permission_applications p ON s.id = p.staff_id 
            AND p.permission_date BETWEEN ? AND ?
        WHERE s.school_id = ? AND s.is_active = 1 {dept_clause[24:] if dept_clause else ''}
        GROUP BY s.id, s.staff_id
    """
    
    permission_params = [start_str, end_str, school_id]
    if department and str(department).strip():
        permission_params.append(department)
    
    permission_data = db.execute(permission_query, tuple(permission_params)).fetchall()
    permission_map = {row['staff_id']: {
        'applications': int(row['approved_permission_count']), 
        'hours': float(row['total_permission_hours'])
    } for row in permission_data}

    # Calculate working days in the period (excluding weekends and holidays)
    total_days = (end_date - start_date).days + 1
    working_days = 0
    current_dt = start_date
    weekend_days = 0
    
    # Get holidays in the period
    holidays_in_period = db.execute("""
        SELECT start_date, end_date, holiday_name 
        FROM holidays 
        WHERE school_id = ? AND is_active = 1
        AND NOT (end_date < ? OR start_date > ?)
    """, (school_id, start_str, end_str)).fetchall()
    
    # Create set of holiday dates
    holiday_dates = set()
    for holiday in holidays_in_period:
        h_start = datetime.datetime.strptime(holiday['start_date'], '%Y-%m-%d').date()
        h_end = datetime.datetime.strptime(holiday['end_date'], '%Y-%m-%d').date()
        curr_date = h_start
        while curr_date <= h_end:
            holiday_dates.add(curr_date)
            curr_date += datetime.timedelta(days=1)
    
    # Count working days (Mon-Fri, excluding holidays)
    while current_dt <= end_date:
        if current_dt.weekday() < 5:  # Monday = 0, Sunday = 6
            if current_dt not in holiday_dates:
                working_days += 1
        else:
            weekend_days += 1
        current_dt += datetime.timedelta(days=1)
    
    holiday_days = len(holiday_dates)

    # Combine all data into final performance report rows
    final_performance_rows = []
    for perf in performance_rows:
        staff_id = perf['staff_id']
        
        # Get leave data for this staff
        leave_info = leave_map.get(staff_id, {'applications': 0, 'days': 0})
        
        # Get OD data for this staff
        od_info = od_map.get(staff_id, {'applications': 0, 'days': 0})
        
        # Get permission data for this staff
        permission_info = permission_map.get(staff_id, {'applications': 0, 'hours': 0})
        permission_days = round(float(permission_info['hours']) / 8, 1) if permission_info['hours'] else 0
        
        final_performance_rows.append({
            'staff_id': perf['staff_id'],
            'staff_name': perf['staff_name'],
            'department': perf['department'],
            'position': perf['position'],
            'days_present': perf['days_present'],
            'days_absent': perf['days_absent'],
            'days_on_od_applied': int(od_info['days']),
            'days_on_leave_applied': int(leave_info['days']),
            'days_with_permission_applied': permission_days,
            'total_working_days': working_days,
            'total_attendance_records': perf['total_attendance_records'],
            'days_late': perf['days_late'],
            'days_present_only': perf['days_present_only']
        })

    # CSV export
    if format_type == 'csv':
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([f"Staff Performance Metrics Report"])
        writer.writerow([f"Period: {period_label}"])
        if department:
            writer.writerow([f"Department: {department}"])
        writer.writerow([])
        writer.writerow(['Staff ID', 'Staff Name', 'Department', 'Position', 'Days Present', 'Days Absent', 'Days Late', 'Days on OD Applied', 'Days on Leave Applied', 'Days with Permission Applied', 'Total Working Days', 'Attendance Records'])
        
        for row in final_performance_rows:
            writer.writerow([
                row['staff_id'], row['staff_name'], row['department'], row['position'],
                row['days_present'], row['days_absent'], row['days_late'], row['days_on_od_applied'],
                row['days_on_leave_applied'], row['days_with_permission_applied'],
                row['total_working_days'], row['total_attendance_records']
            ])
        
        resp = make_response(output.getvalue())
        fname = f"staff_performance_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
        resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
        resp.headers['Content-Type'] = 'text/csv'
        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return resp

    # PDF export
    if format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=18)
            styles = getSampleStyleSheet()
            elements = []
            
            elements.append(Paragraph(f"Staff Performance Metrics Report", styles['Title']))
            elements.append(Paragraph(f"Period: {period_label}", styles['Normal']))
            if department:
                elements.append(Paragraph(f"Department: {department}", styles['Normal']))
            elements.append(Spacer(1, 12))
            
            # Create table data
            headers = ['Staff ID', 'Staff Name', 'Department', 'Position', 'Days Present', 'Days Absent', 'Days Late', 'Days on OD', 'Days on Leave', 'Permission (Days)', 'Working Days', 'Records']
            table_data = [headers]
            
            for row in final_performance_rows:
                table_data.append([
                    row['staff_id'], row['staff_name'], row['department'], row['position'],
                    str(row['days_present']), str(row['days_absent']), str(row['days_late']),
                    str(row['days_on_od_applied']), str(row['days_on_leave_applied']), 
                    str(row['days_with_permission_applied']), str(row['total_working_days']),
                    str(row['total_attendance_records'])
                ])
            
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            elements.append(table)
            
            doc.build(elements)
            pdf = buffer.getvalue()
            buffer.close()
            
            resp = make_response(pdf)
            fname = f"staff_performance_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
            resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
            resp.headers['Content-Type'] = 'application/pdf'
            resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            return resp
        except ImportError:
            return jsonify({'success': False, 'error': 'PDF generation requires reportlab. Install with: pip install reportlab'})
        except Exception as e:
            return jsonify({'success': False, 'error': f'PDF generation failed: {str(e)}'})

    # Default: Excel export  
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Staff Performance Report'

    # Styling
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Report title and metadata
    ws.cell(row=1, column=1, value=f'Staff Performance Metrics Report').font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f'Period: {period_label}').font = Font(bold=True, size=11)
    if department:
        ws.cell(row=3, column=1, value=f'Department: {department}').font = Font(bold=True, size=11)
        header_row = 5
    else:
        header_row = 4

    # Column headers
    headers = ['Staff ID', 'Staff Name', 'Department', 'Position', 'Days Present', 'Days Absent', 'Days Late', 'Days on OD', 'Days on Leave', 'Permission (Days)', 'Total Working Days', 'Attendance Records']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    data_start_row = header_row + 1
    for row_idx, data in enumerate(final_performance_rows, data_start_row):
        values = [
            data['staff_id'], data['staff_name'], data['department'], data['position'],
            data['days_present'], data['days_absent'], data['days_late'],
            data['days_on_od_applied'], data['days_on_leave_applied'], 
            data['days_with_permission_applied'], data['total_working_days'],
            data['total_attendance_records']
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if col >= 5:  # Numeric columns - center align
                cell.alignment = Alignment(horizontal='center')

    # Auto-adjust column widths
    column_widths = [12, 25, 18, 18, 12, 12, 12, 12, 14, 16, 16, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Save workbook and return response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    resp = make_response(output.getvalue())
    fname = f"staff_performance_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp


def generate_daily_attendance_report(school_id, date_str=None, department=None, format_type='excel'):
    """Generate comprehensive Daily Attendance Report (multi-format).
    Includes per-staff records for selected date, department summary, and overall stats.
    """
    import datetime, io

    db = get_db()

    # Resolve date
    try:
        if not date_str:
            date_obj = datetime.date.today()
        else:
            date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        date_obj = datetime.date.today()
    date_str = date_obj.strftime('%Y-%m-%d')

    # Check attendance table columns to avoid SQL errors on older schemas
    try:
        cols = {row['name'] for row in db.execute("PRAGMA table_info(attendance)").fetchall()}
    except Exception:
        cols = set()
    has_work_hours = 'work_hours' in cols
    has_overtime = 'overtime_hours' in cols
    has_late_min = 'late_duration_minutes' in cols
    has_early_min = 'early_departure_minutes' in cols

    # Optional department filter
    dept_clause = ''
    params = [date_str, school_id]
    if department and str(department).strip():
        dept_clause = ' AND COALESCE(s.department, "") = ? '
        params.append(department)

    # Pull per-staff daily records
    select_fields = [
        "s.staff_id", "s.full_name", "s.department", "COALESCE(s.position,'') AS position",
        "a.time_in", "a.time_out", "a.status",
    ]
    if has_work_hours:
        select_fields.append("a.work_hours")
    else:
        select_fields.append("0 AS work_hours")
    if has_overtime:
        select_fields.append("a.overtime_hours")
    else:
        select_fields.append("0 AS overtime_hours")
    if has_late_min:
        select_fields.append("a.late_duration_minutes")
    else:
        select_fields.append("0 AS late_duration_minutes")
    if has_early_min:
        select_fields.append("a.early_departure_minutes")
    else:
        select_fields.append("0 AS early_departure_minutes")

    query = f"""
        SELECT {', '.join(select_fields)}
        FROM staff s
        LEFT JOIN attendance a ON a.staff_id = s.id AND a.date = ?
        WHERE s.school_id = ? {dept_clause}
        ORDER BY s.department, s.full_name
    """
    rows = db.execute(query, tuple(params)).fetchall()

    # Normalize rows and build summaries
    def fmt_time(t):
        return t if (t and isinstance(t, str)) else (t.strftime('%H:%M') if t else '')

    staff_records = []
    dept_summary = {}
    overall = {k: 0 for k in ['present','absent','late','leave','on_duty','holiday']}
    total_staff = 0

    for r in rows:
        total_staff += 1
        status = (r['status'] or '').lower() if r['status'] else 'absent'
        if status not in overall:
            # Map unknown to absent by default
            status = 'absent'
        overall[status] += 1

        dept = r['department'] or 'Unassigned'
        ds = dept_summary.setdefault(dept, {k: 0 for k in ['present','absent','late','leave','on_duty','holiday']})
        ds[status] += 1

        staff_records.append({
            'staff_id': r['staff_id'],
            'full_name': r['full_name'],
            'department': dept,
            'position': r['position'],
            'time_in': r['time_in'] or '',
            'time_out': r['time_out'] or '',
            'status': status.capitalize(),
            'late_minutes': int(r['late_duration_minutes'] or 0),
            'early_departure_minutes': int(r['early_departure_minutes'] or 0),
            'work_hours': float(r['work_hours'] or 0),
            'overtime_hours': float(r['overtime_hours'] or 0),
        })

    # Compute department rows with totals and present rate
    dept_rows = []
    for dept, ds in sorted(dept_summary.items()):
        total = sum(ds.values())
        present = ds['present'] + ds['on_duty'] + ds['late']  # treat on_duty/late as attended
        present_rate = (present / total * 100.0) if total else 0.0
        dept_rows.append({
            'department': dept,
            'total': total,
            **ds,
            'present_rate': round(present_rate, 2)
        })

    date_label = date_obj.strftime('%Y-%m-%d')

    # CSV export
    if format_type == 'csv':
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([f"Daily Attendance Report - {date_label}"])
        writer.writerow([])
        # Overall summary
        writer.writerow(['Total Staff','Present','Late','Leave','On Duty','Holiday','Absent'])
        writer.writerow([total_staff, overall['present'], overall['late'], overall['leave'], overall['on_duty'], overall['holiday'], overall['absent']])
        writer.writerow([])
        # Department summary
        writer.writerow(['Department','Total','Present','Late','Leave','On Duty','Holiday','Absent','Present Rate (%)'])
        for d in dept_rows:
            writer.writerow([d['department'], d['total'], d['present'], d['late'], d['leave'], d['on_duty'], d['holiday'], d['absent'], d['present_rate']])
        writer.writerow([])
        # Staff records
        writer.writerow(['Staff ID','Full Name','Department','Position','Time In','Time Out','Status','Late (min)','Early Dep (min)','Work Hrs','OT Hrs'])
        for s in staff_records:
            writer.writerow([s['staff_id'], s['full_name'], s['department'], s['position'], s['time_in'], s['time_out'], s['status'], s['late_minutes'], s['early_departure_minutes'], s['work_hours'], s['overtime_hours']])
        resp = make_response(output.getvalue())
        fname = f"daily_attendance_report_{date_obj.strftime('%Y_%m_%d')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
        resp.headers['Content-Type'] = 'text/csv'
        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return resp

    # PDF export
    if format_type == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=18)
            styles = getSampleStyleSheet()
            elements = []
            elements.append(Paragraph(f"Daily Attendance Report - {date_label}", styles['Title']))
            elements.append(Spacer(1, 12))
            # Overall summary table
            ov_headers = ['Total','Present','Late','Leave','On Duty','Holiday','Absent']
            ov_data = [ov_headers, [total_staff, overall['present'], overall['late'], overall['leave'], overall['on_duty'], overall['holiday'], overall['absent']]]
            ov_table = Table(ov_data, repeatRows=1)
            ov_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            elements.append(ov_table)
            elements.append(Spacer(1, 12))
            # Department summary table
            dep_headers = ['Department','Total','Present','Late','Leave','On Duty','Holiday','Absent','%Present']
            dep_data = [dep_headers]
            for d in dept_rows:
                dep_data.append([d['department'], d['total'], d['present'], d['late'], d['leave'], d['on_duty'], d['holiday'], d['absent'], d['present_rate']])
            dep_table = Table(dep_data, repeatRows=1)
            dep_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            elements.append(dep_table)
            elements.append(Spacer(1, 12))
            # Staff table (compact)
            st_headers = ['Staff ID','Name','Dept','Pos','In','Out','Status','Late','Early','Hrs','OT']
            st_data = [st_headers]
            for s in staff_records:
                st_data.append([s['staff_id'], s['full_name'], s['department'], s['position'], s['time_in'], s['time_out'], s['status'], s['late_minutes'], s['early_departure_minutes'], s['work_hours'], s['overtime_hours']])
            st_table = Table(st_data, repeatRows=1)
            st_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
            ]))
            elements.append(st_table)
            doc.build(elements)
            pdf = buffer.getvalue(); buffer.close()
            resp = make_response(pdf)
            fname = f"daily_attendance_report_{date_obj.strftime('%Y_%m_%d')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
            resp.headers['Content-Type'] = 'application/pdf'
            resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            return resp
        except ImportError:
            return jsonify({'success': False, 'error': 'PDF generation requires reportlab. Install with: pip install reportlab'})
        except Exception as e:
            return jsonify({'success': False, 'error': f'PDF generation failed: {str(e)}'})

    # Default: Excel export
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()
    
    # Create Daily Records sheet FIRST as the active sheet
    ws_daily = wb.active
    ws_daily.title = 'Daily Records'

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # DAILY RECORDS SHEET (MAIN DATA - First sheet for immediate visibility)
    rec_headers = ['Staff ID','Full Name','Department','Position','Time In','Time Out','Status','Late (min)','Early Dep (min)','Work Hrs','OT Hrs']
    
    # Add title
    ws_daily.cell(row=1, column=1, value=f'Daily Attendance Report - Individual Staff Records - {date_label}').font = Font(bold=True, size=12)
    ws_daily.merge_cells('A1:K1')
    ws_daily.append([])  # Empty row
    
    # Add headers
    for col, h in enumerate(rec_headers, 1):
        c = ws_daily.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
    
    # Add staff data
    r = 4
    for s in staff_records:
        vals = [s['staff_id'], s['full_name'], s['department'], s['position'], s['time_in'], s['time_out'], s['status'], s['late_minutes'], s['early_departure_minutes'], s['work_hours'], s['overtime_hours']]
        for cidx, val in enumerate(vals, 1):
            cell = ws_daily.cell(row=r, column=cidx, value=val)
            cell.border = border
        r += 1
    
    # Set column widths
    widths = [14, 26, 18, 18, 12, 12, 12, 12, 14, 12, 12]
    for col, w in enumerate(widths, 1):
        ws_daily.column_dimensions[get_column_letter(col)].width = w

    # SUMMARY SHEET (Second sheet)
    ws_summary = wb.create_sheet(title='Summary')
    ws_summary.cell(row=1, column=1, value=f'Daily Attendance Report - Summary - {date_label}').font = Font(bold=True, size=12)
    ws_summary.append([])

    # Overall summary
    sum_headers = ['Total Staff','Present','Late','Leave','On Duty','Holiday','Absent']
    for col, h in enumerate(sum_headers, 1):
        c = ws_summary.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
    vals = [total_staff, overall['present'], overall['late'], overall['leave'], overall['on_duty'], overall['holiday'], overall['absent']]
    for cidx, val in enumerate(vals, 1):
        cell = ws_summary.cell(row=4, column=cidx, value=val)
        cell.border = border
    for col in range(1, len(sum_headers)+1):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18

    # Department Summary sheet (Third sheet)
    ws_dept = wb.create_sheet(title='Department Summary')
    dep_headers = ['Department','Total','Present','Late','Leave','On Duty','Holiday','Absent','Present Rate (%)']
    for col, h in enumerate(dep_headers, 1):
        c = ws_dept.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
    r = 2
    for d in dept_rows:
        vals = [d['department'], d['total'], d['present'], d['late'], d['leave'], d['on_duty'], d['holiday'], d['absent'], d['present_rate']]
        for cidx, val in enumerate(vals, 1):
            cell = ws_dept.cell(row=r, column=cidx, value=val)
            cell.border = border
        r += 1
    for col in range(1, len(dep_headers)+1):
        ws_dept.column_dimensions[get_column_letter(col)].width = 20

    output = BytesIO(); wb.save(output); output.seek(0)
    resp = make_response(output.getvalue())
    fname = f"daily_attendance_report_{date_obj.strftime('%Y_%m_%d')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp


    resp = make_response(output.getvalue())
    fname = f"performance_report_{year}_{str(month).zfill(2)}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp

def generate_overtime_report(school_id, year, month, format_type):
    """Generate overtime report - placeholder"""
    excel_generator = ExcelReportGenerator()
    if month:
        return excel_generator.create_monthly_report(school_id, year, month)
    else:
        today = datetime.datetime.now().strftime('%Y-%m-%d')
        return excel_generator.create_staff_attendance_report(school_id, today, today)



# Analytics Dashboard Routes
@app.route('/analytics_dashboard')
def analytics_dashboard():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return redirect(url_for('index'))

    today = datetime.datetime.now()
    thirty_days_ago = today - datetime.timedelta(days=30)

    return render_template('analytics_dashboard.html',
                         start_date=thirty_days_ago.strftime('%Y-%m-%d'),
                         end_date=today.strftime('%Y-%m-%d'))

@app.route('/get_analytics_data')
def get_analytics_data():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    department = request.args.get('department')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range is required'})

    try:
        data_viz = DataVisualization()

        # Generate all chart data
        analytics_data = {
            'attendance_pie': data_viz.generate_attendance_pie_chart(school_id, start_date, end_date),
            'daily_trends': data_viz.generate_daily_trends_chart(school_id, start_date, end_date),
            'department_comparison': data_viz.generate_department_comparison_chart(school_id, start_date, end_date),
            'weekly_pattern': data_viz.generate_weekly_pattern_chart(school_id, start_date, end_date),
            'overtime_analysis': data_viz.generate_overtime_analysis_chart(school_id, start_date, end_date)
        }

        return jsonify({'success': True, 'data': analytics_data})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_performance_metrics')
def get_performance_metrics():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range is required'})

    try:
        data_viz = DataVisualization()
        metrics = data_viz.generate_performance_metrics(school_id, start_date, end_date)

        return jsonify({'success': True, 'data': metrics})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_heatmap_data')
def get_heatmap_data():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    if not year or not month:
        return jsonify({'success': False, 'error': 'Year and month are required'})

    try:
        data_viz = DataVisualization()
        heatmap_data = data_viz.generate_monthly_heatmap_data(school_id, year, month)

        return jsonify({'success': True, 'data': heatmap_data})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/export_analytics_data')
def export_analytics_data():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    export_type = request.args.get('export_type', 'analytics')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Date range is required'})

    # Use the Excel generator to create analytics report
    excel_generator = ExcelReportGenerator()
    return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)

# Notification System Routes
@app.route('/get_notifications')
def get_notifications():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    user_id = session['user_id']
    user_type = session['user_type']
    limit = request.args.get('limit', 50, type=int)
    unread_only = request.args.get('unread_only', False, type=bool)

    notification_manager = NotificationManager()
    result = notification_manager.get_user_notifications(user_id, user_type, limit, unread_only)

    return jsonify(result)

@app.route('/mark_notification_read', methods=['POST'])
def mark_notification_read():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    notification_id = request.form.get('notification_id', type=int)
    user_id = session['user_id']

    if not notification_id:
        return jsonify({'success': False, 'error': 'Notification ID required'})

    notification_manager = NotificationManager()
    result = notification_manager.mark_notification_read(notification_id, user_id)

    return jsonify(result)

@app.route('/send_system_notification', methods=['POST'])
def send_system_notification():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    title = request.form.get('title')
    message = request.form.get('message')
    user_type = request.form.get('user_type', 'all')
    school_id = session['school_id']

    if not title or not message:
        return jsonify({'success': False, 'error': 'Title and message are required'})

    notification_manager = NotificationManager()
    result = notification_manager.send_system_notification(user_type, title, message, school_id)

    return jsonify(result)

@app.route('/get_notification_count')
def get_notification_count():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    user_id = session['user_id']
    user_type = session['user_type']

    db = get_db()
    unread_count = db.execute('''
        SELECT COUNT(*) as count FROM notifications
        WHERE user_id = ? AND user_type = ? AND is_read = 0
    ''', (user_id, user_type)).fetchone()

    return jsonify({
        'success': True,
        'unread_count': unread_count['count']
    })

# Integrate notifications with existing attendance processing
@app.route('/process_attendance_with_notifications', methods=['POST'])
def process_attendance_with_notifications():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id', type=int)
    verification_type = request.form.get('verification_type')
    timestamp_str = request.form.get('timestamp')

    if not all([staff_id, verification_type, timestamp_str]):
        return jsonify({'success': False, 'error': 'Missing required parameters'})

    try:
        timestamp = datetime.datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
        school_id = session['school_id']

        # Process attendance
        attendance_manager = AdvancedAttendanceManager()
        attendance_result = attendance_manager.process_attendance_with_overtime(
            staff_id, school_id, verification_type, timestamp
        )

        # Send notifications based on attendance result
        if attendance_result['success']:
            notification_manager = NotificationManager()

            if attendance_result.get('status') == 'late':
                # Send late arrival alert
                notification_manager.send_attendance_alert(
                    staff_id, 'late_arrival', {
                        'time_in': attendance_result.get('time_in'),
                        'late_minutes': attendance_result.get('late_minutes', 0)
                    }
                )
            elif attendance_result.get('overtime_hours', 0) > 0:
                # Send overtime alert
                notification_manager.send_attendance_alert(
                    staff_id, 'overtime_alert', {
                        'overtime_hours': attendance_result.get('overtime_hours')
                    }
                )

        return jsonify(attendance_result)

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Integrate notifications with leave processing
@app.route('/process_leave_with_notifications', methods=['POST'])
def process_leave_with_notifications():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    leave_id = request.form.get('leave_id', type=int)
    action = request.form.get('action')  # approve or reject
    admin_reason = request.form.get('admin_reason')

    if not leave_id or not action:
        return jsonify({'success': False, 'error': 'Leave ID and action are required'})

    try:
        db = get_db()
        admin_id = session['user_id']

        # Update leave status
        status = 'approved' if action == 'approve' else 'rejected'
        db.execute('''
            UPDATE leave_applications
            SET status = ?, processed_by = ?, processed_at = ?
            WHERE id = ?
        ''', (status, admin_id, datetime.datetime.now(), leave_id))
        db.commit()

        # Send notification
        notification_manager = NotificationManager()
        notification_result = notification_manager.send_leave_notification(
            leave_id, status, admin_reason
        )

        return jsonify({
            'success': True,
            'message': f'Leave application {status} successfully',
            'notification_sent': notification_result['success']
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/staff/dashboard')
def staff_dashboard():
    if 'user_id' not in session or session['user_type'] != 'staff':
        return redirect(url_for('index'))

    db = get_db()
    staff_db_id = session['user_id']

    # Get staff information including staff_id
    staff_info = db.execute('''
        SELECT staff_id, full_name, shift_type FROM staff WHERE id = ?
    ''', (staff_db_id,)).fetchone()

    if not staff_info:
        return redirect(url_for('index'))

    # Get attendance records for the current month, include entry time and shift type
    today = datetime.date.today()
    first_day = today.replace(day=1)
    last_day = (today.replace(day=28) + datetime.timedelta(days=4)).replace(day=1) - datetime.timedelta(days=1)

    attendance_records = db.execute('''
        SELECT a.date, a.time_in, a.time_out, a.status, s.shift_type
        FROM attendance a
        JOIN staff s ON a.staff_id = s.id
        WHERE a.staff_id = ? AND a.date BETWEEN ? AND ?
        ORDER BY a.date DESC
    ''', (staff_db_id, first_day, last_day)).fetchall()

    # For each record, determine present/late based on shift type
    attendance = []
    for record in attendance_records:
        entry_time = record['time_in']
        shift_type = record['shift_type']
        status = record['status']
        # Attendance logic based on shift type
        if entry_time:
            entry_dt = datetime.datetime.strptime(entry_time, '%H:%M:%S')
            if shift_type == 'general':
                cutoff = datetime.time(9, 0)
            elif shift_type == 'over':
                cutoff = datetime.time(9, 0)
            else:
                cutoff = datetime.time(9, 0)
            present = entry_dt.time() <= cutoff
            status_display = 'Present' if present else 'Late'
        else:
            status_display = status.capitalize() if status else 'Absent'
        attendance.append({
            'date': record['date'],
            'entry_time': entry_time or '--:--:--',
            'shift_type': shift_type.capitalize() if shift_type else 'General',
            'status': status_display
        })

    # Get leave applications
    leaves = db.execute('''
        SELECT id, leave_type, start_date, end_date, reason, status
        FROM leave_applications
        WHERE staff_id = ?
        ORDER BY start_date DESC
    ''', (staff_db_id,)).fetchall()

    # Get on-duty applications
    on_duty_applications = db.execute('''
        SELECT id, duty_type, start_date, end_date, start_time, end_time, location, purpose, reason, status
        FROM on_duty_applications
        WHERE staff_id = ?
        ORDER BY start_date DESC
    ''', (staff_db_id,)).fetchall()

    # Get permission applications
    permission_applications = db.execute('''
        SELECT id, permission_type, permission_date, start_time, end_time, duration_hours, reason, status
        FROM permission_applications
        WHERE staff_id = ?
        ORDER BY permission_date DESC
    ''', (staff_db_id,)).fetchall()

    return render_template('staff_dashboard.html',
                         attendance=attendance,
                         leaves=leaves,
                         on_duty_applications=on_duty_applications,
                         permission_applications=permission_applications,
                         today=today,
                         staff_info=staff_info)

@app.route('/admin/department_shifts')
def department_shifts():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return redirect(url_for('index'))

    db = get_db()
    school_id = session['school_id']

    # Get current department shift mappings
    mappings = db.execute('''
        SELECT department, default_shift_type, created_at, updated_at
        FROM department_shift_mappings
        WHERE school_id = ?
        ORDER BY department
    ''', (school_id,)).fetchall()

    # Get all departments currently in use
    departments = db.execute('''
        SELECT DISTINCT department
        FROM staff
        WHERE school_id = ? AND department IS NOT NULL AND department != ''
        ORDER BY department
    ''', (school_id,)).fetchall()

    return render_template('department_shifts.html', mappings=mappings, departments=departments)

@app.route('/api/department_shifts', methods=['GET', 'POST', 'DELETE'])
def api_department_shifts():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized access'}), 401

    db = get_db()
    school_id = session['school_id']

    if request.method == 'GET':
        try:
            # Get all department shift mappings
            mappings = db.execute('''
                SELECT department, default_shift_type, created_at, updated_at
                FROM department_shift_mappings
                WHERE school_id = ?
                ORDER BY department
            ''', [school_id]).fetchall()

            mappings_list = []
            for mapping in mappings:
                mappings_list.append({
                    'department': mapping['department'],
                    'default_shift_type': mapping['default_shift_type'],
                    'created_at': mapping['created_at'],
                    'updated_at': mapping['updated_at']
                })

            return jsonify({
                'success': True,
                'mappings': mappings_list
            })
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Error loading department shifts: {str(e)}'
            }), 500

    elif request.method == 'POST':
        try:
            data = request.get_json()
            department = data.get('department', '').strip()
            shift_type = data.get('shift_type', '').strip()

            if not department or not shift_type:
                return jsonify({
                    'success': False,
                    'message': 'Department and shift type are required'
                }), 400

            # Check if mapping already exists
            existing = db.execute('''
                SELECT id FROM department_shift_mappings
                WHERE school_id = ? AND department = ?
            ''', [school_id, department]).fetchone()

            if existing:
                return jsonify({
                    'success': False,
                    'message': f'Department mapping for "{department}" already exists'
                }), 400

            # Create new mapping
            current_time = datetime.datetime.now().isoformat()
            db.execute('''
                INSERT INTO department_shift_mappings
                (school_id, department, default_shift_type, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
            ''', [school_id, department, shift_type, current_time, current_time])

            db.commit()

            return jsonify({
                'success': True,
                'message': f'Department shift mapping for "{department}" created successfully'
            })

        except Exception as e:
            db.rollback()
            return jsonify({
                'success': False,
                'message': f'Error creating department mapping: {str(e)}'
            }), 500

    elif request.method == 'DELETE':
        try:
            data = request.get_json()
            department = data.get('department', '').strip()

            if not department:
                return jsonify({
                    'success': False,
                    'message': 'Department is required'
                }), 400

            # Delete the mapping
            result = db.execute('''
                DELETE FROM department_shift_mappings
                WHERE school_id = ? AND department = ?
            ''', [school_id, department])

            if result.rowcount == 0:
                return jsonify({
                    'success': False,
                    'message': f'Department mapping for "{department}" not found'
                }), 404

            db.commit()

            return jsonify({
                'success': True,
                'message': f'Department shift mapping for "{department}" deleted successfully'
            })

        except Exception as e:
            db.rollback()
            return jsonify({
                'success': False,
                'message': f'Error deleting department mapping: {str(e)}'
            }), 500

# Test route for department shifts debugging
@app.route('/test/department_shifts')
def test_department_shifts():
    return render_template('test_department_shifts.html')

# Debug route to check database tables
@app.route('/api/debug/tables')
def debug_tables():
    try:
        db = get_db()

        # Get all tables
        tables = db.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
        table_list = [row['name'] for row in tables]

        result = f"Available tables: {table_list}\n\n"

        # Check if department_shift_mappings table exists
        if 'department_shift_mappings' in table_list:
            result += "department_shift_mappings table EXISTS\n"

            # Get table structure
            schema = db.execute("PRAGMA table_info(department_shift_mappings)").fetchall()
            result += "Table structure:\n"
            for col in schema:
                result += f"  {col['name']} ({col['type']})\n"

            # Get record count
            count = db.execute("SELECT COUNT(*) as count FROM department_shift_mappings").fetchone()
            result += f"\nRecord count: {count['count']}\n"

            # Get sample records
            if count['count'] > 0:
                records = db.execute("SELECT * FROM department_shift_mappings LIMIT 5").fetchall()
                result += "\nSample records:\n"
                for record in records:
                    result += f"  {dict(record)}\n"
        else:
            result += "department_shift_mappings table does NOT exist\n"

        return result, 200, {'Content-Type': 'text/plain'}

    except Exception as e:
        return f"Error checking database: {str(e)}", 500, {'Content-Type': 'text/plain'}

@app.route('/admin/staff_management')
def staff_management():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return redirect(url_for('index'))

    db = get_db()
    school_id = session['school_id']

    # Get all staff with comprehensive details
    staff = db.execute('''
        SELECT id, staff_id, full_name, first_name, last_name,
               date_of_birth, date_of_joining, department, destination,
               position, gender, phone, email, shift_type, photo_url
        FROM staff
        WHERE school_id = ?
        ORDER BY full_name
    ''', (school_id,)).fetchall()

    # Get department shift mappings for reference
    dept_mappings = db.execute('''
        SELECT department, default_shift_type
        FROM department_shift_mappings
        WHERE school_id = ?
    ''', (school_id,)).fetchall()

    dept_shift_map = {mapping['department']: mapping['default_shift_type'] for mapping in dept_mappings}

    return render_template('staff_management.html', staff=staff, dept_shift_map=dept_shift_map)

@app.route('/admin/work_time_assignment')
def work_time_assignment():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return redirect(url_for('index'))

    db = get_db()
    school_id = session['school_id']

    # Get total staff count for sidebar
    total_staff_count = db.execute('''
        SELECT COUNT(*) as count FROM staff WHERE school_id = ?
    ''', (school_id,)).fetchone()['count']

    return render_template('work_time_assignment.html', total_staff_count=total_staff_count)

@app.route('/admin/dashboard')
def admin_dashboard():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return redirect(url_for('index'))

    db = get_db()
    school_id = session['school_id']

    # Get all staff
    staff = db.execute('''
        SELECT id, staff_id, full_name, department, position
        FROM staff
        WHERE school_id = ?
        ORDER BY full_name
    ''', (school_id,)).fetchall()

    # Get pending leave applications
    pending_leaves = db.execute('''
        SELECT l.id, s.full_name, l.leave_type, l.start_date, l.end_date, l.reason
        FROM leave_applications l
        JOIN staff s ON l.staff_id = s.id
        WHERE l.school_id = ? AND l.status = 'pending'
        ORDER BY l.applied_at
    ''', (school_id,)).fetchall()

    # Get pending on-duty applications
    pending_on_duty = db.execute('''
        SELECT od.id, s.full_name, od.duty_type, od.start_date, od.end_date, od.start_time, od.end_time, od.location, od.purpose, od.reason
        FROM on_duty_applications od
        JOIN staff s ON od.staff_id = s.id
        WHERE od.school_id = ? AND od.status = 'pending'
        ORDER BY od.applied_at
    ''', (school_id,)).fetchall()

    # Get pending permission applications
    pending_permissions = db.execute('''
        SELECT p.id, s.full_name, p.permission_type, p.permission_date, p.start_time, p.end_time, p.duration_hours, p.reason
        FROM permission_applications p
        JOIN staff s ON p.staff_id = s.id
        WHERE p.school_id = ? AND p.status = 'pending'
        ORDER BY p.applied_at
    ''', (school_id,)).fetchall()

    # Get today's attendance summary
    today = datetime.date.today()
    attendance_summary = db.execute('''
        SELECT
            COUNT(*) as total_staff,
            SUM(CASE WHEN a.status = 'present' THEN 1 ELSE 0 END) as present,
            SUM(CASE WHEN a.status = 'absent' THEN 1 ELSE 0 END) as absent,
            SUM(CASE WHEN a.status = 'late' THEN 1 ELSE 0 END) as late,
            SUM(CASE WHEN a.status = 'leave' THEN 1 ELSE 0 END) as on_leave,
            SUM(CASE WHEN a.status = 'on_duty' THEN 1 ELSE 0 END) as on_duty
        FROM (
            SELECT s.id, COALESCE(a.status, 'absent') as status
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id AND a.date = ?
            WHERE s.school_id = ?
        ) a
    ''', (today, school_id)).fetchone()

    # Get today's attendance details for all staff
    today_attendance = db.execute('''
        SELECT s.id as staff_id, s.staff_id as staff_number, s.full_name, s.department,
               a.time_in, a.time_out,
               COALESCE(a.status, 'absent') as status
        FROM staff s
        LEFT JOIN attendance a ON s.id = a.staff_id AND a.date = ?
        WHERE s.school_id = ?
        ORDER BY s.full_name
    ''', (today, school_id)).fetchall()

    # Check if user wants modern UI (can be a session variable or parameter)
    use_modern_ui = request.args.get('modern', 'false').lower() == 'true' or session.get('use_modern_ui', False)

    if use_modern_ui:
        return render_template('admin_dashboard_modern.html',
                             staff=staff,
                             pending_leaves=pending_leaves,
                             pending_on_duty=pending_on_duty,
                             pending_permissions=pending_permissions,
                             attendance_summary=attendance_summary,
                             today_attendance=today_attendance,
                             today=today,
                             recent_activities=[],  # Add recent activities data
                             performance={},  # Add performance metrics
                             biometric_status={},  # Add biometric status
                             last_backup='Today')  # Add backup info
    else:
        return render_template('admin_dashboard.html',
                             staff=staff,
                             pending_leaves=pending_leaves,
                             pending_on_duty=pending_on_duty,
                             pending_permissions=pending_permissions,
                             attendance_summary=attendance_summary,
                             today_attendance=today_attendance,
                             today=today)


@app.route('/admin/export_dashboard_data')
def export_dashboard_data():
    """Comprehensive admin dashboard data export with multiple format support"""
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    try:
        school_id = session['school_id']
        export_format = request.args.get('format', 'excel').lower()
        export_type = request.args.get('type', 'all').lower()  # all, staff, attendance, applications

        # Use the ExcelReportGenerator for comprehensive reports
        excel_generator = ExcelReportGenerator()

        if export_type == 'staff':
            # Export only staff data using the existing method
            today = datetime.datetime.now().strftime('%Y-%m-%d')
            return excel_generator.create_staff_profile_report(school_id)
        elif export_type == 'attendance':
            # Export attendance data for current month
            today = datetime.date.today()
            start_date = today.replace(day=1).strftime('%Y-%m-%d')
            end_date = today.strftime('%Y-%m-%d')
            return excel_generator.create_staff_attendance_report(school_id, start_date, end_date)
        elif export_type == 'applications':
            # Export applications data
            return export_applications_data(school_id, export_format)
        else:
            # Export comprehensive dashboard data
            return export_comprehensive_dashboard_data(school_id, export_format)

    except Exception as e:
        print(f"Dashboard export error: {str(e)}")
        return jsonify({'success': False, 'error': f'Export failed: {str(e)}'})

def export_applications_data(school_id, export_format='excel'):
    """Export applications data (leave, on-duty, permission)"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    db = get_db()

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 1. Leave Applications Sheet
    ws_leave = wb.create_sheet("Leave Applications")
    leave_headers = ['S.No', 'Staff Name', 'Leave Type', 'Start Date', 'End Date', 'Days', 'Reason', 'Status', 'Applied Date']

    # Add headers
    for col, header in enumerate(leave_headers, 1):
        cell = ws_leave.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Get leave applications data
    leave_apps = db.execute('''
        SELECT l.*, s.full_name
        FROM leave_applications l
        JOIN staff s ON l.staff_id = s.id
        WHERE l.school_id = ?
        ORDER BY l.applied_at DESC
    ''', (school_id,)).fetchall()

    for row_idx, app in enumerate(leave_apps, 2):
        ws_leave.cell(row=row_idx, column=1, value=row_idx-1)
        ws_leave.cell(row=row_idx, column=2, value=app['full_name'])
        ws_leave.cell(row=row_idx, column=3, value=app['leave_type'])
        ws_leave.cell(row=row_idx, column=4, value=app['start_date'])
        ws_leave.cell(row=row_idx, column=5, value=app['end_date'])
        ws_leave.cell(row=row_idx, column=6, value=app.get('total_days', 'N/A'))
        ws_leave.cell(row=row_idx, column=7, value=app['reason'] or 'N/A')
        ws_leave.cell(row=row_idx, column=8, value=app['status'])
        ws_leave.cell(row=row_idx, column=9, value=app['applied_at'])

    # 2. On-Duty Applications Sheet
    ws_duty = wb.create_sheet("On-Duty Applications")
    duty_headers = ['S.No', 'Staff Name', 'Duty Type', 'Start Date', 'End Date', 'Start Time', 'End Time', 'Location', 'Purpose', 'Status']

    for col, header in enumerate(duty_headers, 1):
        cell = ws_duty.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    duty_apps = db.execute('''
        SELECT od.*, s.full_name
        FROM on_duty_applications od
        JOIN staff s ON od.staff_id = s.id
        WHERE od.school_id = ?
        ORDER BY od.applied_at DESC
    ''', (school_id,)).fetchall()

    for row_idx, app in enumerate(duty_apps, 2):
        ws_duty.cell(row=row_idx, column=1, value=row_idx-1)
        ws_duty.cell(row=row_idx, column=2, value=app['full_name'])
        ws_duty.cell(row=row_idx, column=3, value=app['duty_type'])
        ws_duty.cell(row=row_idx, column=4, value=app['start_date'])
        ws_duty.cell(row=row_idx, column=5, value=app['end_date'])
        ws_duty.cell(row=row_idx, column=6, value=app['start_time'])
        ws_duty.cell(row=row_idx, column=7, value=app['end_time'])
        ws_duty.cell(row=row_idx, column=8, value=app['location'])
        ws_duty.cell(row=row_idx, column=9, value=app['purpose'])
        ws_duty.cell(row=row_idx, column=10, value=app['status'])

    # 3. Permission Applications Sheet
    ws_perm = wb.create_sheet("Permission Applications")
    perm_headers = ['S.No', 'Staff Name', 'Permission Type', 'Date', 'Start Time', 'End Time', 'Duration (Hours)', 'Reason', 'Status']

    for col, header in enumerate(perm_headers, 1):
        cell = ws_perm.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    perm_apps = db.execute('''
        SELECT p.*, s.full_name
        FROM permission_applications p
        JOIN staff s ON p.staff_id = s.id
        WHERE p.school_id = ?
        ORDER BY p.applied_at DESC
    ''', (school_id,)).fetchall()

    for row_idx, app in enumerate(perm_apps, 2):
        ws_perm.cell(row=row_idx, column=1, value=row_idx-1)
        ws_perm.cell(row=row_idx, column=2, value=app['full_name'])
        ws_perm.cell(row=row_idx, column=3, value=app['permission_type'])
        ws_perm.cell(row=row_idx, column=4, value=app['permission_date'])
        ws_perm.cell(row=row_idx, column=5, value=app['start_time'])
        ws_perm.cell(row=row_idx, column=6, value=app['end_time'])
        ws_perm.cell(row=row_idx, column=7, value=app['duration_hours'])
        ws_perm.cell(row=row_idx, column=8, value=app['reason'])
        ws_perm.cell(row=row_idx, column=9, value=app['status'])

    # Auto-adjust column widths for all sheets
    for ws in [ws_leave, ws_duty, ws_perm]:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.getvalue())
    filename = f'applications_report_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

    return response

def export_comprehensive_dashboard_data(school_id, export_format='excel'):
    """Export comprehensive dashboard data including all sections"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    db = get_db()
    today = datetime.date.today()

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16, color="2F5597")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 1. Dashboard Summary Sheet
    ws_summary = wb.create_sheet("Dashboard Summary")

    # Title
    ws_summary.merge_cells('A1:F1')
    title_cell = ws_summary['A1']
    title_cell.value = f"Admin Dashboard Summary - {today.strftime('%Y-%m-%d')}"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')

    # Attendance Summary
    attendance_summary = db.execute('''
        SELECT
            COUNT(*) as total_staff,
            SUM(CASE WHEN a.status = 'present' THEN 1 ELSE 0 END) as present,
            SUM(CASE WHEN a.status = 'absent' THEN 1 ELSE 0 END) as absent,
            SUM(CASE WHEN a.status = 'late' THEN 1 ELSE 0 END) as late,
            SUM(CASE WHEN a.status = 'leave' THEN 1 ELSE 0 END) as on_leave,
            SUM(CASE WHEN a.status = 'on_duty' THEN 1 ELSE 0 END) as on_duty
        FROM (
            SELECT s.id, COALESCE(a.status, 'absent') as status
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id AND a.date = ?
            WHERE s.school_id = ?
        ) a
    ''', (today, school_id)).fetchone()

    # Add attendance summary
    ws_summary['A3'] = "Today's Attendance Summary"
    ws_summary['A3'].font = Font(bold=True, size=14)

    summary_data = [
        ['Metric', 'Count'],
        ['Total Staff', attendance_summary['total_staff']],
        ['Present', attendance_summary['present']],
        ['Absent', attendance_summary['absent']],
        ['Late', attendance_summary['late']],
        ['On Leave', attendance_summary['on_leave']],
        ['On Duty', attendance_summary['on_duty']]
    ]

    for row_idx, row_data in enumerate(summary_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == 4:  # Header row
                cell.font = header_font
                cell.fill = header_fill
            cell.border = border

    # 2. Today's Attendance Sheet
    ws_attendance = wb.create_sheet("Today's Attendance")

    attendance_headers = ['S.No', 'Staff ID', 'Staff Name', 'Department', 'Status', 'Time In', 'Time Out']

    for col, header in enumerate(attendance_headers, 1):
        cell = ws_attendance.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    today_attendance = db.execute('''
        SELECT s.staff_id, s.full_name, s.department,
               COALESCE(a.status, 'absent') as status,
               a.time_in, a.time_out
        FROM staff s
        LEFT JOIN attendance a ON s.id = a.staff_id AND a.date = ?
        WHERE s.school_id = ?
        ORDER BY s.full_name
    ''', (today, school_id)).fetchall()

    for row_idx, record in enumerate(today_attendance, 2):
        ws_attendance.cell(row=row_idx, column=1, value=row_idx-1)
        ws_attendance.cell(row=row_idx, column=2, value=record['staff_id'] or 'N/A')
        ws_attendance.cell(row=row_idx, column=3, value=record['full_name'])
        ws_attendance.cell(row=row_idx, column=4, value=record['department'] or 'N/A')
        ws_attendance.cell(row=row_idx, column=5, value=record['status'])
        ws_attendance.cell(row=row_idx, column=6, value=record['time_in'] or 'N/A')
        ws_attendance.cell(row=row_idx, column=7, value=record['time_out'] or 'N/A')

    # Auto-adjust column widths
    for ws in [ws_summary, ws_attendance]:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.getvalue())
    filename = f'dashboard_comprehensive_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

    return response


@app.route('/toggle_modern_ui')
def toggle_modern_ui():
    """Toggle between modern and legacy UI"""
    if 'user_id' not in session:
        return redirect(url_for('index'))

    # Toggle the modern UI setting
    session['use_modern_ui'] = not session.get('use_modern_ui', False)

    # Redirect back to the referring page or dashboard
    return redirect(request.referrer or url_for('admin_dashboard'))


@app.route('/company/dashboard')
def company_dashboard():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return redirect(url_for('index'))

    db = get_db()

    # Get all schools
    schools = db.execute('''
        SELECT s.id, s.name, s.address, s.contact_email, s.contact_phone,
               COUNT(a.id) as admin_count,
               COUNT(st.id) as staff_count
        FROM schools s
        LEFT JOIN admins a ON s.id = a.school_id
        LEFT JOIN staff st ON s.id = st.school_id
        GROUP BY s.id
        ORDER BY s.name
    ''').fetchall()

    return render_template('company_dashboard.html', schools=schools)

@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    """
    LEGACY ROUTE - DEPRECATED
    This route is kept for backward compatibility but should not be used.
    Use /biometric_attendance instead for proper user-controlled verification.
    """
    if 'user_id' not in session or session['user_type'] != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Return error to prevent automatic updates
    return jsonify({
        'success': False,
        'error': 'This route is deprecated. Please use biometric verification system for attendance marking.'
    })

def validate_verification_rules(verification_type, existing_attendance, current_time, staff_shift_type='general'):
    """Validate business rules for biometric verification using shift management"""

    if verification_type == 'check-in':
        # Check if user has already checked in today
        if existing_attendance and existing_attendance['time_in']:
            return 'You have already checked in today. Multiple check-ins are not allowed.'
        return None

    elif verification_type == 'check-out':
        # Check if user has already checked out today
        if existing_attendance and existing_attendance['time_out']:
            return 'You have already checked out today. Multiple check-outs are not allowed.'

        # Check if user has checked in first
        if not existing_attendance or not existing_attendance['time_in']:
            return 'You must check in first before checking out.'
        return None

    return 'Invalid verification type'

@app.route('/biometric_attendance', methods=['POST'])
def biometric_attendance():
    """
    DEPRECATED: Handle biometric attendance verification with manual type selection

    This route is now deprecated. Staff should use the biometric device directly,
    and the system will automatically poll for new verifications.
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})

    # Return error message directing users to use the device directly
    return jsonify({
        'success': False,
        'error': 'Please use the biometric device directly to select your verification type and verify.'
    })

@app.route('/check_device_verification', methods=['POST'])
def check_device_verification():
    """Check for recent biometric verification from the device for a specific staff member"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})

    # Allow both staff and admin to use this
    if session['user_type'] not in ['staff', 'admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Handle both staff and admin users
    if session['user_type'] == 'staff':
        staff_id = session['user_id']
    else:
        # For admin, get staff_id from form
        staff_id = request.form.get('staff_id')
        if not staff_id:
            return jsonify({'success': False, 'error': 'Staff ID required for admin verification'})

    school_id = session['school_id']
    device_ip = request.form.get('device_ip', '192.168.1.201')

    current_datetime = datetime.datetime.now()
    db = get_db()

    # Get staff information
    staff_info = db.execute('''
        SELECT staff_id, full_name FROM staff WHERE id = ?
    ''', (staff_id,)).fetchone()

    if not staff_info:
        return jsonify({'success': False, 'error': 'Staff not found'})

    try:
        # Check for recent verification from the device (within last 30 seconds)
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({
                'success': False,
                'error': 'Failed to connect to biometric device'
            })

        # Look for recent attendance records for this staff member
        recent_cutoff = current_datetime - datetime.timedelta(seconds=30)
        recent_records = zk_device.get_new_attendance_records(recent_cutoff)

        staff_recent_record = None
        for record in recent_records:
            if str(record['user_id']) == str(staff_info['staff_id']):
                staff_recent_record = record
                break

        zk_device.disconnect()

        if not staff_recent_record:
            return jsonify({
                'success': False,
                'error': 'No recent biometric verification found. Please use the biometric device to verify your attendance.'
            })

        # Process the verification from the device
        verification_type = staff_recent_record['verification_type']
        verification_time = staff_recent_record['timestamp']

        # Validate business rules
        today = verification_time.date()
        existing_attendance = db.execute('''
            SELECT * FROM attendance WHERE staff_id = ? AND date = ?
        ''', (staff_id, today)).fetchone()

        # Get staff shift type for validation
        staff_info = db.execute('SELECT shift_type FROM staff WHERE id = ?', (staff_id,)).fetchone()
        staff_shift_type = staff_info['shift_type'] if staff_info and staff_info['shift_type'] else 'general'

        validation_error = validate_verification_rules(verification_type, existing_attendance, verification_time.time(), staff_shift_type)
        if validation_error:
            return jsonify({'success': False, 'error': validation_error})

        # Log successful verification
        db.execute('''
            INSERT INTO biometric_verifications
            (staff_id, school_id, verification_type, verification_time, device_ip, biometric_method, verification_status)
            VALUES (?, ?, ?, ?, ?, 'fingerprint', 'success')
        ''', (staff_id, school_id, verification_type, verification_time, device_ip))

        # Get staff shift type
        staff_info = db.execute('SELECT shift_type FROM staff WHERE id = ?', (staff_id,)).fetchone()
        staff_shift_type = staff_info['shift_type'] if staff_info and staff_info['shift_type'] else 'general'

        # Update attendance based on verification type using enhanced shift management
        current_time = verification_time.strftime('%H:%M:%S')
        shift_manager = ShiftManager()

        if verification_type == 'check-in':
            # Calculate attendance status using shift management
            attendance_result = shift_manager.calculate_attendance_status(
                staff_shift_type, verification_time.time()
            )

            # Only create new record if no check-in exists (validation already checked this)
            if existing_attendance:
                return jsonify({
                    'success': False,
                    'error': 'You have already checked in today.'
                })
            else:
                db.execute('''
                    INSERT INTO attendance (staff_id, school_id, date, time_in, status,
                                          late_duration_minutes, shift_start_time, shift_end_time)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (staff_id, school_id, today, current_time, attendance_result['status'],
                      attendance_result['late_duration_minutes'],
                      attendance_result['shift_start_time'].strftime('%H:%M:%S'),
                      attendance_result['shift_end_time'].strftime('%H:%M:%S')))

        elif verification_type == 'check-out':
            # Only update check-out if no check-out exists (validation already checked this)
            if existing_attendance and existing_attendance['time_out']:
                return jsonify({
                    'success': False,
                    'error': 'You have already checked out today.'
                })
            elif existing_attendance:
                # Calculate if this is early departure
                attendance_result = shift_manager.calculate_attendance_status(
                    staff_shift_type,
                    datetime.datetime.strptime(existing_attendance['time_in'], '%H:%M:%S').time() if existing_attendance['time_in'] else verification_time.time(),
                    verification_time.time()
                )

                # Update status if early departure detected
                final_status = existing_attendance['status']
                if attendance_result['early_departure_minutes'] and attendance_result['early_departure_minutes'] > 0:
                    if final_status != 'late':  # Don't override late status
                        final_status = 'left_soon'

                db.execute('''
                    UPDATE attendance SET time_out = ?, status = ?, early_departure_minutes = ?
                    WHERE staff_id = ? AND date = ?
                ''', (current_time, final_status, attendance_result['early_departure_minutes'], staff_id, today))
            else:
                return jsonify({
                    'success': False,
                    'error': 'You must check in first before checking out.'
                })

        db.commit()

        # Format times for display in 12-hour format
        current_time_12hr = format_time_to_12hr(current_time)
        verification_time_12hr = verification_time.strftime('%Y-%m-%d %I:%M %p')

        return jsonify({
            'success': True,
            'message': f'{verification_type.title()} recorded successfully at {current_time_12hr}',
            'verification_time': verification_time_12hr,
            'verification_type': verification_type,
            'biometric_method': 'fingerprint',
            'time_recorded': current_time_12hr
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Verification error: {str(e)}'
        })

@app.route('/apply_leave', methods=['POST'])
def apply_leave():
    if 'user_id' not in session or session['user_type'] != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    school_id = session['school_id']
    leave_type = request.form.get('leave_type')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    reason = request.form.get('reason')

    db = get_db()

    db.execute('''
        INSERT INTO leave_applications
        (staff_id, school_id, leave_type, start_date, end_date, reason)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (staff_id, school_id, leave_type, start_date, end_date, reason))
    db.commit()

    return jsonify({'success': True})

@app.route('/apply_on_duty', methods=['POST'])
def apply_on_duty():
    if 'user_id' not in session or session['user_type'] != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    school_id = session['school_id']
    duty_type = request.form.get('duty_type')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')
    location = request.form.get('location')
    purpose = request.form.get('purpose')
    reason = request.form.get('reason')

    # Validate required fields
    if not all([duty_type, start_date, end_date, purpose]):
        return jsonify({'success': False, 'error': 'Please fill all required fields'})

    db = get_db()

    try:
        db.execute('''
            INSERT INTO on_duty_applications
            (staff_id, school_id, duty_type, start_date, end_date, start_time, end_time, location, purpose, reason)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (staff_id, school_id, duty_type, start_date, end_date, start_time, end_time, location, purpose, reason))
        db.commit()

        return jsonify({'success': True, 'message': 'On-duty application submitted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to submit application: {str(e)}'})

@app.route('/apply_permission', methods=['POST'])
def apply_permission():
    if 'user_id' not in session or session['user_type'] != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    school_id = session['school_id']
    permission_type = request.form.get('permission_type')
    permission_date = request.form.get('permission_date')
    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')
    reason = request.form.get('reason')

    # Validate required fields
    if not all([permission_type, permission_date, start_time, end_time, reason]):
        return jsonify({'success': False, 'error': 'Please fill all required fields'})

    # Calculate duration in hours
    try:
        from datetime import datetime
        start_dt = datetime.strptime(start_time, '%H:%M')
        end_dt = datetime.strptime(end_time, '%H:%M')
        duration = (end_dt - start_dt).total_seconds() / 3600

        if duration <= 0:
            return jsonify({'success': False, 'error': 'End time must be after start time'})

    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid time format'})

    db = get_db()

    try:
        db.execute('''
            INSERT INTO permission_applications
            (staff_id, school_id, permission_type, permission_date, start_time, end_time, duration_hours, reason)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (staff_id, school_id, permission_type, permission_date, start_time, end_time, duration, reason))
        db.commit()

        return jsonify({'success': True, 'message': 'Permission application submitted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to submit application: {str(e)}'})

@app.route('/process_leave', methods=['POST'])
def process_leave():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    leave_id = request.form.get('leave_id')
    decision = request.form.get('decision')  # 'approve' or 'reject'
    admin_id = session['user_id']
    processed_at = datetime.datetime.now()

    db = get_db()

    status = 'approved' if decision == 'approve' else 'rejected'

    db.execute('''
        UPDATE leave_applications
        SET status = ?, processed_by = ?, processed_at = ?
        WHERE id = ?
    ''', (status, admin_id, processed_at, leave_id))
    db.commit()

    return jsonify({'success': True})

@app.route('/process_on_duty', methods=['POST'])
def process_on_duty():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    application_id = request.form.get('application_id')
    decision = request.form.get('decision')  # 'approve' or 'reject'
    admin_remarks = request.form.get('admin_remarks', '')
    admin_id = session['user_id']
    processed_at = datetime.datetime.now()

    if not application_id or not decision:
        return jsonify({'success': False, 'error': 'Missing required parameters'})

    db = get_db()

    status = 'approved' if decision == 'approve' else 'rejected'

    try:
        # Get the on-duty application details before updating
        on_duty_app = db.execute('''
            SELECT staff_id, school_id, start_date, end_date, duty_type, location, purpose
            FROM on_duty_applications
            WHERE id = ?
        ''', (application_id,)).fetchone()

        if not on_duty_app:
            return jsonify({'success': False, 'error': 'On-duty application not found'})

        # Update the application status
        db.execute('''
            UPDATE on_duty_applications
            SET status = ?, processed_by = ?, processed_at = ?, admin_remarks = ?
            WHERE id = ?
        ''', (status, admin_id, processed_at, admin_remarks, application_id))

        # If approved, mark attendance as "On Duty" for the date range
        if status == 'approved':
            staff_id = on_duty_app['staff_id']
            school_id = on_duty_app['school_id']
            start_date = datetime.datetime.strptime(on_duty_app['start_date'], '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(on_duty_app['end_date'], '%Y-%m-%d').date()

            # Mark attendance for each day in the date range
            current_date = start_date
            while current_date <= end_date:
                date_str = current_date.strftime('%Y-%m-%d')

                # Check if attendance record already exists
                existing_attendance = db.execute('''
                    SELECT id, status FROM attendance
                    WHERE staff_id = ? AND date = ?
                ''', (staff_id, date_str)).fetchone()

                if existing_attendance:
                    # Update existing record to "On Duty" status
                    db.execute('''
                        UPDATE attendance
                        SET status = 'on_duty',
                            on_duty_type = ?,
                            on_duty_location = ?,
                            on_duty_purpose = ?
                        WHERE staff_id = ? AND date = ?
                    ''', (on_duty_app['duty_type'], on_duty_app['location'],
                          on_duty_app['purpose'], staff_id, date_str))
                else:
                    # Create new attendance record with "On Duty" status
                    db.execute('''
                        INSERT INTO attendance
                        (staff_id, school_id, date, status, on_duty_type, on_duty_location, on_duty_purpose)
                        VALUES (?, ?, ?, 'on_duty', ?, ?, ?)
                    ''', (staff_id, school_id, date_str, on_duty_app['duty_type'],
                          on_duty_app['location'], on_duty_app['purpose']))

                current_date += datetime.timedelta(days=1)

        db.commit()

        return jsonify({'success': True, 'message': f'On-duty application {status} successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to process application: {str(e)}'})

@app.route('/process_permission', methods=['POST'])
def process_permission():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    application_id = request.form.get('application_id')
    decision = request.form.get('decision')  # 'approve' or 'reject'
    admin_remarks = request.form.get('admin_remarks', '')
    admin_id = session['user_id']
    processed_at = datetime.datetime.now()

    if not application_id or not decision:
        return jsonify({'success': False, 'error': 'Missing required parameters'})

    db = get_db()

    status = 'approved' if decision == 'approve' else 'rejected'

    try:
        db.execute('''
            UPDATE permission_applications
            SET status = ?, processed_by = ?, processed_at = ?, admin_remarks = ?
            WHERE id = ?
        ''', (status, admin_id, processed_at, admin_remarks, application_id))
        db.commit()

        return jsonify({'success': True, 'message': f'Permission application {status} successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to process application: {str(e)}'})














# Weekly Calendar API Endpoints
@app.route('/get_weekly_attendance')
def get_weekly_attendance():
    """Get weekly attendance data for staff profile calendar"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    # Get parameters
    staff_id = request.args.get('staff_id')  # For admin viewing other staff
    week_start = request.args.get('week_start')  # YYYY-MM-DD format

    # Determine which staff to get data for
    if session['user_type'] == 'staff':
        target_staff_id = session['user_id']
    elif session['user_type'] == 'admin' and staff_id:
        target_staff_id = staff_id
    else:
        return jsonify({'success': False, 'error': 'Invalid request'})

    if not week_start:
        return jsonify({'success': False, 'error': 'Week start date required'})

    try:
        # Parse week start date
        week_start_date = datetime.datetime.strptime(week_start, '%Y-%m-%d').date()
        week_end_date = week_start_date + datetime.timedelta(days=6)

        db = get_db()

        # Get staff information including shift type
        staff_info = db.execute('''
            SELECT id, staff_id, full_name, shift_type
            FROM staff
            WHERE id = ?
        ''', (target_staff_id,)).fetchone()

        if not staff_info:
            return jsonify({'success': False, 'error': 'Staff not found'})

        # Get shift definition for this staff member
        shift_type = staff_info['shift_type'] or 'general'
        shift_def = db.execute('''
            SELECT start_time, end_time, grace_period_minutes
            FROM shift_definitions
            WHERE shift_type = ? AND is_active = 1
        ''', (shift_type,)).fetchone()

        if not shift_def:
            # Fallback to general shift
            shift_def = db.execute('''
                SELECT start_time, end_time, grace_period_minutes
                FROM shift_definitions
                WHERE shift_type = 'general' AND is_active = 1
            ''').fetchone()

        # Get attendance records for the week
        attendance_records = db.execute('''
            SELECT date, time_in, time_out, status,
                   late_duration_minutes, early_departure_minutes,
                   on_duty_type, on_duty_location, on_duty_purpose
            FROM attendance
            WHERE staff_id = ? AND date BETWEEN ? AND ?
            ORDER BY date
        ''', (target_staff_id, week_start_date, week_end_date)).fetchall()

        # Create weekly data structure
        weekly_data = []
        current_date = week_start_date

        for i in range(7):  # 7 days in a week
            date_str = current_date.strftime('%Y-%m-%d')
            day_name = current_date.strftime('%A')

            # Find attendance record for this date
            attendance_record = None
            for record in attendance_records:
                if record['date'] == date_str:
                    attendance_record = record
                    break

            # Calculate day data
            day_data = calculate_daily_attendance_data(
                current_date, attendance_record, shift_def, shift_type
            )
            day_data['day_name'] = day_name
            day_data['date'] = date_str

            weekly_data.append(day_data)
            current_date += datetime.timedelta(days=1)

        return jsonify({
            'success': True,
            'staff_info': {
                'id': staff_info['id'],
                'staff_id': staff_info['staff_id'],
                'full_name': staff_info['full_name'],
                'shift_type': shift_type
            },
            'week_start': week_start,
            'week_end': week_end_date.strftime('%Y-%m-%d'),
            'weekly_data': weekly_data
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


def calculate_daily_attendance_data(date, attendance_record, shift_def, shift_type):
    """Calculate daily attendance data with delays and early departures"""

    # Default data structure with shift timing information
    day_data = {
        'present_status': 'Absent',
        'shift_type_display': shift_type.replace('_', ' ').title() + ' Shift',
        'shift_start_time': None,
        'shift_end_time': None,
        'morning_thumb': None,
        'evening_thumb': None,
        'delay_info': None,
        'delay_duration': None,
        'arrived_soon_info': None,
        'arrived_soon_duration': None,
        'left_soon_info': None,
        'left_soon_duration': None
    }

    # Add shift timing information in 12-hour format
    if shift_def:
        day_data['shift_start_time'] = format_time_to_12hr(shift_def['start_time'])
        day_data['shift_end_time'] = format_time_to_12hr(shift_def['end_time'])

    if not attendance_record:
        return day_data

    # Check attendance status
    attendance_status = attendance_record['status'] if attendance_record['status'] else 'absent'

    # Handle on-duty status
    if attendance_status == 'on_duty':
        day_data['present_status'] = 'On Duty'
        # Safely access on-duty fields with fallbacks
        day_data['on_duty_type'] = attendance_record['on_duty_type'] if attendance_record['on_duty_type'] else 'Official Work'
        day_data['on_duty_location'] = attendance_record['on_duty_location'] if attendance_record['on_duty_location'] else 'Not specified'
        day_data['on_duty_purpose'] = attendance_record['on_duty_purpose'] if attendance_record['on_duty_purpose'] else 'Official duty'
        return day_data  # Return early for on-duty status

    # Staff was present if there's a time_in record
    if attendance_record['time_in']:
        day_data['present_status'] = 'Present'
        day_data['morning_thumb'] = format_time_to_12hr(attendance_record['time_in'])

        # Calculate arrival timing relative to shift start
        if shift_def:
            actual_time = datetime.datetime.strptime(attendance_record['time_in'], '%H:%M:%S').time()
            shift_start_time = datetime.datetime.strptime(shift_def['start_time'], '%H:%M:%S').time()
            grace_period = shift_def['grace_period_minutes'] if shift_def['grace_period_minutes'] else 10

            # Convert times to datetime objects for comparison
            actual_datetime = datetime.datetime.combine(datetime.date.today(), actual_time)
            shift_start_datetime = datetime.datetime.combine(datetime.date.today(), shift_start_time)
            grace_cutoff_datetime = shift_start_datetime + datetime.timedelta(minutes=grace_period)

            # Check if arrived early (before shift start time)
            if actual_datetime < shift_start_datetime:
                early_minutes = int((shift_start_datetime - actual_datetime).total_seconds() / 60)
                shift_start_12hr = format_time_to_12hr(shift_def['start_time'])
                actual_time_12hr = day_data['morning_thumb']

                day_data['arrived_soon_info'] = f"Arrived Soon: {format_duration_minutes(early_minutes)} (Expected: {shift_start_12hr}, Actual: {actual_time_12hr})"
                day_data['arrived_soon_duration'] = early_minutes

            # Check if arrived late (after grace period)
            elif actual_datetime > grace_cutoff_datetime:
                if attendance_record['late_duration_minutes'] and attendance_record['late_duration_minutes'] > 0:
                    delay_minutes = attendance_record['late_duration_minutes']
                    grace_cutoff_12hr = format_time_to_12hr(grace_cutoff_datetime.time().strftime('%H:%M:%S'))
                    actual_time_12hr = day_data['morning_thumb']

                    day_data['delay_info'] = f"Delay: {format_duration_minutes(delay_minutes)} (Expected: {grace_cutoff_12hr}, Actual: {actual_time_12hr})"
                    day_data['delay_duration'] = delay_minutes

    # Evening thumb out
    if attendance_record['time_out']:
        day_data['evening_thumb'] = format_time_to_12hr(attendance_record['time_out'])

        # Calculate departure timing relative to shift end
        if shift_def:
            actual_time = datetime.datetime.strptime(attendance_record['time_out'], '%H:%M:%S').time()
            shift_end_time = datetime.datetime.strptime(shift_def['end_time'], '%H:%M:%S').time()

            # Convert times to datetime objects for comparison
            actual_datetime = datetime.datetime.combine(datetime.date.today(), actual_time)
            shift_end_datetime = datetime.datetime.combine(datetime.date.today(), shift_end_time)

            # Check if left early (before shift end time)
            if actual_datetime < shift_end_datetime:
                early_minutes = int((shift_end_datetime - actual_datetime).total_seconds() / 60)
                shift_end_12hr = format_time_to_12hr(shift_def['end_time'])
                actual_time_12hr = day_data['evening_thumb']

                day_data['left_soon_info'] = f"Left Soon: {format_duration_minutes(early_minutes)} (Expected: {shift_end_12hr}, Actual: {actual_time_12hr})"
                day_data['left_soon_duration'] = early_minutes

    # No overtime functionality - removed

    return day_data


def format_time_to_12hr(time_str):
    """Convert 24-hour time string to 12-hour format"""
    if not time_str:
        return None

    try:
        # Handle both HH:MM:SS and HH:MM formats
        if len(time_str.split(':')) == 3:
            time_obj = datetime.datetime.strptime(time_str, '%H:%M:%S').time()
        else:
            time_obj = datetime.datetime.strptime(time_str, '%H:%M').time()

        # Convert to 12-hour format
        return datetime.datetime.combine(datetime.date.today(), time_obj).strftime('%I:%M %p')
    except:
        return time_str  # Return original if conversion fails


def format_duration_minutes(minutes):
    """Format duration in minutes to readable format without prefix"""
    if minutes <= 0:
        return ""

    hours = minutes // 60
    mins = minutes % 60

    if hours > 0:
        if mins > 0:
            return f"{hours}h {mins}m"
        else:
            return f"{hours}h"
    else:
        return f"{mins}m"


def format_duration_for_display(minutes, prefix):
    """Format duration in minutes to display format"""
    if minutes <= 0:
        return None

    duration_str = format_duration_minutes(minutes)
    return f"{prefix}: {duration_str}"


def format_attendance_times_to_12hr(attendance_record):
    """Convert all time fields in an attendance record to 12-hour format"""
    if not attendance_record:
        return attendance_record

    # Create a copy to avoid modifying the original
    formatted_record = dict(attendance_record)

    # Time fields to convert
    time_fields = ['time_in', 'time_out', 'shift_start_time', 'shift_end_time']

    for field in time_fields:
        if field in formatted_record and formatted_record[field]:
            formatted_record[field] = format_time_to_12hr(formatted_record[field])

    return formatted_record


@app.route('/add_staff_enhanced', methods=['POST'])
def add_staff_enhanced():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    staff_id = request.form.get('staff_id')
    first_name = request.form.get('first_name')
    last_name = request.form.get('last_name')
    date_of_birth = request.form.get('date_of_birth')
    date_of_joining = request.form.get('date_of_joining')
    department = request.form.get('department')
    destination = request.form.get('destination')
    gender = request.form.get('gender')
    phone = request.form.get('phone')
    email = request.form.get('email')
    shift_type = request.form.get('shift_type', 'general')
    password = generate_password_hash(request.form.get('password'))

    # Create full_name from first_name and last_name
    full_name = f"{first_name} {last_name}".strip() if first_name or last_name else ""

    # Validate age (must be 18 or older)
    if date_of_birth:
        try:
            birth_date = datetime.datetime.strptime(date_of_birth, '%Y-%m-%d').date()
            today = datetime.date.today()
            age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
            if age < 18:
                return jsonify({'success': False, 'error': 'Staff member must be at least 18 years old'})
        except ValueError:
            return jsonify({'success': False, 'error': 'Invalid date of birth format'})

    db = get_db()

    # Auto-assign shift type based on department if not manually specified
    if not shift_type or shift_type == 'general':
        if department:
            dept_mapping = db.execute('''
                SELECT default_shift_type
                FROM department_shift_mappings
                WHERE school_id = ? AND department = ?
            ''', (school_id, department)).fetchone()

            if dept_mapping:
                shift_type = dept_mapping['default_shift_type']
            else:
                shift_type = 'general'  # fallback

    # Always create staff in both app and device, do not check for staff ID existence
    device_ip = '192.168.1.201'  # Default device IP
    biometric_created = False
    biometric_error = None
    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            # Always overwrite user on device, do not check existence
            biometric_created = zk_device.create_user(user_id=staff_id, name=full_name, overwrite=True)
            zk_device.disconnect()
            if not biometric_created:
                biometric_error = 'Failed to create staff on biometric device. Please check device connection.'
        else:
            biometric_error = 'Cannot connect to biometric device. Staff not added.'
    except Exception as e:
        biometric_error = f'Biometric device error: {e}'

    # Handle photo upload (if any)
    photo_url = None
    if 'photo' in request.files:
        photo = request.files['photo']
        if photo and photo.filename and allowed_file(photo.filename):
            filename = secure_filename(f"{staff_id}_{photo.filename}")
            photo_path = os.path.join(app.config.get('UPLOAD_FOLDER', 'static/uploads'), filename)
            os.makedirs(os.path.dirname(photo_path), exist_ok=True)
            photo.save(photo_path)
            photo_url = f"uploads/{filename}"

    try:
        # Ensure all required columns exist in the table
        columns = db.execute("PRAGMA table_info(staff)").fetchall()
        column_names = [col['name'] for col in columns]

        # Build the insert query dynamically based on available columns
        insert_columns = ['school_id', 'staff_id', 'password_hash', 'full_name']
        insert_values = [school_id, staff_id, password, full_name]

        if 'first_name' in column_names and first_name:
            insert_columns.append('first_name')
            insert_values.append(first_name)
        if 'last_name' in column_names and last_name:
            insert_columns.append('last_name')
            insert_values.append(last_name)
        if 'date_of_birth' in column_names and date_of_birth:
            insert_columns.append('date_of_birth')
            insert_values.append(date_of_birth)
        if 'date_of_joining' in column_names and date_of_joining:
            insert_columns.append('date_of_joining')
            insert_values.append(date_of_joining)
        if 'department' in column_names and department:
            insert_columns.append('department')
            insert_values.append(department)
        if 'destination' in column_names and destination:
            insert_columns.append('destination')
            insert_values.append(destination)
        if 'gender' in column_names and gender:
            insert_columns.append('gender')
            insert_values.append(gender)
        if 'phone' in column_names and phone:
            insert_columns.append('phone')
            insert_values.append(phone)
        if 'email' in column_names and email:
            insert_columns.append('email')
            insert_values.append(email)
        if 'shift_type' in column_names:
            insert_columns.append('shift_type')
            insert_values.append(shift_type)
        if 'photo_url' in column_names and photo_url:
            insert_columns.append('photo_url')
            insert_values.append(photo_url)

        # Add created_at timestamp
        if 'created_at' in column_names:
            insert_columns.append('created_at')
            insert_values.append(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        # Build and execute the query
        placeholders = ', '.join(['?' for _ in insert_values])
        query = f"INSERT INTO staff ({', '.join(insert_columns)}) VALUES ({placeholders})"

        db.execute(query, insert_values)
        db.commit()

        return jsonify({'success': True, 'biometric_created': biometric_created, 'biometric_error': biometric_error})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/add_staff', methods=['POST'])
def add_staff():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    staff_id = request.form.get('staff_id')
    full_name = request.form.get('full_name')
    password = generate_password_hash(request.form.get('password'))
    email = request.form.get('email')
    phone = request.form.get('phone')
    department = request.form.get('department')
    position = request.form.get('position')

    # Get salary information
    basic_salary = float(request.form.get('basic_salary', 0))
    hra = float(request.form.get('hra', 0))
    transport_allowance = float(request.form.get('transport_allowance', 0))
    other_allowances = float(request.form.get('other_allowances', 0))
    pf_deduction = float(request.form.get('pf_deduction', 0))
    esi_deduction = float(request.form.get('esi_deduction', 0))
    professional_tax = float(request.form.get('professional_tax', 0))
    other_deductions = float(request.form.get('other_deductions', 0))

    # biometric_enrolled = request.form.get('biometric_enrolled', 'false').lower() == 'true'  # Not used currently

    db = get_db()

    # Always create staff in both app and device, do not check for staff ID existence
    device_ip = '192.168.1.201'  # Default device IP
    biometric_created = False
    biometric_error = None
    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            # Always overwrite user on device, do not check existence
            biometric_created = zk_device.create_user(user_id=staff_id, name=full_name, overwrite=True)
            zk_device.disconnect()
            if not biometric_created:
                biometric_error = 'Failed to create staff on biometric device. Please check device connection.'
        else:
            biometric_error = 'Cannot connect to biometric device. Staff not added.'
    except Exception as e:
        biometric_error = f'Biometric device error: {e}'

    # Handle photo upload
    photo_url = None
    if 'photo' in request.files:
        photo = request.files['photo']
        if photo.filename != '' and allowed_file(photo.filename):
            try:
                # Ensure uploads directory exists
                upload_dir = os.path.join(app.static_folder, 'uploads')
                os.makedirs(upload_dir, exist_ok=True)

                # Generate unique filename
                ext = os.path.splitext(photo.filename)[1]
                filename = f"staff_{staff_id}_{int(time.time())}{ext}"
                # Save only the relative path for static serving
                photo_url = f"uploads/{filename}"
                photo.save(os.path.join(upload_dir, filename))
            except Exception as e:
                print(f"Error saving photo: {e}")
                return jsonify({'success': False, 'error': 'Error saving photo'})
        elif photo.filename != '':
            return jsonify({'success': False, 'error': 'Invalid file type. Only PNG, JPG, JPEG, and GIF files are allowed.'})
    # Always set a default photo if none uploaded
    if not photo_url:
        photo_url = 'images/default_profile.png'  # Make sure this file exists in static/images/

    try:
        # Ensure 'photo_url' column exists in the table
        columns = db.execute("PRAGMA table_info(staff)").fetchall()
        has_photo_url = any(col['name'] == 'photo_url' for col in columns)

        if not has_photo_url:
            db.execute("ALTER TABLE staff ADD COLUMN photo_url TEXT")

        created_at = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        db.execute('''
            INSERT INTO staff
            (school_id, staff_id, password_hash, full_name, email, phone, department, position, photo_url, created_at,
             basic_salary, hra, transport_allowance, other_allowances, pf_deduction, esi_deduction, professional_tax, other_deductions)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (school_id, staff_id, password, full_name, email, phone, department, position, photo_url, created_at,
              basic_salary, hra, transport_allowance, other_allowances, pf_deduction, esi_deduction, professional_tax, other_deductions))

        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/check_staff_id_availability', methods=['POST'])
def check_staff_id_availability():
    """Check if a staff ID is available in the current school"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id')
    school_id = session.get('school_id')

    if not staff_id:
        return jsonify({'success': False, 'error': 'Staff ID is required'})

    if not school_id:
        return jsonify({'success': False, 'error': 'School ID not found in session'})

    try:
        db = get_db()
        existing_staff = db.execute('''
            SELECT full_name, created_at FROM staff
            WHERE school_id = ? AND staff_id = ?
        ''', (school_id, staff_id)).fetchone()

        if existing_staff:
            return jsonify({
                'success': True,
                'available': False,
                'message': f'Staff ID "{staff_id}" is already used by {existing_staff["full_name"]}',
                'suggestion': 'Please choose a different Staff ID'
            })
        else:
            return jsonify({
                'success': True,
                'available': True,
                'message': f'Staff ID "{staff_id}" is available'
            })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/delete_school', methods=['POST'])
def delete_school():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = request.form.get('school_id')

    db = get_db()

    try:
        # Delete all related records first
        db.execute('DELETE FROM admins WHERE school_id = ?', (school_id,))
        db.execute('DELETE FROM staff WHERE school_id = ?', (school_id,))
        db.execute('DELETE FROM schools WHERE id = ?', (school_id,))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_staff_details_enhanced')
def get_staff_details_enhanced():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.args.get('id')
    db = get_db()

    staff = db.execute('''
        SELECT id, staff_id, full_name, first_name, last_name,
               date_of_birth, date_of_joining, department, destination,
               position, gender, phone, email, shift_type, photo_url,
               basic_salary, hra, transport_allowance, other_allowances,
               pf_deduction, esi_deduction, professional_tax, other_deductions
        FROM staff
        WHERE id = ?
    ''', (staff_id,)).fetchone()

    if not staff:
        return jsonify({'success': False, 'error': 'Staff not found'})

    return jsonify({
        'success': True,
        'staff': dict(staff)
    })

@app.route('/update_staff_enhanced', methods=['POST'])
def update_staff_enhanced():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_db_id = request.form.get('staff_db_id')
    staff_id = request.form.get('staff_id')
    first_name = request.form.get('first_name')
    last_name = request.form.get('last_name')
    date_of_birth = request.form.get('date_of_birth')
    date_of_joining = request.form.get('date_of_joining')
    department = request.form.get('department')
    destination = request.form.get('destination')
    gender = request.form.get('gender')
    phone = request.form.get('phone')
    email = request.form.get('email')
    shift_type = request.form.get('shift_type', 'general')

    # Salary fields
    basic_salary = request.form.get('basic_salary', type=float)
    hra = request.form.get('hra', type=float)
    transport_allowance = request.form.get('transport_allowance', type=float)
    other_allowances = request.form.get('other_allowances', type=float)
    pf_deduction = request.form.get('pf_deduction', type=float)
    esi_deduction = request.form.get('esi_deduction', type=float)
    professional_tax = request.form.get('professional_tax', type=float)
    other_deductions = request.form.get('other_deductions', type=float)

    school_id = session['school_id']

    # Create full_name from first_name and last_name
    full_name = f"{first_name} {last_name}".strip() if first_name or last_name else ""

    if not staff_db_id or not staff_id or not first_name or not last_name:
        return jsonify({'success': False, 'error': 'Staff ID, first name, and last name are required'})

    db = get_db()

    # Handle photo upload (if any)
    photo_url = None
    if 'photo' in request.files:
        photo = request.files['photo']
        if photo and photo.filename and allowed_file(photo.filename):
            filename = secure_filename(f"{staff_id}_{photo.filename}")
            photo_path = os.path.join(app.config.get('UPLOAD_FOLDER', 'static/uploads'), filename)
            os.makedirs(os.path.dirname(photo_path), exist_ok=True)
            photo.save(photo_path)
            photo_url = f"uploads/{filename}"

    try:
        # Build update query dynamically based on available columns
        columns = db.execute("PRAGMA table_info(staff)").fetchall()
        column_names = [col['name'] for col in columns]

        update_parts = ['staff_id = ?', 'full_name = ?']
        update_values = [staff_id, full_name]

        if 'first_name' in column_names:
            update_parts.append('first_name = ?')
            update_values.append(first_name)
        if 'last_name' in column_names:
            update_parts.append('last_name = ?')
            update_values.append(last_name)
        if 'date_of_birth' in column_names:
            update_parts.append('date_of_birth = ?')
            update_values.append(date_of_birth)
        if 'date_of_joining' in column_names:
            update_parts.append('date_of_joining = ?')
            update_values.append(date_of_joining)
        if 'department' in column_names:
            update_parts.append('department = ?')
            update_values.append(department)
        if 'destination' in column_names:
            update_parts.append('destination = ?')
            update_values.append(destination)
        if 'gender' in column_names:
            update_parts.append('gender = ?')
            update_values.append(gender)
        if 'phone' in column_names:
            update_parts.append('phone = ?')
            update_values.append(phone)
        if 'email' in column_names:
            update_parts.append('email = ?')
            update_values.append(email)
        if 'shift_type' in column_names:
            update_parts.append('shift_type = ?')
            update_values.append(shift_type)
        if 'photo_url' in column_names and photo_url:
            update_parts.append('photo_url = ?')
            update_values.append(photo_url)

        # Add salary fields
        if 'basic_salary' in column_names and basic_salary is not None:
            update_parts.append('basic_salary = ?')
            update_values.append(basic_salary)
        if 'hra' in column_names and hra is not None:
            update_parts.append('hra = ?')
            update_values.append(hra)
        if 'transport_allowance' in column_names and transport_allowance is not None:
            update_parts.append('transport_allowance = ?')
            update_values.append(transport_allowance)
        if 'other_allowances' in column_names and other_allowances is not None:
            update_parts.append('other_allowances = ?')
            update_values.append(other_allowances)
        if 'pf_deduction' in column_names and pf_deduction is not None:
            update_parts.append('pf_deduction = ?')
            update_values.append(pf_deduction)
        if 'esi_deduction' in column_names and esi_deduction is not None:
            update_parts.append('esi_deduction = ?')
            update_values.append(esi_deduction)
        if 'professional_tax' in column_names and professional_tax is not None:
            update_parts.append('professional_tax = ?')
            update_values.append(professional_tax)
        if 'other_deductions' in column_names and other_deductions is not None:
            update_parts.append('other_deductions = ?')
            update_values.append(other_deductions)

        # Add WHERE clause values
        update_values.extend([staff_db_id, school_id])

        # Build and execute the query
        query = f"UPDATE staff SET {', '.join(update_parts)} WHERE id = ? AND school_id = ?"

        db.execute(query, update_values)
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/export_staff_excel')
def export_staff_excel():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    try:
        school_id = session['school_id']
        db = get_db()

        # Get all staff with comprehensive details
        staff = db.execute('''
            SELECT staff_id, full_name, first_name, last_name,
                   date_of_birth, date_of_joining, department, destination,
                   position, gender, phone, email, shift_type, created_at
            FROM staff
            WHERE school_id = ?
            ORDER BY full_name
        ''', (school_id,)).fetchall()

        # Create proper Excel file using openpyxl
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Staff Details"

        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add title
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f"Staff Details Report - Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(bold=True, size=16, color="2F5597")
        title_cell.alignment = Alignment(horizontal='center')

        # Add headers
        headers = [
            'S.No', 'Staff ID', 'First Name', 'Last Name', 'Full Name',
            'Date of Birth', 'Date of Joining', 'Department', 'Destination/Position',
            'Gender', 'Phone Number', 'Email ID', 'Shift Type', 'Created Date'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Add data
        for row_idx, staff_member in enumerate(staff, 4):
            row_data = [
                row_idx - 3,  # S.No
                staff_member['staff_id'] or 'N/A',
                staff_member['first_name'] or 'N/A',
                staff_member['last_name'] or 'N/A',
                staff_member['full_name'] or 'N/A',
                staff_member['date_of_birth'] or 'N/A',
                staff_member['date_of_joining'] or 'N/A',
                staff_member['department'] or 'N/A',
                staff_member['destination'] or staff_member['position'] or 'N/A',
                staff_member['gender'] or 'N/A',
                staff_member['phone'] or 'N/A',
                staff_member['email'] or 'N/A',
                staff_member['shift_type'] or 'General',
                staff_member['created_at'] or 'N/A'
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(vertical='center')

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15

        # Freeze header row
        ws.freeze_panes = 'A4'

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response with proper Excel headers
        response = make_response(output.getvalue())
        filename = f'staff_details_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'

        return response

    except Exception as e:
        # Log the error and return JSON error response
        print(f"Excel export error: {str(e)}")
        return jsonify({'success': False, 'error': f'Excel export failed: {str(e)}'})

    return jsonify({'success': False, 'error': 'Unknown error occurred'})

@app.route('/admin/add_department_shift', methods=['POST'])
def add_department_shift():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    department = request.form.get('department')
    shift_type = request.form.get('shift_type')

    if not department or not shift_type:
        return jsonify({'success': False, 'error': 'Department and shift type are required'})

    db = get_db()

    try:
        db.execute('''
            INSERT OR REPLACE INTO department_shift_mappings
            (school_id, department, default_shift_type, updated_at)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP)
        ''', (school_id, department, shift_type))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/update_department_shift', methods=['POST'])
def update_department_shift():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    department = request.form.get('department')
    shift_type = request.form.get('shift_type')

    if not department or not shift_type:
        return jsonify({'success': False, 'error': 'Department and shift type are required'})

    db = get_db()

    try:
        # Check if mapping already exists
        existing = db.execute('''
            SELECT id FROM department_shift_mappings
            WHERE school_id = ? AND department = ?
        ''', (school_id, department)).fetchone()

        if existing:
            # Update existing mapping
            db.execute('''
                UPDATE department_shift_mappings
                SET default_shift_type = ?, updated_at = CURRENT_TIMESTAMP
                WHERE school_id = ? AND department = ?
            ''', (shift_type, school_id, department))
        else:
            # Insert new mapping
            db.execute('''
                INSERT INTO department_shift_mappings (school_id, department, default_shift_type, created_at, updated_at)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ''', (school_id, department, shift_type))

        # Update all staff members in this department to the new shift type
        staff_updated = db.execute('''
            UPDATE staff
            SET shift_type = ?
            WHERE school_id = ? AND department = ?
        ''', (shift_type, school_id, department))

        affected_rows = staff_updated.rowcount

        db.commit()

        message = f'Department {department} assigned to {shift_type} shift successfully. {affected_rows} staff members updated.'
        return jsonify({'success': True, 'message': message, 'affected_staff': affected_rows})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/delete_department_shift', methods=['POST'])
def delete_department_shift():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    department = request.form.get('department')

    if not department:
        return jsonify({'success': False, 'error': 'Department is required'})

    db = get_db()

    try:
        db.execute('''
            DELETE FROM department_shift_mappings
            WHERE school_id = ? AND department = ?
        ''', (school_id, department))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/bulk_update_staff_shifts', methods=['POST'])
def bulk_update_staff_shifts():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    db = get_db()

    try:
        # Get all staff with departments that have shift mappings
        staff_to_update = db.execute('''
            SELECT s.id, s.full_name, s.department, s.shift_type, dsm.default_shift_type
            FROM staff s
            JOIN department_shift_mappings dsm ON s.department = dsm.department AND s.school_id = dsm.school_id
            WHERE s.school_id = ? AND s.department IS NOT NULL AND s.department != ''
            AND (s.shift_type IS NULL OR s.shift_type != dsm.default_shift_type)
        ''', (school_id,)).fetchall()

        updated_count = 0
        for staff in staff_to_update:
            db.execute('''
                UPDATE staff
                SET shift_type = ?
                WHERE id = ?
            ''', (staff['default_shift_type'], staff['id']))
            updated_count += 1

        db.commit()
        return jsonify({'success': True, 'updated_count': updated_count})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/preview_staff_shift_changes')
def preview_staff_shift_changes():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    db = get_db()

    try:
        # Get all staff with departments that have shift mappings and would be changed
        changes = db.execute('''
            SELECT s.full_name as staff_name, s.department,
                   COALESCE(s.shift_type, 'general') as current_shift,
                   dsm.default_shift_type as new_shift
            FROM staff s
            JOIN department_shift_mappings dsm ON s.department = dsm.department AND s.school_id = dsm.school_id
            WHERE s.school_id = ? AND s.department IS NOT NULL AND s.department != ''
            AND (s.shift_type IS NULL OR s.shift_type != dsm.default_shift_type)
            ORDER BY s.department, s.full_name
        ''', (school_id,)).fetchall()

        changes_list = [dict(change) for change in changes]
        return jsonify({'success': True, 'changes': changes_list})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/update_staff', methods=['POST'])
def update_staff():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id')
    full_name = request.form.get('full_name')
    email = request.form.get('email')
    phone = request.form.get('phone')
    department = request.form.get('department')
    position = request.form.get('position')
    shift_type = request.form.get('shift_type', 'general')
    # status = request.form.get('status')  # Not used currently
    school_id = session['school_id']

    if not staff_id or not full_name:
        return jsonify({'success': False, 'error': 'Staff ID and full name are required'})

    db = get_db()

    # Handle photo upload
    photo_url = None
    if 'photo' in request.files:
        photo = request.files['photo']
        if photo.filename != '' and allowed_file(photo.filename):
            try:
                # Ensure uploads directory exists
                upload_dir = os.path.join(app.static_folder, 'uploads')
                os.makedirs(upload_dir, exist_ok=True)

                # Generate unique filename
                ext = os.path.splitext(photo.filename)[1]
                filename = f"staff_{staff_id}_{int(time.time())}{ext}"
                photo_url = os.path.join('uploads', filename)
                photo.save(os.path.join(app.static_folder, photo_url))
            except Exception as e:
                print(f"Error saving photo: {e}")
                return jsonify({'success': False, 'error': 'Error saving photo'})
        elif photo.filename != '':
            return jsonify({'success': False, 'error': 'Invalid file type. Only PNG, JPG, JPEG, and GIF files are allowed.'})

    try:
        # Update staff record
        if photo_url:
            db.execute('''
                UPDATE staff
                SET full_name = ?, email = ?, phone = ?, department = ?, position = ?, shift_type = ?, photo_url = ?
                WHERE id = ? AND school_id = ?
            ''', (full_name, email, phone, department, position, shift_type, photo_url, staff_id, school_id))
        else:
            db.execute('''
                UPDATE staff
                SET full_name = ?, email = ?, phone = ?, department = ?, position = ?, shift_type = ?
                WHERE id = ? AND school_id = ?
            ''', (full_name, email, phone, department, position, shift_type, staff_id, school_id))

        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/delete_staff', methods=['POST'])
def delete_staff():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id')
    school_id = session['school_id']

    db = get_db()

    # Get staff record to retrieve biometric user ID
    staff = db.execute('SELECT staff_id FROM staff WHERE id = ? AND school_id = ?', (staff_id, school_id)).fetchone()
    biometric_deleted = False
    biometric_error = None
    if staff:
        device_ip = '192.168.1.201'  # Default device IP, adjust if needed
        try:
            zk_device = ZKBiometricDevice(device_ip)
            if zk_device.connect():
                biometric_deleted = zk_device.delete_user(staff['staff_id'])
                zk_device.disconnect()
            else:
                biometric_error = 'Failed to connect to biometric device'
        except Exception as e:
            biometric_error = str(e)

    db.execute('DELETE FROM staff WHERE id = ? AND school_id = ?', (staff_id, school_id))
    db.commit()

    return jsonify({'success': True, 'biometric_deleted': biometric_deleted, 'biometric_error': biometric_error})

@app.route('/reset_staff_password', methods=['POST'])
def reset_staff_password():
    """Reset staff password (Admin only)"""
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_db_id = request.form.get('staff_id')  # This is the database ID
    new_password = request.form.get('new_password')
    school_id = session['school_id']

    if not staff_db_id:
        return jsonify({'success': False, 'error': 'Staff ID is required'})

    if not new_password:
        return jsonify({'success': False, 'error': 'New password is required'})

    if len(new_password) < 6:
        return jsonify({'success': False, 'error': 'Password must be at least 6 characters long'})

    db = get_db()

    try:
        # Verify staff exists in the same school
        staff = db.execute('''
            SELECT staff_id, full_name FROM staff
            WHERE id = ? AND school_id = ?
        ''', (staff_db_id, school_id)).fetchone()

        if not staff:
            return jsonify({'success': False, 'error': 'Staff not found'})

        # Generate new password hash
        password_hash = generate_password_hash(new_password)

        # Update staff password
        db.execute('''
            UPDATE staff
            SET password_hash = ?
            WHERE id = ? AND school_id = ?
        ''', (password_hash, staff_db_id, school_id))

        db.commit()

        return jsonify({
            'success': True,
            'message': f'Password reset successfully for {staff["full_name"]} (Staff ID: {staff["staff_id"]})'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/set_default_passwords', methods=['POST'])
def set_default_passwords():
    """Set default passwords for all staff without passwords (for testing)"""
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    default_password = 'password123'

    db = get_db()

    try:
        # Find staff without passwords
        staff_without_passwords = db.execute('''
            SELECT id, staff_id, full_name FROM staff
            WHERE school_id = ? AND (password_hash IS NULL OR password_hash = '')
        ''', (school_id,)).fetchall()

        if not staff_without_passwords:
            return jsonify({'success': True, 'message': 'All staff already have passwords set'})

        # Generate password hash
        password_hash = generate_password_hash(default_password)

        # Update all staff without passwords
        updated_count = 0
        for staff in staff_without_passwords:
            db.execute('''
                UPDATE staff SET password_hash = ? WHERE id = ?
            ''', (password_hash, staff['id']))
            updated_count += 1

        db.commit()

        return jsonify({
            'success': True,
            'message': f'Set default password for {updated_count} staff members. Default password: {default_password}',
            'updated_count': updated_count
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/staff/change_password', methods=['GET', 'POST'])
def staff_change_password():
    """Allow staff to change their own password"""
    if 'user_id' not in session or session['user_type'] != 'staff':
        return redirect(url_for('index'))

    if request.method == 'GET':
        return render_template('staff_change_password.html')

    # Handle POST request
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    if not all([current_password, new_password, confirm_password]):
        return jsonify({'success': False, 'error': 'All fields are required'})

    if new_password != confirm_password:
        return jsonify({'success': False, 'error': 'New passwords do not match'})

    if len(new_password) < 6:
        return jsonify({'success': False, 'error': 'Password must be at least 6 characters long'})

    db = get_db()
    staff_id = session['user_id']

    try:
        # Get current staff record
        staff = db.execute('SELECT * FROM staff WHERE id = ?', (staff_id,)).fetchone()

        if not staff:
            return jsonify({'success': False, 'error': 'Staff record not found'})

        # Verify current password
        if not check_password_hash(staff['password_hash'], current_password):
            return jsonify({'success': False, 'error': 'Current password is incorrect'})

        # Update password
        new_password_hash = generate_password_hash(new_password)
        db.execute('UPDATE staff SET password_hash = ? WHERE id = ?', (new_password_hash, staff_id))
        db.commit()

        return jsonify({'success': True, 'message': 'Password changed successfully'})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/add_school', methods=['POST'])
def add_school():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    name = request.form.get('name')
    address = request.form.get('address')
    contact_email = request.form.get('contact_email')
    contact_phone = request.form.get('contact_phone')

    admin_username = request.form.get('admin_username')
    admin_password = generate_password_hash(request.form.get('admin_password'))
    admin_full_name = request.form.get('admin_full_name')
    admin_email = request.form.get('admin_email')

    db = get_db()

    # Handle logo upload
    logo_url = None
    if 'logo' in request.files:
        logo = request.files['logo']
        if logo.filename != '' and allowed_file(logo.filename):
            try:
                upload_dir = os.path.join(app.static_folder, 'school_logos')
                os.makedirs(upload_dir, exist_ok=True)

                ext = os.path.splitext(logo.filename)[1]
                filename = f"school_{int(time.time())}{ext}"
                logo_url = os.path.join('school_logos', filename)
                logo.save(os.path.join(app.static_folder, logo_url))
            except Exception as e:
                print(f"Error saving logo: {e}")
                return jsonify({'success': False, 'error': 'Failed to save school logo'})
        elif logo.filename != '':
            return jsonify({'success': False, 'error': 'Invalid file type. Only PNG, JPG, JPEG, and GIF files are allowed.'})

    try:
        cursor = db.cursor()

        # Add school
        cursor.execute('''
            INSERT INTO schools (name, address, contact_email, contact_phone, logo_url)
            VALUES (?, ?, ?, ?, ?)
        ''', (name, address, contact_email, contact_phone, logo_url))
        school_id = cursor.lastrowid

        # Add initial admin for the school
        cursor.execute('''
            INSERT INTO admins (school_id, username, password, full_name, email)
            VALUES (?, ?, ?, ?, ?)
        ''', (school_id, admin_username, admin_password, admin_full_name, admin_email))

        db.commit()
        return jsonify({'success': True})

    except sqlite3.IntegrityError:
        db.rollback()
        return jsonify({'success': False, 'error': 'School or admin username already exists'})

    except Exception as e:
        db.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/staff/<int:id>')
def staff_profile(id):
    if 'user_id' not in session or session['user_type'] != 'admin':
        return redirect(url_for('index'))

    db = get_db()
    staff = db.execute('SELECT * FROM staff WHERE id = ? AND school_id = ?',
                      (id, session['school_id'])).fetchone()

    if not staff:
        return redirect(url_for('admin_dashboard'))

    # Get attendance summary for this staff member
    attendance_summary = db.execute('''
        SELECT
            COUNT(CASE WHEN status = 'present' THEN 1 END) as present_count,
            COUNT(CASE WHEN status = 'late' THEN 1 END) as late_count,
            COUNT(CASE WHEN status = 'absent' THEN 1 END) as absent_count,
            COUNT(CASE WHEN status = 'leave' THEN 1 END) as leave_count
        FROM attendance
        WHERE staff_id = ? AND date >= date('now', '-30 days')
    ''', (id,)).fetchone()



    # Get recent biometric verifications (latest per day per type)
    recent_verifications = db.execute('''
        SELECT verification_type, verification_time, biometric_method, verification_status
        FROM biometric_verifications bv1
        WHERE staff_id = ?
          AND verification_time = (
            SELECT MAX(verification_time)
            FROM biometric_verifications bv2
            WHERE bv2.staff_id = bv1.staff_id
              AND bv2.verification_type = bv1.verification_type
              AND DATE(bv2.verification_time) = DATE(bv1.verification_time)
          )
        ORDER BY verification_time DESC
        LIMIT 20
    ''', (id,)).fetchall()

    return render_template('staff_profile.html',
                         staff=staff,
                         attendance_summary=attendance_summary,
                         recent_verifications=recent_verifications)

@app.route('/admin/search_staff')
def search_staff():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    search_term = request.args.get('q', '')
    db = get_db()

    staff = db.execute('''
        SELECT id, staff_id, full_name, department, position
        FROM staff
        WHERE school_id = ? AND full_name LIKE ?
        ORDER BY full_name
    ''', (session['school_id'], f"%{search_term}%")).fetchall()

    # Get pending leave applications
    pending_leaves = db.execute('''
        SELECT l.id, s.full_name, l.leave_type, l.start_date, l.end_date, l.reason
        FROM leave_applications l
        JOIN staff s ON l.staff_id = s.id
        WHERE l.school_id = ? AND l.status = 'pending'
        ORDER BY l.applied_at
    ''', (session['school_id'],)).fetchall()

    # Get today's attendance summary
    today = datetime.date.today()
    attendance_summary = db.execute('''
        SELECT
            COUNT(*) as total_staff,
            SUM(CASE WHEN a.status = 'present' THEN 1 ELSE 0 END) as present,
            SUM(CASE WHEN a.status = 'absent' THEN 1 ELSE 0 END) as absent,
            SUM(CASE WHEN a.status = 'late' THEN 1 ELSE 0 END) as late,
            SUM(CASE WHEN a.status = 'leave' THEN 1 ELSE 0 END) as on_leave
        FROM (
            SELECT s.id, COALESCE(a.status, 'absent') as status
            FROM staff s
            LEFT JOIN attendance a ON s.id = a.staff_id AND a.date = ?
            WHERE s.school_id = ?
        ) a
    ''', (today, session['school_id'])).fetchone()

    return render_template('admin_dashboard.html',
                         staff=staff,
                         pending_leaves=pending_leaves,
                         attendance_summary=attendance_summary,
                         today=today)

# Update in app.py
@app.route('/toggle_school_visibility', methods=['POST'])
def toggle_school_visibility():
    if 'user_id' not in session or session['user_type'] != 'company_admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = request.form.get('school_id')
    db = get_db()

    # Ensure column exists
    columns = db.execute("PRAGMA table_info(schools)").fetchall()
    has_is_hidden = any(col['name'] == 'is_hidden' for col in columns)

    if not has_is_hidden:
        db.execute('ALTER TABLE schools ADD COLUMN is_hidden BOOLEAN DEFAULT 0')
        db.commit()

    # Toggle visibility
    db.execute('''
        UPDATE schools
        SET is_hidden = CASE WHEN is_hidden = 1 THEN 0 ELSE 1 END
        WHERE id = ?
    ''', (school_id,))
    db.commit()

    return jsonify({'success': True})




# ZK Biometric Device Integration Routes
@app.route('/sync_biometric_attendance', methods=['POST'])
def sync_biometric_attendance():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    device_ip = request.form.get('device_ip', '192.168.1.201')

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({'success': False, 'error': 'Failed to connect to biometric device'})

        # Use the correct method name
        attendance_records = zk_device.get_attendance_records()
        if not attendance_records:
            zk_device.disconnect()
            return jsonify({'success': False, 'error': 'No attendance records found'})

        db = get_db()
        synced_count = 0

        for record in attendance_records:
            try:
                # Get staff database ID from staff_id (biometric user_id)
                staff_record = db.execute('''
                    SELECT id FROM staff WHERE staff_id = ? AND school_id = ?
                ''', (str(record['user_id']), school_id)).fetchone()

                if not staff_record:
                    print(f"Staff with ID {record['user_id']} not found in database")
                    continue

                staff_db_id = staff_record['id']
                attendance_date = record['timestamp'].date()
                attendance_time = record['timestamp'].strftime('%H:%M:%S')
                verification_type = record.get('verification_type', 'check-in')

                # Check if attendance record exists for this date
                existing_record = db.execute('''
                    SELECT * FROM attendance WHERE staff_id = ? AND date = ?
                ''', (staff_db_id, attendance_date)).fetchone()

                if existing_record:
                    # Update existing record based on verification type
                    if verification_type == 'check-in' and not existing_record['time_in']:
                        db.execute('''
                            UPDATE attendance SET time_in = ?, status = 'present'
                            WHERE staff_id = ? AND date = ?
                        ''', (attendance_time, staff_db_id, attendance_date))
                        synced_count += 1
                    elif verification_type == 'check-out' and not existing_record['time_out']:
                        db.execute('''
                            UPDATE attendance SET time_out = ?
                            WHERE staff_id = ? AND date = ?
                        ''', (attendance_time, staff_db_id, attendance_date))
                        synced_count += 1
                else:
                    # Create new attendance record
                    if verification_type == 'check-in':
                        db.execute('''
                            INSERT INTO attendance (staff_id, school_id, date, time_in, status)
                            VALUES (?, ?, ?, ?, 'present')
                        ''', (staff_db_id, school_id, attendance_date, attendance_time))
                        synced_count += 1

            except Exception as record_error:
                print(f"Error processing record {record}: {record_error}")
                continue

        db.commit()
        zk_device.disconnect()

        return jsonify({
            'success': True,
            'message': f'Successfully synced {synced_count} attendance records',
            'synced_count': synced_count,
            'total_records': len(attendance_records)
        })

    except Exception as e:
        return jsonify({'success': False, 'error': f'Sync failed: {str(e)}'})

@app.route('/process_device_attendance', methods=['POST'])
def process_device_attendance():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    device_ip = request.form.get('device_ip', '192.168.1.201')

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({'success': False, 'error': 'Failed to connect to biometric device'})

        process_device_attendance_automatically(zk_device, school_id, db=get_db())
        return jsonify({'success': True, 'message': 'Device attendance processed successfully'})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/verify_staff_biometric', methods=['POST'])
def verify_staff_biometric_route():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id')
    device_ip = request.form.get('device_ip', '192.168.1.201')

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({'success': False, 'error': 'Failed to connect to biometric device'})

        is_valid = verify_staff_biometric(zk_device, staff_id)
        return jsonify({'success': True, 'is_valid': is_valid})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Duplicate route removed - already defined above

@app.route('/apply_on_duty_permission', methods=['POST'])
def apply_on_duty_permission():
    if 'user_id' not in session or session['user_type'] != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    school_id = session['school_id']
    permission_type = request.form.get('permission_type', 'On Duty')
    start_datetime = request.form.get('start_datetime')
    end_datetime = request.form.get('end_datetime')
    reason = request.form.get('reason')

    if not start_datetime or not end_datetime:
        return jsonify({'success': False, 'error': 'Start and end datetime required'})

    db = get_db()
    try:
        db.execute('''
            INSERT INTO on_duty_permissions
            (staff_id, school_id, permission_type, start_datetime, end_datetime, reason)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (staff_id, school_id, permission_type, start_datetime, end_datetime, reason))
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Duplicate route removed - already defined above

@app.route('/test_biometric_connection', methods=['POST'])
def test_biometric_connection():
    """Test connection to ZK biometric device (Ethernet or Cloud)"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip = request.form.get('device_ip', '192.168.1.201')
    device_id = request.form.get('device_id', f"ZK_{device_ip.replace('.', '_')}")
    use_cloud = request.form.get('use_cloud', 'false').lower() == 'true'

    try:
        # Determine connection mode
        if CLOUD_ENABLED and use_cloud:
            # Test cloud connection
            device_config = get_device_config(device_id)
            if not device_config:
                return jsonify({
                    'success': False,
                    'message': f'Device {device_id} not configured for cloud connectivity',
                    'device_id': device_id
                })

            connector = get_cloud_connector()
            if not connector.running:
                return jsonify({
                    'success': False,
                    'message': 'Cloud connector is not running',
                    'device_id': device_id
                })

            status = connector.get_device_status(device_id)
            if status['status'] == 'connected':
                return jsonify({
                    'success': True,
                    'message': 'Cloud connection successful',
                    'device_id': device_id,
                    'device_ip': device_config.local_ip,
                    'total_users': status.get('user_count', 0),
                    'connection_type': 'cloud'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': f'Device status: {status["status"]}',
                    'device_id': device_id,
                    'connection_type': 'cloud'
                })
        else:
            # Test Ethernet connection
            port = 4370 if device_ip == '192.168.1.201' else 32150
            zk_device = ZKBiometricDevice(device_ip, port=port, timeout=15, device_id='1', use_cloud=False)
            if zk_device.connect():
                # Get device info
                users = zk_device.get_users()
                zk_device.disconnect()

                return jsonify({
                    'success': True,
                    'message': 'Ethernet connection successful',
                    'device_ip': device_ip,
                    'total_users': len(users),
                    'connection_type': 'ethernet'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Failed to connect to device via Ethernet',
                    'device_ip': device_ip,
                    'connection_type': 'ethernet'
                })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Connection test failed: {str(e)}',
            'device_ip': device_ip,
            'device_id': device_id
        })

@app.route('/cloud_config', methods=['GET', 'POST'])
def cloud_config():
    """Manage cloud configuration"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    if not CLOUD_ENABLED:
        return jsonify({'success': False, 'error': 'Cloud functionality not available'})

    if request.method == 'GET':
        try:
            config = get_cloud_config()
            connector = get_cloud_connector()

            # Return safe configuration (no sensitive data)
            safe_config = {
                'cloud_provider': config.cloud_provider,
                'api_base_url': config.api_base_url,
                'websocket_url': config.websocket_url,
                'mqtt_broker': config.mqtt_broker,
                'mqtt_port': config.mqtt_port,
                'organization_id': config.organization_id,
                'connection_timeout': config.connection_timeout,
                'retry_attempts': config.retry_attempts,
                'heartbeat_interval': config.heartbeat_interval,
                'use_ssl': config.use_ssl,
                'verify_ssl': config.verify_ssl,
                'encryption_enabled': config.encryption_enabled,
                'auto_sync': config.auto_sync,
                'sync_interval': config.sync_interval,
                'batch_size': config.batch_size,
                'local_backup': config.local_backup,
                'backup_retention_days': config.backup_retention_days,
                'connector_running': connector.running if connector else False,
                'has_api_key': bool(config.api_key),
                'has_secret_key': bool(config.secret_key)
            }

            return jsonify({
                'success': True,
                'config': safe_config
            })

        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})

    elif request.method == 'POST':
        try:
            # Update configuration
            data = request.get_json()
            if not data:
                return jsonify({'success': False, 'error': 'No data provided'})

            from cloud_config import config_manager

            # Update configuration fields
            config = config_manager.config

            # Update safe fields only
            safe_fields = [
                'cloud_provider', 'api_base_url', 'websocket_url', 'mqtt_broker', 'mqtt_port',
                'organization_id', 'connection_timeout', 'retry_attempts', 'heartbeat_interval',
                'use_ssl', 'verify_ssl', 'encryption_enabled', 'auto_sync', 'sync_interval',
                'batch_size', 'local_backup', 'backup_retention_days'
            ]

            for field in safe_fields:
                if field in data:
                    setattr(config, field, data[field])

            # Handle sensitive fields separately
            if 'api_key' in data and data['api_key']:
                config.api_key = data['api_key']

            if 'secret_key' in data and data['secret_key']:
                config.secret_key = data['secret_key']

            # Save configuration
            config_manager.save_config()

            return jsonify({
                'success': True,
                'message': 'Configuration updated successfully'
            })

        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})

@app.route('/cloud_status', methods=['GET'])
def cloud_status():
    """Get cloud connector status"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    if not CLOUD_ENABLED:
        return jsonify({
            'success': True,
            'cloud_enabled': False,
            'message': 'Cloud functionality not available'
        })

    try:
        connector = get_cloud_connector()
        config = get_cloud_config()

        from cloud_config import get_all_devices
        devices = get_all_devices()

        # Get device statuses
        device_statuses = []
        for device in devices:
            if device.cloud_enabled:
                status = connector.get_device_status(device.device_id)
                device_statuses.append({
                    'device_id': device.device_id,
                    'device_name': device.device_name,
                    'local_ip': device.local_ip,
                    'status': status['status'],
                    'last_sync': device.last_sync,
                    'user_count': status.get('user_count', 0)
                })

        return jsonify({
            'success': True,
            'cloud_enabled': True,
            'connector_running': connector.running,
            'websocket_connected': connector.websocket is not None and
                                 connector.websocket.sock and
                                 connector.websocket.sock.connected,
            'last_heartbeat': connector.last_heartbeat.isoformat() if connector.last_heartbeat else None,
            'message_queue_size': len(connector.message_queue),
            'device_count': len(device_statuses),
            'devices': device_statuses,
            'config_valid': bool(config.api_key and config.api_base_url and config.organization_id)
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/cloud_dashboard')
def cloud_dashboard():
    """Cloud monitoring dashboard"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return redirect(url_for('login'))

    if not CLOUD_ENABLED:
        return render_template('error.html',
                             error_message="Cloud functionality not available",
                             error_details="Please install cloud dependencies to access this feature.")

    return render_template('cloud_dashboard.html')

@app.route('/get_biometric_users', methods=['GET'])
def get_biometric_users():
    """Get users from ZK biometric device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip = request.args.get('device_ip', '192.168.1.201')

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            users = zk_device.get_users()
            zk_device.disconnect()

            return jsonify({
                'success': True,
                'users': users,
                'total_users': len(users)
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to get users: {str(e)}'
        })

@app.route('/enroll_biometric_user', methods=['POST'])
def enroll_biometric_user():
    """Enroll a user in the ZK biometric device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip = request.form.get('device_ip', '192.168.1.201')
    user_id = request.form.get('user_id')
    name = request.form.get('name')
    privilege = int(request.form.get('privilege', 0))
    overwrite = request.form.get('overwrite', 'false').lower() == 'true'

    if not user_id or not name:
        return jsonify({'success': False, 'message': 'User ID and name are required'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            # First check if user already exists
            users = zk_device.get_users()
            existing_user = None
            for user in users:
                # Handle both dict and object formats
                user_id_value = user.get('user_id') if isinstance(user, dict) else getattr(user, 'user_id', None)
                if str(user_id_value) == str(user_id):
                    existing_user = user
                    break

            if existing_user and not overwrite:
                zk_device.disconnect()
                return jsonify({
                    'success': False,
                    'message': f'User ID {user_id} already exists on biometric device',
                    'user_exists': True,
                    'existing_user': {
                        'user_id': existing_user.get('user_id') if isinstance(existing_user, dict) else getattr(existing_user, 'user_id', 'Unknown'),
                        'name': existing_user.get('name') if isinstance(existing_user, dict) else getattr(existing_user, 'name', 'Unknown'),
                        'privilege': existing_user.get('privilege') if isinstance(existing_user, dict) else getattr(existing_user, 'privilege', 0)
                    },
                    'suggestion': 'You can either:\n1. Use a different User ID\n2. Enable "Overwrite existing user" option\n3. Create staff account for the existing biometric user'
                })

            result = zk_device.enroll_user(user_id, name, privilege, overwrite=overwrite)
            zk_device.disconnect()

            if result['success']:
                return jsonify({
                    'success': True,
                    'message': result['message'],
                    'action': result.get('action', 'enrolled'),
                    'user_exists': result.get('user_exists', False)
                })
            else:
                return jsonify({
                    'success': False,
                    'message': result['message'],
                    'user_exists': result.get('user_exists', False),
                    'existing_user': result.get('existing_user')
                })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Enrollment failed: {str(e)}'
        })

@app.route('/check_biometric_user', methods=['POST'])
def check_biometric_user():
    """Check if a user already exists on the ZK biometric device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip = request.form.get('device_ip', '192.168.1.201')
    user_id = request.form.get('user_id')

    if not user_id:
        return jsonify({'success': False, 'message': 'User ID is required'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            users = zk_device.get_users()
            zk_device.disconnect()

            # Check if user exists
            for user in users:
                if user['user_id'] == user_id:
                    return jsonify({
                        'success': True,
                        'user_exists': True,
                        'user_data': {
                            'user_id': user['user_id'],
                            'name': user['name'],
                            'privilege': user['privilege']
                        }
                    })

            return jsonify({
                'success': True,
                'user_exists': False,
                'message': f'User {user_id} not found on device'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Check failed: {str(e)}'
        })

@app.route('/start_biometric_enrollment', methods=['POST'])
def start_biometric_enrollment():
    """Start biometric enrollment mode on device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip = request.form.get('device_ip', '192.168.1.201')

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            success = zk_device.start_enrollment_mode()
            # Don't disconnect here - keep connection for enrollment

            if success:
                return jsonify({
                    'success': True,
                    'message': 'Device ready for biometric enrollment'
                })
            else:
                zk_device.disconnect()
                return jsonify({
                    'success': False,
                    'message': 'Failed to start enrollment mode'
                })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to start enrollment: {str(e)}'
        })

@app.route('/end_biometric_enrollment', methods=['POST'])
def end_biometric_enrollment():
    """End biometric enrollment mode on device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip = request.form.get('device_ip', '192.168.1.201')

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            success = zk_device.end_enrollment_mode()
            zk_device.disconnect()

            if success:
                return jsonify({
                    'success': True,
                    'message': 'Enrollment mode ended'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Failed to end enrollment mode'
                })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to end enrollment: {str(e)}'
        })

@app.route('/verify_biometric_enrollment', methods=['POST'])
def verify_biometric_enrollment():
    """Verify that biometric data was captured for a user"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip = request.form.get('device_ip', '192.168.1.201')
    user_id = request.form.get('user_id')
    trigger_enrollment = request.form.get('trigger_enrollment', 'false').lower() == 'true'

    if not user_id:
        return jsonify({'success': False, 'message': 'User ID is required'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            # First check if user exists
            users = zk_device.get_users()
            user_exists = False
            user_data = None

            for user in users:
                if user['user_id'] == user_id:
                    user_exists = True
                    user_data = user
                    break

            # If user doesn't exist or we need to trigger enrollment
            if trigger_enrollment and user_exists:
                # Trigger biometric enrollment for the user
                result = zk_device.trigger_biometric_enrollment(user_id)

                if result['success']:
                    return jsonify({
                        'success': True,
                        'enrolled': False,
                        'enrollment_started': True,
                        'manual_mode': result.get('manual_mode', True),
                        'message': result['message']
                    })
                else:
                    zk_device.disconnect()
                    return jsonify({
                        'success': False,
                        'message': result['message']
                    })

            # Check if user exists and has biometric data
            if user_exists:
                zk_device.disconnect()
                return jsonify({
                    'success': True,
                    'enrolled': True,
                    'user_data': user_data,
                    'message': f'User {user_id} biometric data verified'
                })
            else:
                # User not found, try to create and enroll
                if trigger_enrollment:
                    # First create the user
                    name = request.form.get('name', f'User {user_id}')
                    privilege = int(request.form.get('privilege', 0))

                    # Create user first
                    enroll_result = zk_device.enroll_user(user_id, name, privilege)

                    if enroll_result['success']:
                        # Now trigger biometric enrollment
                        result = zk_device.trigger_biometric_enrollment(user_id)

                        if result['success']:
                            return jsonify({
                                'success': True,
                                'enrolled': False,
                                'user_created': True,
                                'enrollment_started': True,
                                'manual_mode': result.get('manual_mode', True),
                                'message': f'User created and {result["message"]}'
                            })

                    zk_device.disconnect()
                    return jsonify({
                        'success': False,
                        'message': f'Failed to create user: {enroll_result.get("message", "Unknown error")}'
                    })

                zk_device.disconnect()
                return jsonify({
                    'success': True,
                    'enrolled': False,
                    'message': f'User {user_id} not found or biometric data not captured'
                })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Verification failed: {str(e)}'
        })

@app.route('/delete_biometric_user', methods=['POST'])
def delete_biometric_user():
    """Delete a user from the ZK biometric device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip = request.form.get('device_ip', '192.168.1.201')
    user_id = request.form.get('user_id')

    if not user_id:
        return jsonify({'success': False, 'message': 'User ID is required'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if zk_device.connect():
            success = zk_device.delete_user(user_id)
            zk_device.disconnect()

            if success:
                return jsonify({
                    'success': True,
                    'message': f'User {user_id} deleted successfully from biometric device'
                })
            else:
                return jsonify({
                    'success': False,
                    'message': f'User {user_id} not found on device or deletion failed'
                })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to connect to biometric device'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Deletion failed: {str(e)}'
        })

@app.route('/resolve_biometric_conflict', methods=['POST'])
def resolve_biometric_conflict():
    """Resolve biometric user conflicts with multiple resolution options"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip = request.form.get('device_ip', '192.168.1.201')
    user_id = request.form.get('user_id')
    action = request.form.get('action')  # 'overwrite', 'delete', 'check'
    new_name = request.form.get('new_name', '')

    if not user_id:
        return jsonify({'success': False, 'message': 'User ID is required'})

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({'success': False, 'message': 'Failed to connect to biometric device'})

        if action == 'check':
            # Just check if user exists and return details
            users = zk_device.get_users()
            for user in users:
                if str(user['user_id']) == str(user_id):
                    zk_device.disconnect()
                    return jsonify({
                        'success': True,
                        'user_exists': True,
                        'user_data': {
                            'user_id': user['user_id'],
                            'name': user['name'],
                            'privilege': user['privilege'],
                            'uid': user['uid']
                        }
                    })

            zk_device.disconnect()
            return jsonify({
                'success': True,
                'user_exists': False,
                'message': f'User {user_id} not found on device'
            })

        elif action == 'delete':
            # Delete the existing user
            success = zk_device.delete_user(user_id)
            zk_device.disconnect()

            return jsonify({
                'success': success,
                'message': f'User {user_id} {"deleted successfully" if success else "deletion failed"}'
            })

        elif action == 'overwrite':
            # Overwrite the existing user
            if not new_name:
                zk_device.disconnect()
                return jsonify({'success': False, 'message': 'New name is required for overwrite'})

            result = zk_device.enroll_user(user_id, new_name, overwrite=True)
            zk_device.disconnect()

            return jsonify(result)

        else:
            zk_device.disconnect()
            return jsonify({'success': False, 'message': 'Invalid action specified'})

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Conflict resolution failed: {str(e)}'
        })

@app.route('/poll_device_attendance', methods=['POST'])
def poll_device_attendance():
    """Poll ZK device for new attendance records and process them automatically"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip = request.form.get('device_ip', '192.168.1.201')
    school_id = session.get('school_id', 1)

    try:
        result = process_device_attendance_automatically(device_ip, school_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Polling failed: {str(e)}',
            'processed_count': 0
        })

@app.route('/get_latest_device_verifications')
def get_latest_device_verifications():
    """Get the latest biometric verifications from the device for real-time updates"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    device_ip = request.args.get('device_ip', '192.168.1.201')
    since_minutes = int(request.args.get('since_minutes', 5))  # Default to last 5 minutes

    try:
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({
                'success': False,
                'error': 'Failed to connect to biometric device',
                'verifications': []
            })

        # Get records from the last few minutes
        since_timestamp = datetime.datetime.now() - datetime.timedelta(minutes=since_minutes)
        recent_records = zk_device.get_new_attendance_records(since_timestamp)

        zk_device.disconnect()

        # Format the records for the frontend
        verifications = []
        for record in recent_records:
            verifications.append({
                'user_id': record['user_id'],
                'verification_type': record['verification_type'],
                'timestamp': record['timestamp'].strftime('%Y-%m-%d %I:%M %p'),
                'time_only': record['timestamp'].strftime('%I:%M %p')
            })

        return jsonify({
            'success': True,
            'verifications': verifications,
            'count': len(verifications)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Failed to get latest verifications: {str(e)}',
            'verifications': []
        })

@app.route('/get_comprehensive_staff_profile')
def get_comprehensive_staff_profile():
    """Get comprehensive staff profile data for admin dashboard modal"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.args.get('id')
    if not staff_id:
        return jsonify({'success': False, 'error': 'Staff ID required'})

    db = get_db()

    try:
        # Get staff information
        staff = db.execute('''
            SELECT s.*, sc.name as school_name
            FROM staff s
            LEFT JOIN schools sc ON s.school_id = sc.id
            WHERE s.id = ?
        ''', (staff_id,)).fetchone()

        if not staff:
            return jsonify({'success': False, 'error': 'Staff not found'})

        # Get attendance records (last 30 days)
        thirty_days_ago = (datetime.datetime.now() - datetime.timedelta(days=30)).date()
        attendance = db.execute('''
            SELECT date, time_in, time_out, status,
                   on_duty_type, on_duty_location, on_duty_purpose
            FROM attendance
            WHERE staff_id = ? AND date >= ?
            ORDER BY date DESC
        ''', (staff_id, thirty_days_ago)).fetchall()

        # Get biometric verifications (last 30 days)
        verifications = db.execute('''
            SELECT verification_type, verification_time, verification_status, device_ip
            FROM biometric_verifications
            WHERE staff_id = ? AND DATE(verification_time) >= ?
            ORDER BY verification_time DESC
            LIMIT 50
        ''', (staff_id, thirty_days_ago)).fetchall()

        # Get leave applications
        leaves = db.execute('''
            SELECT leave_type, start_date, end_date, reason, status, applied_at
            FROM leave_applications
            WHERE staff_id = ?
            ORDER BY applied_at DESC
            LIMIT 20
        ''', (staff_id,)).fetchall()

        # Get on-duty applications
        on_duty_applications = db.execute('''
            SELECT duty_type, start_date, end_date, start_time, end_time, location, purpose, reason, status, applied_at, admin_remarks
            FROM on_duty_applications
            WHERE staff_id = ?
            ORDER BY applied_at DESC
            LIMIT 20
        ''', (staff_id,)).fetchall()

        # Get permission applications
        permission_applications = db.execute('''
            SELECT permission_type, permission_date, start_time, end_time, duration_hours, reason, status, applied_at, admin_remarks
            FROM permission_applications
            WHERE staff_id = ?
            ORDER BY applied_at DESC
            LIMIT 20
        ''', (staff_id,)).fetchall()

        # Calculate attendance statistics (excluding holidays from total days)
        total_days = len(attendance)
        present_days = len([a for a in attendance if a['status'] in ['present', 'late', 'on_duty']])
        absent_days = len([a for a in attendance if a['status'] == 'absent'])
        late_days = len([a for a in attendance if a['status'] == 'late'])
        on_duty_days = len([a for a in attendance if a['status'] == 'on_duty'])
        holiday_days = len([a for a in attendance if a['status'] == 'holiday'])

        # Calculate working days (excluding holidays)
        working_days = total_days - holiday_days

        attendance_stats = {
            'total_days': total_days,
            'working_days': working_days,
            'present_days': present_days,
            'absent_days': absent_days,
            'late_days': late_days,
            'on_duty_days': on_duty_days,
            'holiday_days': holiday_days,
            'attendance_rate': round((present_days / working_days * 100) if working_days > 0 else 0, 1)
        }

        # Format attendance times to 12-hour format
        formatted_attendance = [format_attendance_times_to_12hr(dict(a)) for a in attendance]

        return jsonify({
            'success': True,
            'staff': dict(staff),
            'attendance': formatted_attendance,
            'verifications': [dict(v) for v in verifications],
            'leaves': [dict(l) for l in leaves],
            'on_duty_applications': [dict(od) for od in on_duty_applications],
            'permission_applications': [dict(p) for p in permission_applications],
            'attendance_stats': attendance_stats
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Failed to get staff profile: {str(e)}'
        })

@app.route('/staff/profile')
def staff_profile_page():
    """Staff profile page with personal information and attendance history"""
    if 'user_id' not in session or session['user_type'] != 'staff':
        return redirect(url_for('index'))

    db = get_db()
    staff_id = session['user_id']

    # Get staff information
    staff = db.execute('''
        SELECT * FROM staff WHERE id = ?
    ''', (staff_id,)).fetchone()

    if not staff:
        return redirect(url_for('index'))

    # Get attendance summary for current month
    today = datetime.date.today()
    first_day = today.replace(day=1)
    last_day = (today.replace(day=28) + datetime.timedelta(days=4)).replace(day=1) - datetime.timedelta(days=1)

    attendance_summary = db.execute('''
        SELECT
            COUNT(*) as total_days,
            SUM(CASE WHEN status = 'present' THEN 1 ELSE 0 END) as present_days,
            SUM(CASE WHEN status = 'absent' THEN 1 ELSE 0 END) as absent_days,
            SUM(CASE WHEN status = 'late' THEN 1 ELSE 0 END) as late_days,
            SUM(CASE WHEN status = 'leave' THEN 1 ELSE 0 END) as leave_days,
            SUM(CASE WHEN status = 'on_duty' THEN 1 ELSE 0 END) as on_duty_days
        FROM attendance
        WHERE staff_id = ? AND date BETWEEN ? AND ?
    ''', (staff_id, first_day, last_day)).fetchone()



    # Get leave applications
    leave_applications = db.execute('''
        SELECT id, leave_type, start_date, end_date, reason, status, applied_at
        FROM leave_applications
        WHERE staff_id = ?
        ORDER BY applied_at DESC
        LIMIT 10
    ''', (staff_id,)).fetchall()

    # Get recent biometric verifications (latest per day per type)
    recent_verifications = db.execute('''
        SELECT verification_type, verification_time, biometric_method, verification_status
        FROM biometric_verifications bv1
        WHERE staff_id = ?
          AND verification_time = (
            SELECT MAX(verification_time)
            FROM biometric_verifications bv2
            WHERE bv2.staff_id = bv1.staff_id
              AND bv2.verification_type = bv1.verification_type
              AND DATE(bv2.verification_time) = DATE(bv1.verification_time)
          )
        ORDER BY verification_time DESC
        LIMIT 20
    ''', (staff_id,)).fetchall()

    return render_template('staff_my_profile.html',
                         staff=staff,
                         attendance_summary=attendance_summary,
                         leave_applications=leave_applications,
                         recent_verifications=recent_verifications,
                         today=today,
                         current_month=today.strftime('%B %Y'))

@app.route('/staff/update_profile', methods=['POST'])
def update_staff_profile():
    """Update staff profile information"""
    if 'user_id' not in session or session['user_type'] != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    first_name = request.form.get('first_name')
    last_name = request.form.get('last_name')
    date_of_birth = request.form.get('date_of_birth')
    gender = request.form.get('gender')
    email = request.form.get('email')
    phone = request.form.get('phone')

    db = get_db()

    try:
        # Handle photo upload
        photo_url = None
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo.filename != '' and allowed_file(photo.filename):
                try:
                    # Ensure uploads directory exists
                    upload_dir = os.path.join(app.static_folder, 'uploads')
                    os.makedirs(upload_dir, exist_ok=True)

                    # Generate unique filename
                    ext = os.path.splitext(photo.filename)[1]
                    filename = f"staff_{staff_id}_{int(time.time())}{ext}"
                    photo.save(os.path.join(upload_dir, filename))
                    photo_url = f"uploads/{filename}"
                except Exception as e:
                    return jsonify({'success': False, 'error': 'Error saving photo'})

        # Build update query dynamically based on available columns
        columns = db.execute("PRAGMA table_info(staff)").fetchall()
        column_names = [col['name'] for col in columns]

        update_parts = []
        update_values = []

        if 'first_name' in column_names and first_name:
            update_parts.append('first_name = ?')
            update_values.append(first_name)
        if 'last_name' in column_names and last_name:
            update_parts.append('last_name = ?')
            update_values.append(last_name)
        if 'date_of_birth' in column_names and date_of_birth:
            update_parts.append('date_of_birth = ?')
            update_values.append(date_of_birth)
        if 'gender' in column_names and gender:
            update_parts.append('gender = ?')
            update_values.append(gender)
        if 'email' in column_names:
            update_parts.append('email = ?')
            update_values.append(email)
        if 'phone' in column_names:
            update_parts.append('phone = ?')
            update_values.append(phone)
        if 'photo_url' in column_names and photo_url:
            update_parts.append('photo_url = ?')
            update_values.append(photo_url)

        # Update full_name if first_name or last_name changed
        if first_name or last_name:
            # Get current names if only one is being updated
            current_staff = db.execute('SELECT first_name, last_name FROM staff WHERE id = ?', (staff_id,)).fetchone()
            current_first = current_staff['first_name'] if current_staff else ''
            current_last = current_staff['last_name'] if current_staff else ''

            new_first = first_name if first_name else current_first
            new_last = last_name if last_name else current_last
            full_name = f"{new_first} {new_last}".strip()

            update_parts.append('full_name = ?')
            update_values.append(full_name)

        if update_parts:
            # Add WHERE clause value
            update_values.append(staff_id)

            # Build and execute the query
            query = f"UPDATE staff SET {', '.join(update_parts)} WHERE id = ?"

            db.execute(query, update_values)

        db.commit()
        return jsonify({'success': True, 'message': 'Profile updated successfully'})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/staff/change_password', methods=['POST'])
def change_staff_password():
    """Change staff password"""
    if 'user_id' not in session or session['user_type'] != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    if new_password != confirm_password:
        return jsonify({'success': False, 'error': 'New passwords do not match'})

    db = get_db()

    # Get current password hash
    staff = db.execute('SELECT password_hash FROM staff WHERE id = ?', (staff_id,)).fetchone()

    if not staff or not check_password_hash(staff['password_hash'], current_password):
        return jsonify({'success': False, 'error': 'Current password is incorrect'})

    try:
        # Update password
        new_password_hash = generate_password_hash(new_password)
        db.execute('''
            UPDATE staff SET password_hash = ?
            WHERE id = ?
        ''', (new_password_hash, staff_id))

        db.commit()
        return jsonify({'success': True, 'message': 'Password changed successfully'})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/staff/attendance_calendar')
def staff_attendance_calendar():
    """Get attendance data for calendar view"""
    if 'user_id' not in session or session['user_type'] != 'staff':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = session['user_id']
    start_date = request.args.get('start')
    end_date = request.args.get('end')

    db = get_db()

    # Get attendance records for the date range with enhanced information
    attendance = db.execute('''
        SELECT a.date, a.time_in, a.time_out, a.status, a.notes,
               s.shift_type, a.late_duration_minutes, a.early_departure_minutes,
               a.shift_start_time, a.shift_end_time
        FROM attendance a
        JOIN staff s ON a.staff_id = s.id
        WHERE a.staff_id = ? AND a.date BETWEEN ? AND ?
        ORDER BY a.date
    ''', (staff_id, start_date, end_date)).fetchall()

    # Get leave applications for the date range
    leaves = db.execute('''
        SELECT start_date, end_date, leave_type, status
        FROM leave_applications
        WHERE staff_id = ? AND status = 'approved'
        AND ((start_date BETWEEN ? AND ?) OR (end_date BETWEEN ? AND ?)
        OR (start_date <= ? AND end_date >= ?))
    ''', (staff_id, start_date, end_date, start_date, end_date, start_date, end_date)).fetchall()

    # Format attendance times to 12-hour format
    formatted_attendance = [format_attendance_times_to_12hr(dict(a)) for a in attendance]

    return jsonify({
        'success': True,
        'attendance': formatted_attendance,
        'leaves': [dict(l) for l in leaves]
    })

@app.route('/create_staff_from_device_user', methods=['POST'])
def create_staff_from_device_user():
    """Create staff account for user who already exists on biometric device"""
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session['school_id']
    device_ip = request.form.get('device_ip', '192.168.1.201')
    device_user_id = request.form.get('device_user_id')
    full_name = request.form.get('full_name')
    # Defer hashing until we know if device user has a password
    provided_password = request.form.get('password')
    email = request.form.get('email', '')
    phone = request.form.get('phone', '')
    department = request.form.get('department', '')
    position = request.form.get('position', '')

    if not device_user_id or not full_name:
        return jsonify({'success': False, 'error': 'Device User ID and full name are required'})

    try:
        # Verify user exists on device
        zk_device = ZKBiometricDevice(device_ip)
        if not zk_device.connect():
            return jsonify({'success': False, 'error': 'Failed to connect to biometric device'})

        users = zk_device.get_users()
        device_user = None
        for user in users:
            if str(user['user_id']) == str(device_user_id):
                device_user = user
                break

        zk_device.disconnect()

        if not device_user:
            return jsonify({'success': False, 'error': f'User {device_user_id} not found on biometric device'})

        # Choose password: prefer the biometric device user's password if present
        device_password = ''
        if device_user:
            # Handle both dict and object formats gracefully
            try:
                device_password = device_user.get('password') if isinstance(device_user, dict) else getattr(device_user, 'password', '')
            except Exception:
                device_password = ''

        selected_password = device_password or provided_password or 'password123'
        password_hash = generate_password_hash(selected_password)

        # Check if staff already exists in database
        db = get_db()
        existing_staff = db.execute('SELECT id FROM staff WHERE staff_id = ? AND school_id = ?',
                                  (device_user_id, school_id)).fetchone()

        if existing_staff:
            return jsonify({'success': False, 'error': f'Staff with ID {device_user_id} already exists in database'})

        # Create staff account
        db.execute('''
            INSERT INTO staff
            (school_id, staff_id, password_hash, full_name, email, phone, department, position)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (school_id, device_user_id, password_hash, full_name, email, phone, department, position))

        db.commit()

        return jsonify({
            'success': True,
            'message': f'Staff account created successfully for biometric user {device_user_id} ({full_name})'
        })

    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'error': 'Staff ID already exists'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/resolve_user_conflict', methods=['POST'])
def resolve_user_conflict():
    """Resolve conflicts when user ID already exists on biometric device"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    action = request.form.get('action')  # 'overwrite', 'use_different_id', 'create_from_existing'
    device_ip = request.form.get('device_ip', '192.168.1.201')

    if action == 'overwrite':
        # Overwrite existing user on device
        user_id = request.form.get('user_id')
        name = request.form.get('name')
        privilege = int(request.form.get('privilege', 0))

        try:
            zk_device = ZKBiometricDevice(device_ip)
            if zk_device.connect():
                result = zk_device.enroll_user(user_id, name, privilege, overwrite=True)
                zk_device.disconnect()

                if result['success']:
                    return jsonify({
                        'success': True,
                        'message': f'User {user_id} has been overwritten on the device. You can now proceed with biometric enrollment.',
                        'action': 'overwritten'
                    })
                else:
                    return jsonify({
                        'success': False,
                        'message': result['message']
                    })
            else:
                return jsonify({
                    'success': False,
                    'message': 'Failed to connect to biometric device'
                })
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Error overwriting user: {str(e)}'
            })

    elif action == 'create_from_existing':
        # Create staff account using existing biometric user
        school_id = session.get('school_id')
        if not school_id:
            return jsonify({'success': False, 'error': 'School ID not found'})

        existing_user_id = request.form.get('existing_user_id')
        full_name = request.form.get('full_name')
        password = generate_password_hash(request.form.get('password', 'default123'))
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')
        department = request.form.get('department', '')
        position = request.form.get('position', '')

        try:
            db = get_db()

            # Create staff account
            db.execute('''
                INSERT INTO staff
                (school_id, staff_id, password_hash, full_name, email, phone, department, position)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (school_id, existing_user_id, password, full_name, email, phone, department, position))

            db.commit()

            return jsonify({
                'success': True,
                'message': f'Staff account created for existing biometric user {existing_user_id}',
                'action': 'created_from_existing'
            })

        except sqlite3.IntegrityError:
            return jsonify({'success': False, 'error': 'Staff ID already exists in database'})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})

    else:
        return jsonify({
            'success': False,
            'error': 'Invalid action specified'
        })

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

# Database migration - Add created_at column to staff table (commented out to prevent errors)
# This should be run only once during initial setup
# import sqlite3
# conn = sqlite3.connect('vishnorex.db')
# try:
#     conn.execute("ALTER TABLE staff ADD COLUMN created_at TEXT")
#     conn.commit()
# except sqlite3.OperationalError:
#     pass  # Column already exists
# conn.close()

# Backup and Data Management Routes
@app.route('/create_backup', methods=['POST'])
def create_backup():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    backup_name = request.form.get('backup_name')
    include_logs = request.form.get('include_logs', True, type=bool)

    backup_manager = BackupManager()
    result = backup_manager.create_database_backup(backup_name, include_logs)

    return jsonify(result)

@app.route('/get_backup_list')
def get_backup_list():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    try:
        backup_manager = BackupManager()
        backup_dir = backup_manager.backup_dir

        backups = []
        if os.path.exists(backup_dir):
            for filename in os.listdir(backup_dir):
                if filename.endswith('.db'):
                    backup_name = filename[:-3]  # Remove .db extension
                    backup_path = os.path.join(backup_dir, filename)
                    metadata_path = os.path.join(backup_dir, f"{backup_name}_metadata.json")

                    backup_info = {
                        'name': backup_name,
                        'size': os.path.getsize(backup_path),
                        'created_at': datetime.datetime.fromtimestamp(os.path.getctime(backup_path)).isoformat()
                    }

                    # Load metadata if available
                    if os.path.exists(metadata_path):
                        try:
                            with open(metadata_path, 'r') as f:
                                metadata = json.load(f)
                                backup_info.update(metadata)
                        except:
                            pass

                    backups.append(backup_info)

        # Sort by creation date (newest first)
        backups.sort(key=lambda x: x['created_at'], reverse=True)

        return jsonify({'success': True, 'backups': backups})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/export_data_backup', methods=['POST'])
def export_data_backup():
    if 'user_id' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'})

    export_type = request.form.get('export_type', 'excel')
    school_id = session.get('school_id')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    tables = request.form.getlist('tables')

    backup_manager = BackupManager()
    result = backup_manager.export_data(export_type, school_id, start_date, end_date, tables)

    if result['success']:
        # Return file for download
        from flask import send_file
        return send_file(result['export_path'], as_attachment=True)
    else:
        return jsonify(result)

# Salary Calculation Routes
@app.route('/calculate_salary', methods=['POST'])
def calculate_salary():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id', type=int)
    year = request.form.get('year', type=int)
    month = request.form.get('month', type=int)

    if not all([staff_id, year, month]):
        return jsonify({'success': False, 'error': 'Staff ID, year, and month are required'})

    school_id = session.get('school_id')
    salary_calculator = SalaryCalculator(school_id=school_id)
    result = salary_calculator.calculate_monthly_salary(staff_id, year, month)

    return jsonify(result)

@app.route('/generate_salary_report', methods=['POST'])
def generate_salary_report():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    staff_id = request.form.get('staff_id', type=int)
    year = request.form.get('year', type=int)
    month = request.form.get('month', type=int)

    if not all([staff_id, year, month]):
        return jsonify({'success': False, 'error': 'Staff ID, year, and month are required'})

    school_id = session.get('school_id')
    salary_calculator = SalaryCalculator(school_id=school_id)
    result = salary_calculator.generate_salary_report(staff_id, year, month)

    return jsonify(result)

@app.route('/bulk_salary_calculation', methods=['POST'])
def bulk_salary_calculation():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    year = request.form.get('year', type=int)
    month = request.form.get('month', type=int)
    department = request.form.get('department')
    school_id = session.get('school_id')

    if not all([year, month]):
        return jsonify({'success': False, 'error': 'Year and month are required'})

    try:
        # Get staff list based on filters
        query = 'SELECT id, staff_id, full_name, department FROM staff WHERE school_id = ?'
        params = [school_id]

        if department:
            query += ' AND department = ?'
            params.append(department)

        staff_list = get_db().execute(query, params).fetchall()

        school_id = session.get('school_id')
        salary_calculator = SalaryCalculator(school_id=school_id)
        results = []

        for staff in staff_list:
            salary_result = salary_calculator.calculate_monthly_salary(staff['id'], year, month)
            if salary_result['success']:
                results.append({
                    'id': staff['id'],  # Add database ID
                    'staff_id': staff['staff_id'],
                    'staff_name': staff['full_name'],
                    'department': staff['department'],
                    'net_salary': salary_result['salary_breakdown']['net_salary'],
                    'total_earnings': salary_result['salary_breakdown']['earnings']['total_earnings'],
                    'total_deductions': salary_result['salary_breakdown']['deductions']['total_deductions'],
                    'present_days': salary_result['salary_breakdown']['attendance_summary']['present_days'],
                    'absent_days': salary_result['salary_breakdown']['attendance_summary']['absent_days']
                })

        return jsonify({
            'success': True,
            'calculation_period': f"{calendar.month_name[month]} {year}",
            'total_staff': len(results),
            'salary_calculations': results
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/update_salary_rules', methods=['POST'])
def update_salary_rules():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized - Admin access required'})

    try:
        new_rules = {}

        # Get salary rule updates from form
        rule_fields = [
            'early_arrival_bonus_per_hour',
            'early_departure_penalty_per_hour',
            'late_arrival_penalty_per_hour',
            'absent_day_deduction_rate',
            'overtime_rate_multiplier',
            'on_duty_rate',
            'bonus_rate_percentage',
            'minimum_hours_for_bonus'
        ]

        for field in rule_fields:
            value = request.form.get(field, type=float)
            if value is not None:
                new_rules[field] = value

        school_id = session.get('school_id')
        salary_calculator = SalaryCalculator(school_id=school_id)
        result = salary_calculator.update_salary_rules(new_rules)

        return jsonify(result)

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/calculate_hourly_rate', methods=['POST'])
def calculate_hourly_rate_api():
    """Calculate hourly rate from base monthly salary"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized - Admin access required'})

    try:
        from database import calculate_hourly_rate

        base_salary = request.json.get('base_salary', 0)
        if not base_salary or base_salary <= 0:
            return jsonify({'success': False, 'error': 'Valid base salary is required'})

        result = calculate_hourly_rate(base_salary)
        return jsonify({'success': True, 'data': result})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/calculate_enhanced_salary', methods=['POST'])
def calculate_enhanced_salary_api():
    """Calculate enhanced salary based on actual hours worked"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized - Admin access required'})

    try:
        staff_id = request.json.get('staff_id')
        year = request.json.get('year')
        month = request.json.get('month')

        if not all([staff_id, year, month]):
            return jsonify({'success': False, 'error': 'Staff ID, year, and month are required'})

        school_id = session.get('school_id')
        salary_calculator = SalaryCalculator(school_id=school_id)
        result = salary_calculator.calculate_enhanced_monthly_salary(int(staff_id), int(year), int(month))

        return jsonify(result)

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/get_staff_hourly_rate/<int:staff_id>')
def get_staff_hourly_rate(staff_id):
    """Get hourly rate for a specific staff member"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized - Admin access required'})

    try:
        from database import calculate_hourly_rate

        db = get_db()
        school_id = session['school_id']

        # Get staff basic salary
        staff = db.execute('''
            SELECT basic_salary, full_name
            FROM staff
            WHERE id = ? AND school_id = ?
        ''', (staff_id, school_id)).fetchone()

        if not staff:
            return jsonify({'success': False, 'error': 'Staff member not found'})

        basic_salary = float(staff['basic_salary'] or 0)
        if basic_salary <= 0:
            return jsonify({'success': False, 'error': 'Base salary not set for this staff member'})

        hourly_rate_info = calculate_hourly_rate(basic_salary)

        return jsonify({
            'success': True,
            'staff_name': staff['full_name'],
            'basic_salary': basic_salary,
            'hourly_rate_info': hourly_rate_info
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_salary_rules')
def get_salary_rules():
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    school_id = session.get('school_id')
    salary_calculator = SalaryCalculator(school_id=school_id)

    return jsonify({
        'success': True,
        'salary_rules': salary_calculator.salary_rules
    })

@app.route('/export_salary_calculation_results', methods=['POST'])
def export_salary_calculation_results():
    """Export salary calculation results to Excel"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    try:
        # Get the calculation parameters from the request
        year = request.form.get('year', type=int)
        month = request.form.get('month', type=int)
        department = request.form.get('department')
        export_format = request.form.get('format', 'excel').lower()

        if not year or not month:
            return jsonify({'success': False, 'error': 'Year and month are required'})

        school_id = session.get('school_id')

        # Re-run the bulk salary calculation to get fresh data
        salary_calculator = SalaryCalculator(school_id=school_id)

        # Get staff list based on filters
        query = 'SELECT id, staff_id, full_name, department FROM staff WHERE school_id = ?'
        params = [school_id]

        if department:
            query += ' AND department = ?'
            params.append(department)

        db = get_db()
        staff_list = db.execute(query, params).fetchall()

        # Calculate salaries for all staff
        salary_results = []
        for staff in staff_list:
            salary_result = salary_calculator.calculate_monthly_salary(staff['id'], year, month)
            if salary_result['success']:
                salary_results.append({
                    'id': staff['id'],
                    'staff_id': staff['staff_id'],
                    'staff_name': staff['full_name'],
                    'department': staff['department'],
                    'salary_data': salary_result
                })

        # Generate report in requested format
        if export_format == 'excel':
            return generate_salary_calculation_excel(salary_results, year, month, department)
        elif export_format == 'csv':
            return generate_salary_calculation_csv(salary_results, year, month, department)
        elif export_format == 'pdf':
            return generate_salary_calculation_pdf(salary_results, year, month, department)
        else:
            return jsonify({'success': False, 'error': 'Supported formats: excel, csv, pdf'})

    except Exception as e:
        return jsonify({'success': False, 'error': f'Export failed: {str(e)}'})

def generate_salary_calculation_excel(salary_results, year, month, department=None):
    """Generate Excel report for salary calculation results"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    import io
    import calendar

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salary Calculation Results"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    title = f"Salary Calculation Results - {calendar.month_name[month]} {year}"
    if department:
        title += f" - {department} Department"

    ws.merge_cells('A1:P1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    headers = [
        'Staff ID', 'Staff Name', 'Department', 'Base Salary', 'Present Days', 'Absent Days',
        'Total Earnings', 'Bonuses', 'Allowances', 'Total Deductions', 'Penalties',
        'Other Deductions', 'Net Salary', 'Hourly Rate', 'Hours Worked', 'Attendance Rate'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    row = 4
    total_earnings = 0
    total_deductions = 0
    total_net_salary = 0

    for result in salary_results:
        salary_data = result['salary_data']
        breakdown = salary_data['salary_breakdown']
        attendance = breakdown['attendance_summary']
        earnings = breakdown['earnings']
        deductions = breakdown['deductions']

        # Calculate attendance rate
        total_days = attendance['present_days'] + attendance['absent_days']
        attendance_rate = (attendance['present_days'] / total_days * 100) if total_days > 0 else 0

        # Row data
        row_data = [
            result['staff_id'],
            result['staff_name'],
            result['department'],
            breakdown.get('base_salary', 0),
            attendance['present_days'],
            attendance['absent_days'],
            earnings['total_earnings'],
            earnings.get('bonuses', 0),
            earnings.get('allowances', 0),
            deductions['total_deductions'],
            deductions.get('penalties', 0),
            deductions.get('other_deductions', 0),
            breakdown['net_salary'],
            salary_data.get('hourly_rate', 0),
            salary_data.get('actual_hours_worked', 0),
            f"{attendance_rate:.1f}%"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col >= 4 and col <= 15:  # Numeric columns
                cell.alignment = Alignment(horizontal='right')
                if isinstance(value, (int, float)) and col != 16:  # Not percentage
                    cell.number_format = '#,##0.00'

        # Update totals
        total_earnings += earnings['total_earnings']
        total_deductions += deductions['total_deductions']
        total_net_salary += breakdown['net_salary']
        row += 1

    # Summary row
    summary_row = row + 1
    ws.cell(row=summary_row, column=1, value="TOTALS").font = Font(bold=True)
    ws.cell(row=summary_row, column=7, value=total_earnings).font = Font(bold=True)
    ws.cell(row=summary_row, column=10, value=total_deductions).font = Font(bold=True)
    ws.cell(row=summary_row, column=13, value=total_net_salary).font = Font(bold=True)

    # Format summary row
    for col in [7, 10, 13]:
        cell = ws.cell(row=summary_row, column=col)
        cell.number_format = '#,##0.00'
        cell.border = border
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.getvalue())
    filename = f'salary_calculation_results_{year}_{month:02d}'
    if department:
        filename += f'_{department.replace(" ", "_")}'
    filename += f'_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

    return response

def generate_salary_calculation_csv(salary_results, year, month, department=None):
    """Generate CSV report for salary calculation results"""
    import csv
    import io
    import calendar

    # Create CSV content
    output = io.StringIO()
    writer = csv.writer(output)

    # Title
    title = f"Salary Calculation Results - {calendar.month_name[month]} {year}"
    if department:
        title += f" - {department} Department"
    writer.writerow([title])
    writer.writerow([])  # Empty row

    # Headers
    headers = [
        'Staff ID', 'Staff Name', 'Department', 'Base Salary', 'Present Days', 'Absent Days',
        'Total Earnings', 'Bonuses', 'Allowances', 'Total Deductions', 'Penalties',
        'Other Deductions', 'Net Salary', 'Hourly Rate', 'Hours Worked', 'Attendance Rate'
    ]
    writer.writerow(headers)

    # Data rows
    total_earnings = 0
    total_deductions = 0
    total_net_salary = 0

    for result in salary_results:
        salary_data = result['salary_data']
        breakdown = salary_data['salary_breakdown']
        attendance = breakdown['attendance_summary']
        earnings = breakdown['earnings']
        deductions = breakdown['deductions']

        # Calculate attendance rate
        total_days = attendance['present_days'] + attendance['absent_days']
        attendance_rate = (attendance['present_days'] / total_days * 100) if total_days > 0 else 0

        # Row data
        row_data = [
            result['staff_id'],
            result['staff_name'],
            result['department'],
            breakdown.get('base_salary', 0),
            attendance['present_days'],
            attendance['absent_days'],
            earnings['total_earnings'],
            earnings.get('bonuses', 0),
            earnings.get('allowances', 0),
            deductions['total_deductions'],
            deductions.get('penalties', 0),
            deductions.get('other_deductions', 0),
            breakdown['net_salary'],
            salary_data.get('hourly_rate', 0),
            salary_data.get('actual_hours_worked', 0),
            f"{attendance_rate:.1f}%"
        ]

        writer.writerow(row_data)

        # Update totals
        total_earnings += earnings['total_earnings']
        total_deductions += deductions['total_deductions']
        total_net_salary += breakdown['net_salary']

    # Summary row
    writer.writerow([])  # Empty row
    summary_row = ['TOTALS', '', '', '', '', '', total_earnings, '', '', total_deductions, '', '', total_net_salary, '', '', '']
    writer.writerow(summary_row)

    # Create response
    csv_content = output.getvalue()
    output.close()

    response = make_response(csv_content)
    filename = f'salary_calculation_results_{year}_{month:02d}'
    if department:
        filename += f'_{department.replace(" ", "_")}'
    filename += f'_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

    return response

def generate_salary_calculation_pdf(salary_results, year, month, department=None):
    """Generate PDF report for salary calculation results"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        import io
        import calendar

        # Create PDF buffer
        buffer = io.BytesIO()

        # Create PDF document
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

        # Container for the 'Flowable' objects
        elements = []

        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )

        # Title
        title = f"Salary Calculation Results<br/>{calendar.month_name[month]} {year}"
        if department:
            title += f"<br/>{department} Department"

        title_para = Paragraph(title, title_style)
        elements.append(title_para)
        elements.append(Spacer(1, 12))

        # Prepare table data
        table_data = [
            ['Staff ID', 'Name', 'Dept', 'Base Salary', 'Present', 'Absent', 'Earnings', 'Deductions', 'Net Salary']
        ]

        total_earnings = 0
        total_deductions = 0
        total_net_salary = 0

        for result in salary_results:
            salary_data = result['salary_data']
            breakdown = salary_data['salary_breakdown']
            attendance = breakdown['attendance_summary']
            earnings = breakdown['earnings']
            deductions = breakdown['deductions']

            row_data = [
                result['staff_id'],
                result['staff_name'][:15] + '...' if len(result['staff_name']) > 15 else result['staff_name'],
                result['department'][:8] + '...' if len(result['department']) > 8 else result['department'],
                f"₹{breakdown.get('base_salary', 0):,.0f}",
                str(attendance['present_days']),
                str(attendance['absent_days']),
                f"₹{earnings['total_earnings']:,.0f}",
                f"₹{deductions['total_deductions']:,.0f}",
                f"₹{breakdown['net_salary']:,.0f}"
            ]

            table_data.append(row_data)

            # Update totals
            total_earnings += earnings['total_earnings']
            total_deductions += deductions['total_deductions']
            total_net_salary += breakdown['net_salary']

        # Add totals row
        table_data.append([
            'TOTALS', '', '', '', '', '',
            f"₹{total_earnings:,.0f}",
            f"₹{total_deductions:,.0f}",
            f"₹{total_net_salary:,.0f}"
        ])

        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)

        # Get PDF content
        pdf_content = buffer.getvalue()
        buffer.close()

        # Create response
        response = make_response(pdf_content)
        filename = f'salary_calculation_results_{year}_{month:02d}'
        if department:
            filename += f'_{department.replace(" ", "_")}'
        filename += f'_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'

        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'

        return response

    except ImportError:
        # If reportlab is not installed, return error
        return jsonify({
            'success': False,
            'error': 'PDF generation requires reportlab library. Please install it with: pip install reportlab'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'PDF generation failed: {str(e)}'})

@app.route('/get_staff_count')
def get_staff_count():
    """Get total staff count for sidebar stats"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'})

    try:
        db = get_db()
        school_id = session.get('school_id', 1)

        count = db.execute(
            'SELECT COUNT(*) as total FROM staff WHERE school_id = ?',
            (school_id,)
        ).fetchone()

        return jsonify({
            'success': True,
            'count': count['total'] if count else 0
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/reports')
def admin_reports():
    """Admin reports page"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return redirect(url_for('index'))

    return render_template('admin_reports.html')

@app.route('/admin/settings')
def admin_settings():
    """Admin settings page"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return redirect(url_for('index'))

    return render_template('admin_settings.html')

@app.route('/api/get_institution_timings', methods=['GET'])
def get_institution_timings():
    """Get current institution check-in and check-out timings"""
    try:
        db = get_db()

        # Check if settings table exists and create if not
        cursor = db.execute("""
            SELECT name FROM sqlite_master
            WHERE type='table' AND name='institution_settings'
        """)

        if not cursor.fetchone():
            # Create the settings table
            db.execute("""
                CREATE TABLE institution_settings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    setting_name TEXT UNIQUE NOT NULL,
                    setting_value TEXT NOT NULL,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Insert default timings
            db.execute("""
                INSERT INTO institution_settings (setting_name, setting_value)
                VALUES ('checkin_time', '09:00'), ('checkout_time', '17:00')
            """)
            db.commit()

        # Fetch current timings
        cursor = db.execute("""
            SELECT setting_name, setting_value
            FROM institution_settings
            WHERE setting_name IN ('checkin_time', 'checkout_time')
        """)

        settings = dict(cursor.fetchall())

        return jsonify({
            'success': True,
            'checkin_time': settings.get('checkin_time', '09:00'),
            'checkout_time': settings.get('checkout_time', '17:00')
        })

    except Exception as e:
        print(f"Error getting institution timings: {e}")
        return jsonify({
            'success': False,
            'message': 'Failed to retrieve institution timings'
        }), 500

@app.route('/api/update_institution_timings', methods=['POST'])
def update_institution_timings():
    """Update institution check-in and check-out timings"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        checkin_time = request.form.get('checkin_time')
        checkout_time = request.form.get('checkout_time')

        # Validate inputs
        if not checkin_time or not checkout_time:
            return jsonify({
                'success': False,
                'message': 'Both check-in and check-out times are required'
            }), 400

        # Validate time format (HH:MM)
        import re
        time_pattern = re.compile(r'^([01]?[0-9]|2[0-3]):[0-5][0-9]$')

        if not time_pattern.match(checkin_time) or not time_pattern.match(checkout_time):
            return jsonify({
                'success': False,
                'message': 'Invalid time format. Please use HH:MM format'
            }), 400

        # Validate that checkout is after checkin
        from datetime import datetime
        checkin_dt = datetime.strptime(checkin_time, '%H:%M')
        checkout_dt = datetime.strptime(checkout_time, '%H:%M')

        if checkout_dt <= checkin_dt:
            return jsonify({
                'success': False,
                'message': 'Check-out time must be later than check-in time'
            }), 400

        db = get_db()

        # Update or insert the timings
        for setting_name, setting_value in [('checkin_time', checkin_time), ('checkout_time', checkout_time)]:
            db.execute("""
                INSERT OR REPLACE INTO institution_settings (setting_name, setting_value, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
            """, (setting_name, setting_value))

        # Sync with shift definitions - update 'general' shift to match institution timings
        try:
            # Check if shift_definitions table exists
            cursor = db.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='shift_definitions'
            """)

            if cursor.fetchone():
                # Update or insert general shift to match institution timings with strict rules
                db.execute("""
                    INSERT OR REPLACE INTO shift_definitions
                    (shift_type, start_time, end_time, grace_period_minutes, description, is_active)
                    VALUES ('general', ?, ?, 0, 'Institution Default Shift (Strict Timing)', 1)
                """, (checkin_time + ':00', checkout_time + ':00'))

                print(f"✅ Synced general shift: {checkin_time} - {checkout_time}")
            else:
                # Create shift_definitions table and insert general shift
                db.execute("""
                    CREATE TABLE IF NOT EXISTS shift_definitions (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        shift_type TEXT UNIQUE NOT NULL,
                        start_time TEXT NOT NULL,
                        end_time TEXT NOT NULL,
                        grace_period_minutes INTEGER DEFAULT 10,
                        description TEXT,
                        is_active INTEGER DEFAULT 1,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)

                db.execute("""
                    INSERT INTO shift_definitions
                    (shift_type, start_time, end_time, grace_period_minutes, description, is_active)
                    VALUES ('general', ?, ?, 0, 'Institution Default Shift (Strict Timing)', 1)
                """, (checkin_time + ':00', checkout_time + ':00'))

                print(f"✅ Created shift_definitions table and synced general shift: {checkin_time} - {checkout_time}")

        except Exception as sync_error:
            print(f"⚠️ Warning: Could not sync shift definitions: {sync_error}")
            # Continue execution even if shift sync fails

        db.commit()

        # Notify all systems to refresh their configurations
        try:
            # Reload shift manager if it exists
            from shift_management import ShiftManager
            if hasattr(app, 'shift_manager'):
                app.shift_manager.reload_shift_definitions()
            else:
                # Create new shift manager to ensure latest timings are loaded
                app.shift_manager = ShiftManager()

            print(f"✅ Institution timings updated and synced across all systems")

        except Exception as reload_error:
            print(f"⚠️ Warning: Could not reload shift manager: {reload_error}")

        return jsonify({
            'success': True,
            'message': 'Institution timings updated successfully',
            'checkin_time': checkin_time,
            'checkout_time': checkout_time
        })

    except Exception as e:
        print(f"Error updating institution timings: {e}")
        return jsonify({
            'success': False,
            'message': 'Failed to update institution timings'
        }), 500

@app.route('/api/debug_session', methods=['GET'])
def debug_session():
    """Debug route to check session status"""
    try:
        return jsonify({
            'success': True,
            'session_data': {
                'user_id': session.get('user_id'),
                'user_type': session.get('user_type'),
                'full_name': session.get('full_name'),
                'has_session': 'user_id' in session
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/test_timing_sync', methods=['GET'])
def test_timing_sync():
    """Test route to verify that all systems are using the same institution timings"""
    try:
        from database import get_institution_timings, calculate_attendance_status
        from shift_management import ShiftManager
        import datetime

        # Get institution timings
        institution_timings = get_institution_timings()

        # Get shift manager timings
        shift_manager = ShiftManager()
        general_shift = shift_manager.get_shift_info('general')

        # Test attendance calculation
        test_time = datetime.time(9, 30)  # 9:30 AM
        status = calculate_attendance_status(test_time, 'check-in')

        return jsonify({
            'success': True,
            'sync_check': {
                'institution_checkin': institution_timings['checkin_time'].strftime('%H:%M'),
                'institution_checkout': institution_timings['checkout_time'].strftime('%H:%M'),
                'institution_is_custom': institution_timings['is_custom'],
                'shift_manager_checkin': general_shift['start_time'].strftime('%H:%M') if general_shift else 'Not found',
                'shift_manager_checkout': general_shift['end_time'].strftime('%H:%M') if general_shift else 'Not found',
                'attendance_status_test': f"9:30 AM check-in status: {status}",
                'systems_synced': (
                    institution_timings['checkin_time'] == general_shift['start_time'] and
                    institution_timings['checkout_time'] == general_shift['end_time']
                ) if general_shift else False
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# Holiday Management API Routes

@app.route('/api/holidays', methods=['GET'])
def get_holidays_api():
    """Get holidays for the current school"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        from database import get_holidays
        import json

        # Get query parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        department = request.args.get('department')

        # Get holidays
        holidays = get_holidays(
            start_date=start_date,
            end_date=end_date,
            department=department
        )

        # Convert to list of dictionaries for JSON serialization
        holidays_list = []
        for holiday in holidays:
            holiday_dict = dict(holiday)
            # Parse departments JSON if present
            if holiday_dict.get('departments'):
                try:
                    holiday_dict['departments'] = json.loads(holiday_dict['departments'])
                except (json.JSONDecodeError, TypeError):
                    holiday_dict['departments'] = []
            else:
                holiday_dict['departments'] = []
            holidays_list.append(holiday_dict)

        return jsonify({
            'success': True,
            'holidays': holidays_list,
            'count': len(holidays_list)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to retrieve holidays'
        }), 500


@app.route('/api/holidays', methods=['POST'])
def create_holiday_api():
    """Create a new holiday"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        from database import create_holiday

        # Get form data
        holiday_data = {
            'holiday_name': request.form.get('holiday_name'),
            'start_date': request.form.get('start_date'),
            'end_date': request.form.get('end_date'),
            'holiday_type': request.form.get('holiday_type', 'institution_wide'),
            'description': request.form.get('description', ''),
            'is_recurring': bool(request.form.get('is_recurring')),
            'recurring_type': request.form.get('recurring_type')
        }

        # Handle departments for department-specific holidays
        if holiday_data['holiday_type'] == 'department_specific':
            departments = request.form.getlist('departments')
            if not departments:
                departments_str = request.form.get('departments')
                if departments_str:
                    departments = [dept.strip() for dept in departments_str.split(',')]
            holiday_data['departments'] = departments

        # Validate required fields
        if not holiday_data['holiday_name'] or not holiday_data['start_date'] or not holiday_data['end_date']:
            return jsonify({
                'success': False,
                'message': 'Holiday name, start date, and end date are required'
            }), 400

        # Create holiday
        result = create_holiday(holiday_data)

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to create holiday'
        }), 500


@app.route('/api/holidays/<int:holiday_id>', methods=['PUT'])
def update_holiday_api(holiday_id):
    """Update an existing holiday"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        from database import update_holiday

        # Get form data
        holiday_data = {
            'holiday_name': request.form.get('holiday_name'),
            'start_date': request.form.get('start_date'),
            'end_date': request.form.get('end_date'),
            'holiday_type': request.form.get('holiday_type'),
            'description': request.form.get('description'),
            'is_recurring': bool(request.form.get('is_recurring')),
            'recurring_type': request.form.get('recurring_type')
        }

        # Handle departments for department-specific holidays
        if holiday_data['holiday_type'] == 'department_specific':
            departments = request.form.getlist('departments')
            if not departments:
                departments_str = request.form.get('departments')
                if departments_str:
                    departments = [dept.strip() for dept in departments_str.split(',')]
            holiday_data['departments'] = departments

        # Update holiday
        result = update_holiday(holiday_id, holiday_data)

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to update holiday'
        }), 500


@app.route('/api/holidays/<int:holiday_id>', methods=['DELETE'])
def delete_holiday_api(holiday_id):
    """Delete a holiday"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        from database import delete_holiday

        # Delete holiday
        result = delete_holiday(holiday_id)

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to delete holiday'
        }), 500


@app.route('/api/departments', methods=['GET'])
def get_departments_api():
    """Get list of departments for the current school"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        from database import get_departments_list

        departments = get_departments_list()

        return jsonify({
            'success': True,
            'departments': departments,
            'count': len(departments)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to retrieve departments'
        }), 500


# Weekly Off Configuration API Routes

@app.route('/api/weekly_off_config', methods=['GET'])
def get_weekly_off_config_api():
    """Get weekly off configuration for the current school"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        from database import get_weekly_off_config

        result = get_weekly_off_config()

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to get weekly off configuration'
        }), 500


@app.route('/api/weekly_off_config', methods=['POST'])
def save_weekly_off_config_api():
    """Save weekly off configuration for the current school"""
    try:
        # Check authorization
        if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        # CSRF Protection
        try:
            csrf.protect()
        except ValidationError as e:
            return jsonify({
                'success': False,
                'message': 'CSRF token validation failed',
                'error': str(e)
            }), 400

        from database import save_weekly_off_config

        # Get weekly off days from form data
        weekly_off_days = []

        # Check for Sunday off
        if request.form.get('sunday_off_enabled') == 'true':
            weekly_off_days.append('sunday')

        result = save_weekly_off_config(weekly_off_days)

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to save weekly off configuration'
        }), 500


@app.route('/api/staff/holidays', methods=['GET'])
def get_staff_holidays_api():
    """Get holidays applicable to a specific staff member based on their department"""
    try:
        # Check authorization - both admin and staff can access this
        if 'user_id' not in session:
            return jsonify({
                'success': False,
                'message': 'Unauthorized access'
            }), 403

        # Get staff_id - either from parameter (admin) or session (staff)
        staff_id = request.args.get('staff_id')
        if session['user_type'] == 'staff':
            staff_id = session['user_id']
        elif session['user_type'] in ['admin', 'company_admin'] and staff_id:
            staff_id = staff_id
        else:
            return jsonify({
                'success': False,
                'message': 'Staff ID required'
            }), 400

        # Get date range parameters
        start_date = request.args.get('start_date')  # YYYY-MM-DD format
        end_date = request.args.get('end_date')      # YYYY-MM-DD format

        db = get_db()

        # Get staff information including department
        staff_info = db.execute('''
            SELECT id, staff_id, full_name, department
            FROM staff
            WHERE id = ?
        ''', (staff_id,)).fetchone()

        if not staff_info:
            return jsonify({
                'success': False,
                'message': 'Staff member not found'
            }), 404

        staff_department = staff_info['department'] or ''

        # Build query to get applicable holidays
        query_conditions = ['h.is_active = 1']
        query_params = []

        # Base query for holidays
        base_query = '''
            SELECT h.id, h.holiday_name, h.start_date, h.end_date, h.holiday_type,
                   h.description, h.is_recurring, h.recurring_type, h.created_at, h.departments
            FROM holidays h
        '''

        # Add date range filter if provided
        if start_date and end_date:
            query_conditions.append('''
                ((h.start_date BETWEEN ? AND ?) OR
                 (h.end_date BETWEEN ? AND ?) OR
                 (h.start_date <= ? AND h.end_date >= ?))
            ''')
            query_params.extend([start_date, end_date, start_date, end_date, start_date, end_date])

        # Combine query
        full_query = base_query + ' WHERE ' + ' AND '.join(query_conditions)
        full_query += ' ORDER BY h.start_date ASC'

        # Execute query to get all holidays in date range
        holidays_data = db.execute(full_query, query_params).fetchall()

        # Filter holidays applicable to this staff member
        holidays = []
        import json
        for holiday in holidays_data:
            # Always include institution_wide and common_leave holidays
            if holiday['holiday_type'] in ['institution_wide', 'common_leave']:
                holiday_dict = {
                    'id': holiday['id'],
                    'name': holiday['holiday_name'],  # Use 'name' to match JS expectations
                    'holiday_name': holiday['holiday_name'],
                    'start_date': holiday['start_date'],
                    'end_date': holiday['end_date'],
                    'type': holiday['holiday_type'],  # Use 'type' to match JS expectations
                    'holiday_type': holiday['holiday_type'],
                    'description': holiday['description'],
                    'is_recurring': bool(holiday['is_recurring']),
                    'recurring_type': holiday['recurring_type'],
                    'created_at': holiday['created_at']
                }
                holidays.append(holiday_dict)

            # For department-specific holidays, check if staff's department is included
            elif holiday['holiday_type'] == 'department_specific' and staff_department:
                try:
                    departments = json.loads(holiday['departments'] or '[]')
                    if staff_department in departments:
                        holiday_dict = {
                            'id': holiday['id'],
                            'name': holiday['holiday_name'],  # Use 'name' to match JS expectations
                            'holiday_name': holiday['holiday_name'],
                            'start_date': holiday['start_date'],
                            'end_date': holiday['end_date'],
                            'type': holiday['holiday_type'],  # Use 'type' to match JS expectations
                            'holiday_type': holiday['holiday_type'],
                            'description': holiday['description'],
                            'is_recurring': bool(holiday['is_recurring']),
                            'recurring_type': holiday['recurring_type'],
                            'created_at': holiday['created_at']
                        }
                        holidays.append(holiday_dict)
                except (json.JSONDecodeError, TypeError):
                    # Skip holidays with invalid department JSON
                    continue

        return jsonify({
            'success': True,
            'holidays': holidays,
            'count': len(holidays),
            'staff_info': {
                'id': staff_info['id'],
                'staff_id': staff_info['staff_id'],
                'full_name': staff_info['full_name'],
                'department': staff_department
            },
            'date_range': {
                'start_date': start_date,
                'end_date': end_date
            }
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to retrieve holidays for staff member'
        }), 500


@app.route('/admin/holiday_management')
def holiday_management():
    """Holiday Management page for administrators"""
    if 'user_id' not in session or session['user_type'] not in ['admin', 'company_admin']:
        return redirect(url_for('index'))

    return render_template('holiday_management.html')


if __name__ == '__main__':
    init_db(app)
    app.run(debug=True, host='0.0.0.0', port=5000)
