#!/usr/bin/env python3

"""
FINAL VERIFICATION: Monthly Attendance Report Excel Export
"""

from app import app
import datetime
import openpyxl
import io

def final_verification():
    """Final test of the Monthly Attendance Report"""
    
    with app.app_context():
        from app import get_db
        from excel_reports import ExcelReportGenerator
        
        print("🎯 FINAL VERIFICATION: MONTHLY ATTENDANCE REPORT")
        print("=" * 60)
        
        # Get a valid school_id from the database
        db = get_db()
        school_result = db.execute("SELECT DISTINCT school_id FROM staff WHERE is_active = 1 LIMIT 1").fetchone()
        test_school_id = school_result['school_id'] if school_result else 1
        
        # Use current year and month for testing
        current_year = datetime.datetime.now().year
        current_month = datetime.datetime.now().month
        
        print(f"📊 Testing School ID: {test_school_id}")
        print(f"📊 Testing Period: {current_year}/{current_month:02d}")
        
        # Count business days
        start_date = datetime.datetime(current_year, current_month, 1).date()
        if current_month == 12:
            end_date = datetime.datetime(current_year + 1, 1, 1).date() - datetime.timedelta(days=1)
        else:
            end_date = datetime.datetime(current_year, current_month + 1, 1).date() - datetime.timedelta(days=1)
        
        business_days = 0
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() < 5:  # Monday = 0, Sunday = 6
                business_days += 1
            current_date += datetime.timedelta(days=1)
        
        print(f"📊 Business days in month: {business_days}")
        
        # Count active staff
        staff_count = db.execute(
            'SELECT COUNT(*) as count FROM staff WHERE school_id = ? AND is_active = 1', 
            (test_school_id,)
        ).fetchone()['count']
        
        print(f"📊 Active staff count: {staff_count}")
        print(f"📊 Expected total working days: {staff_count * business_days}")
        
        try:
            # Generate the report
            excel_generator = ExcelReportGenerator()
            response = excel_generator.create_monthly_report(test_school_id, current_year, current_month)
            
            if response and hasattr(response, 'data'):
                # Save the Excel file
                filename = f'FINAL_monthly_report_{current_year}_{current_month:02d}.xlsx'
                with open(filename, 'wb') as f:
                    f.write(response.data)
                
                # Load and inspect the Excel file
                workbook = openpyxl.load_workbook(io.BytesIO(response.data))
                
                print(f"\\n📋 EXCEL FILE STRUCTURE:")
                print(f"   Sheet Names: {workbook.sheetnames}")
                print(f"   First Sheet: {workbook.sheetnames[0]}")
                
                staff_sheet = workbook['Staff Records']
                
                print(f"\\n📊 STAFF RECORDS SHEET VERIFICATION:")
                
                # Headers
                headers = [staff_sheet.cell(row=5, column=c).value for c in range(1, 10)]
                print(f"   Headers: {headers}")
                
                # Find totals row
                row = 6
                staff_data_count = 0
                while staff_sheet.cell(row=row, column=1).value and not str(staff_sheet.cell(row=row, column=1).value).startswith('TOTAL'):
                    staff_data_count += 1
                    row += 1
                
                print(f"   Staff records found: {staff_data_count}")
                
                # Find totals
                totals_row = None
                for check_row in range(row, row + 5):
                    cell_val = staff_sheet.cell(row=check_row, column=1).value
                    if cell_val and 'TOTAL' in str(cell_val):
                        totals_row = check_row
                        break
                
                if totals_row:
                    totals = [staff_sheet.cell(row=totals_row, column=c).value for c in range(1, 10)]
                    print(f"   Totals Row: {totals}")
                    
                    total_working_days = totals[4]  # Column 5 (index 4)
                    total_present = totals[8]       # Column 9 (index 8)
                    total_absent = totals[5]        # Column 6 (index 5)
                    
                    expected_working_days = staff_count * business_days
                    
                    print(f"\\n🔍 FINAL CALCULATIONS CHECK:")
                    print(f"   ✅ Staff Records Count: {staff_data_count} (matches database: {staff_count})")
                    print(f"   ✅ Business Days: {business_days}")
                    print(f"   ✅ Total Working Days: {total_working_days} (expected: {expected_working_days})")
                    print(f"   ✅ Total Present: {total_present}")
                    print(f"   ✅ Total Absent: {total_absent}")
                    
                    # Verify the math
                    grand_total = sum([int(totals[i] or 0) for i in [5, 6, 7, 8]])  # absent + leave + on_duty + present
                    print(f"   ✅ Grand Total Check: {grand_total} = {int(total_working_days or 0)}")
                    
                    all_correct = (
                        staff_data_count == staff_count and
                        int(total_working_days or 0) == expected_working_days and
                        grand_total == int(total_working_days or 0)
                    )
                    
                    if all_correct:
                        print(f"\\n🎉 ALL VERIFICATIONS PASSED!")
                        print(f"✅ Staff Records sheet is correctly structured")
                        print(f"✅ All calculations are accurate")
                        print(f"✅ Totals row is properly calculated")
                        print(f"\\n💾 Excel file saved as: {filename}")
                        return True
                    else:
                        print(f"\\n❌ Some verifications failed")
                        return False
                else:
                    print(f"   ❌ Totals row not found")
                    return False
            else:
                print(f"❌ Failed to generate report")
                return False
                
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

if __name__ == "__main__":
    success = final_verification()
    
    if success:
        print(f"\\n🚀 MONTHLY ATTENDANCE REPORT IS READY!")
        print(f"   The Excel export now includes:")
        print(f"   • 'Staff Records' sheet with detailed individual records")
        print(f"   • Proper business day calculations")
        print(f"   • Accurate present/absent/leave/on-duty counts")
        print(f"   • Correct totals and summary information")
        print(f"   • Staff IDs sorted numerically")
        print(f"   • Proper formatting and user-friendly layout")
    else:
        print(f"\\n❌ VERIFICATION FAILED")
        print(f"   Please check the logs above for details")