#!/usr/bin/env python3
"""
Test script to diagnose Excel export functionality issues
"""

import sys
import os
import traceback

def test_basic_imports():
    """Test if all required libraries are available"""
    print("Testing basic imports...")
    
    try:
        import openpyxl
        print(f"✓ openpyxl imported successfully (version: {openpyxl.__version__})")
    except ImportError as e:
        print(f"✗ openpyxl import failed: {e}")
        return False
    
    try:
        import pandas
        print(f"✓ pandas imported successfully (version: {pandas.__version__})")
    except ImportError as e:
        print(f"✗ pandas import failed: {e}")
        return False
    
    try:
        import xlsxwriter
        print(f"✓ xlsxwriter imported successfully (version: {xlsxwriter.__version__})")
    except ImportError as e:
        print(f"✗ xlsxwriter import failed: {e}")
        # Not critical since we're using openpyxl
    
    return True

def test_excel_reports_module():
    """Test if excel_reports module can be imported and instantiated"""
    print("\nTesting excel_reports module...")
    
    try:
        from excel_reports import ExcelReportGenerator
        print("✓ ExcelReportGenerator imported successfully")
        
        generator = ExcelReportGenerator()
        print("✓ ExcelReportGenerator instantiated successfully")
        return generator
    except ImportError as e:
        print(f"✗ excel_reports import failed: {e}")
        traceback.print_exc()
        return None
    except Exception as e:
        print(f"✗ ExcelReportGenerator instantiation failed: {e}")
        traceback.print_exc()
        return None

def test_database_connection():
    """Test database connection"""
    print("\nTesting database connection...")
    
    try:
        from database import get_db
        db = get_db()
        print("✓ Database connection successful")
        return db
    except Exception as e:
        print(f"✗ Database connection failed: {e}")
        traceback.print_exc()
        return None

def test_simple_excel_creation():
    """Test creating a simple Excel file"""
    print("\nTesting simple Excel file creation...")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # Create a simple workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Add some data
        ws['A1'] = "Test Header"
        ws['A1'].font = Font(bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        ws['A2'] = "Test Data"
        ws['B2'] = "Value 1"
        
        # Save to a test file
        test_file = "test_excel_output.xlsx"
        wb.save(test_file)
        
        print(f"✓ Simple Excel file created successfully: {test_file}")
        
        # Check file size
        if os.path.exists(test_file):
            size = os.path.getsize(test_file)
            print(f"✓ File size: {size} bytes")
            os.remove(test_file)  # Clean up
            print("✓ Test file cleaned up")
        
        return True
        
    except Exception as e:
        print(f"✗ Simple Excel creation failed: {e}")
        traceback.print_exc()
        return False

def test_csv_alternative():
    """Test CSV export as an alternative"""
    print("\nTesting CSV export alternative...")
    
    try:
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write test data
        writer.writerow(['Header 1', 'Header 2', 'Header 3'])
        writer.writerow(['Data 1', 'Data 2', 'Data 3'])
        writer.writerow(['Test 1', 'Test 2', 'Test 3'])
        
        csv_content = output.getvalue()
        print(f"✓ CSV content generated successfully ({len(csv_content)} characters)")
        
        return csv_content
        
    except Exception as e:
        print(f"✗ CSV export failed: {e}")
        traceback.print_exc()
        return None

def main():
    """Run all tests"""
    print("=" * 60)
    print("Excel Export Functionality Diagnostic Test")
    print("=" * 60)
    
    # Test 1: Basic imports
    if not test_basic_imports():
        print("\nCritical: Basic imports failed. Please install required packages.")
        return
    
    # Test 2: Excel reports module
    excel_generator = test_excel_reports_module()
    
    # Test 3: Database connection
    db = test_database_connection()
    
    # Test 4: Simple Excel creation
    simple_excel_ok = test_simple_excel_creation()
    
    # Test 5: CSV alternative
    csv_content = test_csv_alternative()
    
    print("\n" + "=" * 60)
    print("DIAGNOSTIC SUMMARY")
    print("=" * 60)
    
    if simple_excel_ok:
        print("✓ Basic Excel functionality is working")
    else:
        print("✗ Basic Excel functionality has issues")
    
    if excel_generator:
        print("✓ ExcelReportGenerator module is functional")
    else:
        print("✗ ExcelReportGenerator module has issues")
    
    if db:
        print("✓ Database connection is working")
    else:
        print("✗ Database connection has issues")
    
    if csv_content:
        print("✓ CSV export fallback is available")
    else:
        print("✗ CSV export fallback has issues")
    
    print("\nRecommendations:")
    if not simple_excel_ok:
        print("- Check openpyxl installation and permissions")
    if not excel_generator:
        print("- Check excel_reports.py for syntax/import errors")
    if not db:
        print("- Check database configuration and file permissions")
    
    print("=" * 60)

if __name__ == "__main__":
    main()
