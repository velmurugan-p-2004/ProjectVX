import openpyxl

wb = openpyxl.load_workbook('corrected_monthly_calculations_2025_09.xlsx')
ws = wb['Staff Records']

print('Finding totals row...')
row = 6
while ws.cell(row=row, column=1).value and not str(ws.cell(row=row, column=1).value).startswith('TOTAL'):
    row += 1

print(f'Staff data ends at row {row-1}')
print(f'Total staff count: {row-6}')

# Check for totals row
if ws.cell(row=row, column=1).value:
    print(f'Totals row {row}: {[ws.cell(row=row, column=c).value for c in range(1, 10)]}')
else:
    # Look for totals row a bit further down
    for check_row in range(row, row+5):
        cell_val = ws.cell(row=check_row, column=1).value
        if cell_val and 'TOTAL' in str(cell_val):
            print(f'Found totals at row {check_row}: {[ws.cell(row=check_row, column=c).value for c in range(1, 10)]}')
            break
    else:
        print('No totals row found')

# Show business day calculation logic
print(f'\nBusiness day calculation for September 2025:')
import datetime
start_date = datetime.datetime(2025, 9, 1).date()
end_date = datetime.datetime(2025, 9, 30).date()

business_days = 0
current_date = start_date
while current_date <= end_date:
    if current_date.weekday() < 5:  # Monday = 0, Sunday = 6
        business_days += 1
        if business_days <= 5 or business_days >= 20:  # Show first few and last few
            print(f'  {current_date} ({current_date.strftime("%A")}) - Business day #{business_days}')
    current_date += datetime.timedelta(days=1)

print(f'Total business days in September 2025: {business_days}')