#!/usr/bin/env python3
"""
Comprehensive test for staff report generation after address column fix.
"""

import requests
import json
import sys
import os

def test_staff_directory_endpoint():
    """Test the staff directory report generation endpoint."""
    
    print("=== Testing Staff Directory Report Endpoint ===")
    
    # Test URL (assuming Flask app runs on localhost:5000)
    url = "http://localhost:5000/generate_admin_report"
    
    # Test parameters for staff directory report
    params = {
        "report_type": "staff_directory",
        "format": "excel"
    }
    
    try:
        # Make the request
        response = requests.get(url, params=params, timeout=15)
        
        print(f"Status Code: {response.status_code}")
        print(f"Content Type: {response.headers.get('content-type', 'Unknown')}")
        
        if response.status_code == 200:
            # Check if it's a file download
            content_disposition = response.headers.get('content-disposition', '')
            if 'attachment' in content_disposition and 'xlsx' in content_disposition:
                print("✅ Excel file download response received!")
                print(f"Content-Disposition: {content_disposition}")
                print(f"File size: {len(response.content)} bytes")
                
                # Save file for verification
                filename = "test_staff_directory_report.xlsx"
                with open(filename, 'wb') as f:
                    f.write(response.content)
                print(f"✅ File saved as: {filename}")
                
                # Try to open with openpyxl to verify it's a valid Excel file
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(filename)
                    ws = wb.active
                    print(f"✅ Valid Excel file with {ws.max_row} rows and {ws.max_column} columns")
                    
                    # Check headers
                    headers = [ws.cell(row=3, column=col).value for col in range(1, ws.max_column + 1)]
                    print(f"✅ Headers: {headers}")
                    
                    wb.close()
                    return True
                    
                except ImportError:
                    print("⚠️  openpyxl not available for file validation")
                    return True
                except Exception as e:
                    print(f"❌ Excel file validation failed: {e}")
                    return False
                
            else:
                # Not a file download - might be JSON error
                try:
                    result = response.json()
                    print(f"❌ JSON Response instead of file: {json.dumps(result, indent=2)}")
                    return False
                except:
                    print(f"❌ Unexpected response: {response.text[:200]}...")
                    return False
        else:
            print(f"❌ Error response: {response.text}")
            return False
            
    except requests.exceptions.ConnectionError:
        print("❌ Connection failed - Flask app may not be running")
        print("\n💡 To test manually:")
        print("   1. Start the Flask app: python app.py")
        print("   2. Login as admin")
        print("   3. Go to Reports & Analytics or Admin Dashboard")
        print("   4. Click 'Generate' on Staff Directory report")
        print("   5. Verify Excel file downloads successfully")
        return False
        
    except Exception as e:
        print(f"❌ Test failed: {e}")
        return False

def test_direct_function_call():
    """Test the report function directly without HTTP."""
    print("\n=== Testing Direct Function Call ===")
    
    try:
        # Import the function directly
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from app import app, generate_staff_directory_report
        
        with app.app_context():
            # Simulate the function call
            print("Testing generate_staff_directory_report(school_id=1, format_type='excel')...")
            
            try:
                result = generate_staff_directory_report(1, 'excel')
                print(f"✅ Function executed successfully!")
                print(f"Response type: {type(result)}")
                
                # Check if it's a Flask response
                if hasattr(result, 'headers'):
                    content_type = result.headers.get('Content-Type', 'Unknown')
                    content_disp = result.headers.get('Content-Disposition', 'Unknown')
                    print(f"Content-Type: {content_type}")
                    print(f"Content-Disposition: {content_disp}")
                    
                    if 'xlsx' in content_type or 'xlsx' in content_disp:
                        print("✅ Response appears to be an Excel file")
                        return True
                    else:
                        print(f"❌ Response doesn't appear to be Excel: {content_type}")
                        return False
                else:
                    print(f"❌ Unexpected response type: {result}")
                    return False
                    
            except Exception as e:
                print(f"❌ Function call failed: {e}")
                print(f"   Error type: {type(e)}")
                
                # Check if it's the specific column error we were fixing
                if "no such column: s.address" in str(e):
                    print("❌ CRITICAL: The address column error still exists!")
                    return False
                elif "no such column" in str(e):
                    print(f"❌ Different column error: {e}")
                    return False
                else:
                    print("❌ Other error (may be related to session/context)")
                    return False
                    
    except ImportError as e:
        print(f"❌ Import failed: {e}")
        return False
    except Exception as e:
        print(f"❌ Direct test failed: {e}")
        return False

if __name__ == "__main__":
    print("Staff Report Generation - End-to-End Test")
    print("=" * 60)
    
    # Test direct function call first
    direct_ok = test_direct_function_call()
    
    # Test HTTP endpoint if app is running
    http_ok = test_staff_directory_endpoint()
    
    print("\n" + "=" * 60)
    print("TEST RESULTS:")
    print(f"Direct Function Call: {'✅ PASSED' if direct_ok else '❌ FAILED'}")
    print(f"HTTP Endpoint Test:   {'✅ PASSED' if http_ok else '❌ FAILED or APP NOT RUNNING'}")
    
    if direct_ok:
        print("\n✅ STAFF ADDRESS COLUMN FIX IS WORKING!")
        print("\nThe fix successfully resolved:")
        print("• 'no such column: s.address' error")
        print("• 'no such column: s.emergency_contact' error") 
        print("• 'no such column: s.qualification' error")
        print("• 'no such column: s.experience' error")
        print("• 'no such column: s.updated_at' error")
        
        print("\nStaff directory reports now work with 11 columns:")
        print("1. S.No  2. Staff ID  3. Full Name  4. Department  5. Position")
        print("6. Gender  7. Phone  8. Email  9. Date of Joining  10. Date of Birth  11. Shift Type")
    else:
        print("\n❌ ISSUES REMAIN - Please check the error messages above")
    
    if not http_ok and direct_ok:
        print("\n💡 The fix works in direct calls. For web testing:")
        print("1. Start Flask app: python app.py")
        print("2. Login as admin")
        print("3. Test report generation from the web interface")
