#!/usr/bin/env python3

"""
Test the Daily Attendance Report Excel export to verify individual staff data
"""

from app import app
import datetime
import openpyxl
import io

def test_daily_attendance_export():
    """Test the generate_daily_attendance_report function directly"""
    
    with app.app_context():
        from app import get_db, generate_daily_attendance_report
        
        print("🔍 TESTING DAILY ATTENDANCE REPORT EXCEL EXPORT")
        print("=" * 60)
        
        # Get a valid school_id from the database
        db = get_db()
        school_result = db.execute("SELECT DISTINCT school_id FROM staff WHERE is_active = 1 LIMIT 1").fetchone()
        test_school_id = school_result['school_id'] if school_result else 1
        
        # Use today's date for testing
        today_str = datetime.date.today().strftime('%Y-%m-%d')
        
        print(f"📊 Test Parameters:")
        print(f"   School ID: {test_school_id}")
        print(f"   Date: {today_str}")
        print(f"   Format: Excel")
        
        try:
            # Call the function directly
            response = generate_daily_attendance_report(
                school_id=test_school_id,
                date_str=today_str,
                format_type='excel'
            )
            
            if response and hasattr(response, 'data'):
                print(f"✅ Report generated successfully!")
                print(f"   Response type: {type(response)}")
                print(f"   Data size: {len(response.data)} bytes")
                
                # Save the Excel file for inspection
                filename = f'test_daily_attendance_{today_str}.xlsx'
                with open(filename, 'wb') as f:
                    f.write(response.data)
                
                # Load and inspect the Excel file
                workbook = openpyxl.load_workbook(io.BytesIO(response.data))
                sheet_names = workbook.sheetnames
                
                print(f"\n📋 Excel File Analysis:")
                print(f"   Sheets found: {sheet_names}")
                
                # Check each sheet
                for i, sheet_name in enumerate(sheet_names, 1):
                    sheet = workbook[sheet_name]
                    print(f"\n   {i}. Sheet: '{sheet_name}'")
                    
                    # Get headers from first row
                    headers = []
                    for col in range(1, 20):  # Check first 20 columns
                        cell_value = sheet.cell(row=1, column=col).value
                        if cell_value:
                            headers.append(str(cell_value))
                        else:
                            break
                    
                    print(f"      Headers: {headers}")
                    
                    # Count data rows
                    data_rows = 0
                    for row in range(2, sheet.max_row + 1):
                        if sheet.cell(row=row, column=1).value:
                            data_rows += 1
                    
                    print(f"      Data rows: {data_rows}")
                    
                    # Special check for Daily Records sheet
                    if sheet_name == 'Daily Records':
                        print(f"      ✅ Found 'Daily Records' sheet!")
                        expected_columns = [
                            'Staff ID', 'Full Name', 'Department', 'Position',
                            'Time In', 'Time Out', 'Status', 'Late (min)',
                            'Early Dep (min)', 'Work Hrs', 'OT Hrs'
                        ]
                        
                        missing_columns = []
                        for expected in expected_columns:
                            if expected not in headers:
                                missing_columns.append(expected)
                        
                        if missing_columns:
                            print(f"      ❌ Missing columns: {missing_columns}")
                        else:
                            print(f"      ✅ All required columns present!")
                        
                        # Show sample data
                        if data_rows > 0:
                            print(f"      📊 Sample data (first 3 rows):")
                            for row in range(2, min(5, 2 + data_rows)):
                                staff_id = sheet.cell(row=row, column=1).value
                                name = sheet.cell(row=row, column=2).value
                                status = sheet.cell(row=row, column=7).value
                                print(f"         {staff_id} | {name} | {status}")
                
                print(f"\n✅ Excel file saved as: {filename}")
                return True
                
            else:
                print(f"❌ Failed to generate report")
                print(f"   Response: {response}")
                return False
                
        except Exception as e:
            print(f"❌ Error during test: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

if __name__ == "__main__":
    success = test_daily_attendance_export()
    
    if success:
        print(f"\n🎯 TEST RESULTS:")
        print(f"   The Daily Attendance Report function is working")
        print(f"   If you're not seeing individual staff data, check:")
        print(f"   1. Make sure to open the 'Daily Records' sheet in Excel")
        print(f"   2. Verify there is attendance data for the selected date")
        print(f"   3. Check if staff records exist in the database")
    else:
        print(f"\n❌ TEST FAILED")
        print(f"   The function needs to be debugged")